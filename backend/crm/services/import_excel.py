import openpyxl
from datetime import datetime, timedelta
from django.db import transaction
from crm.models import Lead, Team
from django.contrib.auth import get_user_model

import re

EXCEL_KEYS = {
    "id": "id",
    "fuente": "fuente",
    "fecha": "fecha",
    "activo/s": "activo",
    "activo": "activo",
    "tipo activo": "tipo_activo",
    "precio": "precio",
    "dorm.": "dorm",
    "dorm": "dorm",
    "interés": "interes",
    "interes": "interes",
    "nombre": "nombre",
    "telefono": "telefono",
    "teléfono": "telefono",
    "mail": "email",
    "email": "email",
    "agente": "agente",
    "seatable": "seatable",
    "inmovilla": "inmovilla",
    "contestada": "contestada",
    "contestado": "contestada",
    "visita": "visita",
}

def norm_header(value):
    return str(value or "").strip().lower()

def parse_fecha(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=int(value))).date()
    try:
        if "/" in str(value):
            return datetime.strptime(str(value), "%d/%m/%Y").date()
        if "-" in str(value):
            return datetime.strptime(str(value), "%Y-%m-%d").date()
    except Exception:
        pass
    return None

def parse_bool(value):
    if isinstance(value, bool): return value
    if not value: return False
    return str(value).strip().lower() in ("1", "si", "yes", "true", "x")

def get_agente(queryset, value):
    User = get_user_model()
    return queryset.filter(username=value).first() or queryset.filter(email=value).first()

def normalize_phone(p):
    """Deja solo dígitos para comparar teléfonos."""
    return re.sub(r"\D", "", str(p or ""))


def preview_import(file, team_id):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    # Cabeceras normalizadas
    headers = [norm_header(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    keys = [EXCEL_KEYS.get(h, h) for h in headers]

    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    # Datos existentes del team
    existing_email = set(
        e.lower() for e in
        Lead.objects
        .filter(team_id=team_id)
        .exclude(email="")
        .values_list("email", flat=True)
    )

    existing_tel = set(
        normalize_phone(t) for t in
        Lead.objects
        .filter(team_id=team_id)
        .exclude(telefono="")
        .values_list("telefono", flat=True)
    )

    nuevos = 0
    actualizados = 0
    invalidos = 0
    warnings = []

    for idx, row in enumerate(rows, start=2):
        data = dict(zip(keys, row))

        tel = normalize_phone(data.get("telefono") or data.get("teléfono"))
        email = (data.get("email") or data.get("mail") or "").strip().lower()

        # --- Validación mínima ---
        if not email and not tel:
            invalidos += 1
            warnings.append(f"Fila {idx}: ignorada, requiere email o teléfono.")
            continue

        # --- Coincidencias existentes ---
        if email and email in existing_email:
            actualizados += 1
            continue

        if tel and tel in existing_tel:
            actualizados += 1
            continue

        # --- Nuevo lead ---
        nuevos += 1

    return {
        "nuevos": nuevos,
        "actualizados": actualizados,
        "invalidos": invalidos,
        "warnings": warnings,
    }
    
preview_import_excel = preview_import

@transaction.atomic
def commit_import(file, team_id):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    User = get_user_model()
    keys = [EXCEL_KEYS.get(h, h.lower()) for h in headers]
    new, updated, dups, warn = 0, 0, 0, []

    for row in rows:
        data = dict(zip(keys, row))
        tel = str(data.get("telefono") or "").strip()
        email = (data.get("email") or "").strip().lower()
        defaults = {
            "fuente": data.get("fuente"),
            "fecha": parse_fecha(data.get("fecha")),
            "activo": data.get("activo"),
            "tipo_activo": data.get("tipo_activo"),
            "precio": data.get("precio") or 0,
            "dorm": data.get("dorm"),
            "interes": data.get("interes"),
            "nombre": data.get("nombre"),
            "seatable": parse_bool(data.get("seatable")),
            "inmovilla": parse_bool(data.get("inmovilla")),
            "contestada": parse_bool(data.get("contestada")),
            "contestada_at": parse_fecha(data.get("contestada")),
            "seatable_at": parse_fecha(data.get("seatable")),
            "inmovilla_at": parse_fecha(data.get("inmovilla")),
            "visita": parse_bool(data.get("visita")),
            "visita_at": parse_fecha(data.get("visita")),
            "estado": "nuevo",
        }
        # agente assignment
        agent_name = data.get("agente")
        defaults["agente"] = get_agente(User.objects.all(), agent_name) if agent_name else None

        # deduplicate
        obj = None
        if email:
            obj = Lead.objects.filter(team_id=team_id, email=email).first()
        if not obj and tel:
            obj = Lead.objects.filter(team_id=team_id, telefono=tel).first()
        if obj:
            for k, v in defaults.items():
                setattr(obj, k, v)
            obj.save()
            updated += 1
        elif (email or tel):
            Lead.objects.create(team_id=team_id, email=email, telefono=tel, **defaults)
            new += 1
        else:
            warn.append(f"Fila sin email/telefono: {data}")
            continue
    return {"nuevos": new, "actualizados": updated, "warnings": warn}
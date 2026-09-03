from pathlib import Path
from decimal import Decimal, InvalidOperation
from datetime import datetime, date

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.gestion.models import (
    EmpresaGestionLegacy,
    Proveedor,
    FacturaProveedorGestion,
    AlbaranProveedorGestion,
    FacturaProveedorLineaGestion,
    AlbaranProveedorLineaGestion,
)


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except Exception:
        return None


def clean_decimal(value, places="0.00"):
    if value in (None, ""):
        return Decimal(places)
    try:
        return Decimal(str(value)).quantize(Decimal(places))
    except (InvalidOperation, ValueError):
        return Decimal(places)


def clean_bool(value):
    if value in (True, 1, "1", "true", "True", "Sí", "SI", "si", "SÍ"):
        return True
    return False


def clean_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def read_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [clean(c.value) for c in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or clean(v) == "" for v in row):
            continue
        rows.append(dict(zip(headers, row)))

    return headers, rows


def require_columns(headers, required, label):
    missing = [h for h in required if h not in headers]
    if missing:
        raise CommandError(f"{label}: faltan columnas {missing}")


def raw_data(data):
    return {k: clean(v) for k, v in data.items()}


class Command(BaseCommand):
    help = "Sync incremental Access/XLSX para Gestión. Dry-run por defecto. Commit insert-only fase 1."

    def add_arguments(self, parser):
        parser.add_argument("--source", required=True, help="Directorio origen access_sync_YYYYMMDD_HHMM")
        parser.add_argument("--commit", action="store_true", help="Aplica inserciones nuevas. Nunca actualiza existentes.")
        parser.add_argument("--samples", type=int, default=8)

    def handle(self, *args, **options):
        source = Path(options["source"])
        commit = options["commit"]

        if not source.exists() or not source.is_dir():
            raise CommandError(f"No existe el directorio source: {source}")

        self.samples = options["samples"]

        self.stdout.write("=== sync_access_incremental ===")
        self.stdout.write(f"Source: {source}")
        self.stdout.write(f"Modo: {'COMMIT INSERT-ONLY' if commit else 'DRY-RUN'}")
        self.stdout.write("Regla: no actualiza, no borra, no sobrescribe existentes.")
        self.stdout.write("")

        files = {
            "proveedores": source / "tblProveedores.xlsx",
            "facturas": source / "tblFacturas.xlsx",
            "facturas_lineas": source / "tblFacturasLineas.xlsx",
            "albaranes": source / "tblAlbaranes.xlsx",
            "albaranes_lineas": source / "tblAlbaranesLineas.xlsx",
        }

        for label, path in files.items():
            if not path.exists():
                raise CommandError(f"Falta fichero requerido {label}: {path}")

        empresas = {
            e.legacy_id_empresa: e
            for e in EmpresaGestionLegacy.objects.select_related("team").all()
        }

        if not empresas:
            raise CommandError("No hay EmpresaGestionLegacy.")

        sin_asignar = (
            empresas.get(0)
            or EmpresaGestionLegacy.objects.select_related("team")
            .filter(nombre_empresa__icontains="SIN ASIGNAR")
            .first()
        )

        if not sin_asignar:
            raise CommandError("No existe EmpresaGestionLegacy SIN ASIGNAR LEGACY / legacy_id_empresa=0.")

        unique_teams = []
        seen_team_ids = set()
        for empresa in EmpresaGestionLegacy.objects.select_related("team").order_by("legacy_id_empresa"):
            if empresa.team_id and empresa.team_id not in seen_team_ids:
                unique_teams.append(empresa.team)
                seen_team_ids.add(empresa.team_id)

        self.stdout.write(f"Empresas legacy: {len(empresas)}")
        self.stdout.write(f"Teams destino proveedores: {len(unique_teams)}")
        self.stdout.write(f"SIN ASIGNAR LEGACY: empresa_id={sin_asignar.id}, team_id={sin_asignar.team_id}")
        self.stdout.write("")

        with transaction.atomic():
            resumen = {}

            resumen["proveedores"] = self.sync_proveedores(files["proveedores"], unique_teams, commit)
            resumen["facturas"] = self.sync_facturas(files["facturas"], empresas, sin_asignar, commit)
            resumen["albaranes"] = self.sync_albaranes(files["albaranes"], empresas, sin_asignar, commit)
            resumen["facturas_lineas"] = self.sync_facturas_lineas(files["facturas_lineas"], commit)
            resumen["albaranes_lineas"] = self.sync_albaranes_lineas(files["albaranes_lineas"], commit)

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write("=== RESUMEN FINAL ===")
        for key, data in resumen.items():
            self.stdout.write(f"{key}: {data}")

        self.stdout.write("")
        self.stdout.write("OK: proceso terminado.")

    def sync_proveedores(self, path, teams, commit):
        required = [
            "IdProveedor", "NombreComercial", "NombreFiscal", "Direccion", "CodPostal",
            "Poblacion", "Provincia", "Pais", "CIF", "Email", "Telefono",
            "ContactoComercial", "TelContactoComercial", "ContactoAdmin",
            "TelContactoAdmin", "SP_Iva", "Observaciones", "SubContrata",
            "CodObra", "FueraListado",
        ]

        headers, rows = read_xlsx(path)
        require_columns(headers, required, "tblProveedores.xlsx")

        created = 0
        existing = 0
        skipped = 0
        preview = []

        for row_num, data in enumerate(rows, start=2):
            legacy_id = clean_int(data.get("IdProveedor"))
            nombre = clean(data.get("NombreComercial"))

            if not legacy_id or not nombre:
                skipped += len(teams)
                continue

            payload = {
                "nombre_comercial": nombre,
                "nombre_fiscal": clean(data.get("NombreFiscal")),
                "direccion": clean(data.get("Direccion")),
                "cod_postal": clean(data.get("CodPostal")),
                "poblacion": clean(data.get("Poblacion")),
                "provincia": clean(data.get("Provincia")),
                "pais": clean(data.get("Pais")),
                "cif": clean(data.get("CIF")),
                "email": clean(data.get("Email")),
                "telefono": clean(data.get("Telefono")),
                "contacto_comercial": clean(data.get("ContactoComercial")),
                "tel_contacto_comercial": clean(data.get("TelContactoComercial")),
                "contacto_admin": clean(data.get("ContactoAdmin")),
                "tel_contacto_admin": clean(data.get("TelContactoAdmin")),
                "sp_iva": clean_bool(data.get("SP_Iva")),
                "observaciones": clean(data.get("Observaciones")),
                "es_subcontrata": clean_bool(data.get("SubContrata")),
                "cod_obra_legacy": clean(data.get("CodObra")),
                "fuera_listado": clean_bool(data.get("FueraListado")),
                "activo": not clean_bool(data.get("FueraListado")),
                "raw_data": raw_data(data),
            }

            for team in teams:
                exists = Proveedor.objects.filter(
                    team=team,
                    legacy_id_proveedor=legacy_id,
                ).exists()

                if exists:
                    existing += 1
                    continue

                if commit:
                    Proveedor.objects.create(
                        team=team,
                        legacy_id_proveedor=legacy_id,
                        **payload,
                    )

                created += 1

                if len(preview) < self.samples:
                    preview.append({
                        "team_id": team.id,
                        "legacy_id": legacy_id,
                        "nombre": nombre,
                    })

        self.stdout.write("=== PROVEEDORES ===")
        self.stdout.write(f"Fichero: {path}")
        self.stdout.write(f"Candidatos base: {len(rows)}")
        self.stdout.write(f"Destinos team: {len(teams)}")
        self.stdout.write(f"Existentes omitidos: {existing}")
        self.stdout.write(f"Nuevos {'creados' if commit else 'candidatos'}: {created}")
        self.stdout.write(f"Omitidos sin clave/nombre: {skipped}")
        self.stdout.write(f"Muestra: {preview}")
        self.stdout.write("")

        return {
            "base": len(rows),
            "destinos_team": len(teams),
            "existentes": existing,
            "nuevos": created,
            "omitidos": skipped,
        }

    def sync_facturas(self, path, empresas, sin_asignar, commit):
        required = [
            "CodFactura", "CodObra", "CodProveedor", "NumFacturaProveedor",
            "FechaEmisionFactura", "ImporteBaseImponible", "ImporteIva",
            "ImporteFactura", "FechaAutorizacionGerencia", "FechaPagoSegunContrato",
            "FormasPago", "FechaRealPago", "Estado", "Observaciones", "Asignada",
            "Retencion", "TieneRetencion", "GeneradoAlbaran", "Archivo", "Archivo1",
            "Certificada", "Empresa", "ImportePagado",
        ]

        headers, rows = read_xlsx(path)
        require_columns(headers, required, "tblFacturas.xlsx")

        created = 0
        existing = 0
        skipped_no_cod = 0
        mapped_sin_asignar = 0
        empresa_unknown = 0
        proveedor_missing = 0
        preview = []

        for row_num, data in enumerate(rows, start=2):
            cod_factura = clean(data.get("CodFactura"))
            empresa_raw = clean_int(data.get("Empresa"))
            cod_proveedor = clean_int(data.get("CodProveedor"))

            if not cod_factura:
                skipped_no_cod += 1
                continue

            if empresa_raw in (None, 0):
                empresa = sin_asignar
                mapped_sin_asignar += 1
            else:
                empresa = empresas.get(empresa_raw)

            if not empresa:
                empresa_unknown += 1
                continue

            exists = FacturaProveedorGestion.objects.filter(
                team=empresa.team,
                cod_factura=cod_factura,
            ).exists()

            if exists:
                existing += 1
                continue

            proveedor = None
            if cod_proveedor is not None:
                proveedor = Proveedor.objects.filter(
                    team=empresa.team,
                    legacy_id_proveedor=cod_proveedor,
                ).first()
                if not proveedor:
                    proveedor_missing += 1

            payload = {
                "team": empresa.team,
                "empresa_legacy": empresa,
                "proveedor": proveedor,
                "cod_factura": cod_factura,
                "cod_obra_legacy": clean(data.get("CodObra")),
                "cod_proveedor_legacy": cod_proveedor,
                "empresa_legacy_raw": empresa_raw,
                "num_factura_proveedor": clean(data.get("NumFacturaProveedor")),
                "fecha_emision": clean_date(data.get("FechaEmisionFactura")),
                "fecha_autorizacion_gerencia": clean_date(data.get("FechaAutorizacionGerencia")),
                "fecha_pago_segun_contrato": clean_date(data.get("FechaPagoSegunContrato")),
                "fecha_real_pago": clean_date(data.get("FechaRealPago")),
                "importe_base_imponible": clean_decimal(data.get("ImporteBaseImponible")),
                "importe_iva": clean_decimal(data.get("ImporteIva")),
                "importe_factura": clean_decimal(data.get("ImporteFactura")),
                "retencion": clean_decimal(data.get("Retencion")),
                "importe_pagado": clean_decimal(data.get("ImportePagado")),
                "forma_pago": clean(data.get("FormasPago")),
                "estado": clean(data.get("Estado")),
                "observaciones": clean(data.get("Observaciones")),
                "asignada": clean_bool(data.get("Asignada")),
                "tiene_retencion": clean_bool(data.get("TieneRetencion")),
                "generado_albaran": clean_bool(data.get("GeneradoAlbaran")),
                "certificada": clean_bool(data.get("Certificada")),
                "archivo": clean(data.get("Archivo")),
                "archivo1": clean(data.get("Archivo1")),
                "raw_data": raw_data(data),
            }

            if commit:
                FacturaProveedorGestion.objects.create(**payload)

            created += 1

            if len(preview) < self.samples:
                preview.append({
                    "team_id": empresa.team_id,
                    "empresa_raw": empresa_raw,
                    "cod_factura": cod_factura,
                    "importe": str(payload["importe_factura"]),
                    "proveedor_link": bool(proveedor),
                })

        self.stdout.write("=== FACTURAS CABECERA ===")
        self.stdout.write(f"Fichero: {path}")
        self.stdout.write(f"Candidatos: {len(rows)}")
        self.stdout.write(f"Existentes omitidas: {existing}")
        self.stdout.write(f"Nuevas {'creadas' if commit else 'candidatas'}: {created}")
        self.stdout.write(f"Mapeadas a SIN ASIGNAR LEGACY: {mapped_sin_asignar}")
        self.stdout.write(f"Empresa desconocida omitidas: {empresa_unknown}")
        self.stdout.write(f"Sin CodFactura omitidas: {skipped_no_cod}")
        self.stdout.write(f"Proveedor no encontrado: {proveedor_missing}")
        self.stdout.write(f"Muestra: {preview}")
        self.stdout.write("")

        return {
            "candidatos": len(rows),
            "existentes": existing,
            "nuevos": created,
            "sin_asignar": mapped_sin_asignar,
            "empresa_desconocida": empresa_unknown,
            "sin_codigo": skipped_no_cod,
            "proveedor_missing": proveedor_missing,
        }

    def sync_albaranes(self, path, empresas, sin_asignar, commit):
        required = [
            "CodObra", "CodAlbaran", "Presupuesto", "CodPresupuesto",
            "NumAlbaranProveedor", "FechaAlbaran", "ImporteAlbaran",
            "Descripcion", "FechaEntregaMercaderia", "Recepcionadopor",
            "OK_Presupuesto", "CodProveedor", "AutorizadoJefeObra", "Archivo",
            "AsignadoPartidaObra", "Empresa", "Situacion", "AsignadoFactura",
            "ImporteAsignadoFactura", "LineasAsignadas",
        ]

        headers, rows = read_xlsx(path)
        require_columns(headers, required, "tblAlbaranes.xlsx")

        created = 0
        existing = 0
        skipped_no_cod = 0
        mapped_sin_asignar = 0
        empresa_unknown = 0
        proveedor_missing = 0
        preview = []

        for row_num, data in enumerate(rows, start=2):
            cod_albaran = clean(data.get("CodAlbaran"))
            empresa_raw = clean_int(data.get("Empresa"))
            cod_proveedor = clean_int(data.get("CodProveedor"))

            if not cod_albaran:
                skipped_no_cod += 1
                continue

            if empresa_raw in (None, 0):
                empresa = sin_asignar
                mapped_sin_asignar += 1
            else:
                empresa = empresas.get(empresa_raw)

            if not empresa:
                empresa_unknown += 1
                continue

            exists = AlbaranProveedorGestion.objects.filter(
                team=empresa.team,
                cod_albaran=cod_albaran,
            ).exists()

            if exists:
                existing += 1
                continue

            proveedor = None
            if cod_proveedor is not None:
                proveedor = Proveedor.objects.filter(
                    team=empresa.team,
                    legacy_id_proveedor=cod_proveedor,
                ).first()
                if not proveedor:
                    proveedor_missing += 1

            payload = {
                "team": empresa.team,
                "empresa_legacy": empresa,
                "proveedor": proveedor,
                "cod_albaran": cod_albaran,
                "cod_obra_legacy": clean(data.get("CodObra")),
                "cod_proveedor_legacy": cod_proveedor,
                "empresa_legacy_raw": empresa_raw,
                "num_albaran_proveedor": clean(data.get("NumAlbaranProveedor")),
                "fecha_albaran": clean_date(data.get("FechaAlbaran")),
                "fecha_entrega_mercaderia": clean_date(data.get("FechaEntregaMercaderia")),
                "importe_albaran": clean_decimal(data.get("ImporteAlbaran")),
                "importe_asignado_factura": clean_decimal(data.get("ImporteAsignadoFactura")),
                "descripcion": clean(data.get("Descripcion")),
                "recepcionado_por": clean(data.get("Recepcionadopor")),
                "presupuesto": clean_bool(data.get("Presupuesto")),
                "cod_presupuesto_legacy": clean(data.get("CodPresupuesto")),
                "ok_presupuesto": clean_bool(data.get("OK_Presupuesto")),
                "autorizado_jefe_obra": clean_bool(data.get("AutorizadoJefeObra")),
                "asignado_partida_obra": clean_bool(data.get("AsignadoPartidaObra")),
                "asignado_factura": clean_bool(data.get("AsignadoFactura")),
                "lineas_asignadas": clean_int(data.get("LineasAsignadas")) or 0,
                "situacion": clean(data.get("Situacion")),
                "archivo": clean(data.get("Archivo")),
                "raw_data": raw_data(data),
            }

            if commit:
                AlbaranProveedorGestion.objects.create(**payload)

            created += 1

            if len(preview) < self.samples:
                preview.append({
                    "team_id": empresa.team_id,
                    "empresa_raw": empresa_raw,
                    "cod_albaran": cod_albaran,
                    "importe": str(payload["importe_albaran"]),
                    "proveedor_link": bool(proveedor),
                })

        self.stdout.write("=== ALBARANES CABECERA ===")
        self.stdout.write(f"Fichero: {path}")
        self.stdout.write(f"Candidatos: {len(rows)}")
        self.stdout.write(f"Existentes omitidos: {existing}")
        self.stdout.write(f"Nuevos {'creados' if commit else 'candidatos'}: {created}")
        self.stdout.write(f"Mapeados a SIN ASIGNAR LEGACY: {mapped_sin_asignar}")
        self.stdout.write(f"Empresa desconocida omitidos: {empresa_unknown}")
        self.stdout.write(f"Sin CodAlbaran omitidos: {skipped_no_cod}")
        self.stdout.write(f"Proveedor no encontrado: {proveedor_missing}")
        self.stdout.write(f"Muestra: {preview}")
        self.stdout.write("")

        return {
            "candidatos": len(rows),
            "existentes": existing,
            "nuevos": created,
            "sin_asignar": mapped_sin_asignar,
            "empresa_desconocida": empresa_unknown,
            "sin_codigo": skipped_no_cod,
            "proveedor_missing": proveedor_missing,
        }


    def sync_facturas_lineas(self, path, commit):
        required = [
            "CodFactura", "Linea", "CodArticulo", "CodAlbaran", "LineaAlbaran",
            "Cantidad", "PrecioUnitario", "ImporteLinea", "ImporteDescuento",
            "Descuento", "EnPartida", "CantidadEnPartidas", "EnAlmacen",
        ]

        headers, rows = read_xlsx(path)
        require_columns(headers, required, "tblFacturasLineas.xlsx")

        facturas = {
            f.cod_factura: f
            for f in FacturaProveedorGestion.objects.select_related("team").all()
        }

        albaranes = {
            (a.team_id, a.cod_albaran): a
            for a in AlbaranProveedorGestion.objects.all()
        }

        created = 0
        existing = 0
        sin_cabecera = 0
        sin_clave = 0
        duplicados_source = 0
        albaran_missing = 0
        preview = []
        seen = set()

        for data in rows:
            cod_factura = clean(data.get("CodFactura"))
            linea = clean_int(data.get("Linea"))

            if not cod_factura or linea is None:
                sin_clave += 1
                continue

            key = (cod_factura, linea)
            if key in seen:
                duplicados_source += 1
                continue
            seen.add(key)

            factura = facturas.get(cod_factura)
            if not factura:
                sin_cabecera += 1
                continue

            exists = FacturaProveedorLineaGestion.objects.filter(
                factura=factura,
                linea=linea,
            ).exists()

            if exists:
                existing += 1
                continue

            cod_albaran = clean(data.get("CodAlbaran"))
            albaran = None

            if cod_albaran:
                albaran = albaranes.get((factura.team_id, cod_albaran))
                if not albaran:
                    albaran_missing += 1

            payload = {
                "factura": factura,
                "albaran": albaran,
                "linea": linea,
                "cod_articulo_legacy": clean_int(data.get("CodArticulo")),
                "cod_albaran_legacy": cod_albaran,
                "linea_albaran_legacy": clean_int(data.get("LineaAlbaran")),
                "cantidad": clean_decimal(data.get("Cantidad"), "0.0000"),
                "precio_unitario": clean_decimal(data.get("PrecioUnitario"), "0.0000"),
                "importe_linea": clean_decimal(data.get("ImporteLinea"), "0.00"),
                "importe_descuento": clean_decimal(data.get("ImporteDescuento"), "0.00"),
                "descuento": clean_decimal(data.get("Descuento"), "0.0000"),
                "en_partida": clean_bool(data.get("EnPartida")),
                "cantidad_en_partidas": clean_decimal(data.get("CantidadEnPartidas"), "0.0000"),
                "en_almacen": clean_bool(data.get("EnAlmacen")),
                "raw_data": raw_data(data),
            }

            if commit:
                FacturaProveedorLineaGestion.objects.create(**payload)

            created += 1

            if len(preview) < self.samples:
                preview.append({
                    "team_id": factura.team_id,
                    "cod_factura": cod_factura,
                    "linea": linea,
                    "cod_albaran": cod_albaran,
                    "albaran_link": bool(albaran),
                    "importe": str(payload["importe_linea"]),
                })

        self.stdout.write("=== FACTURAS LINEAS ===")
        self.stdout.write(f"Fichero: {path}")
        self.stdout.write(f"Candidatos: {len(rows)}")
        self.stdout.write(f"Existentes omitidas: {existing}")
        self.stdout.write(f"Nuevas {'creadas' if commit else 'candidatas'}: {created}")
        self.stdout.write(f"Omitidas sin cabecera: {sin_cabecera}")
        self.stdout.write(f"Sin clave: {sin_clave}")
        self.stdout.write(f"Duplicados source: {duplicados_source}")
        self.stdout.write(f"CodAlbaran sin cabecera importada mismo Team: {albaran_missing}")
        self.stdout.write(f"Muestra: {preview}")
        self.stdout.write("")

        return {
            "candidatos": len(rows),
            "existentes": existing,
            "nuevos": created,
            "omitidas_sin_cabecera": sin_cabecera,
            "sin_clave": sin_clave,
            "duplicados_source": duplicados_source,
            "albaran_missing": albaran_missing,
        }


    def sync_albaranes_lineas(self, path, commit):
        required = [
            "CodAlbaran", "Linea", "CodArticulo", "Cantidad", "Unidad",
            "CantidadCompra", "UnidadCompra", "CantidadxUnidad", "PrecioUnitario",
            "ImporteLinea", "Facturado", "Factura", "EnPedido", "EnPartida",
            "FechaEntrega", "RecepcionadoPor", "ImporteDescuento", "Descuento",
            "IdAlmacen", "Observaciones", "TipoRecurso", "CantidadEnPartidas",
            "EnAlmacen",
        ]

        headers, rows = read_xlsx(path)
        require_columns(headers, required, "tblAlbaranesLineas.xlsx")

        albaranes = {
            a.cod_albaran: a
            for a in AlbaranProveedorGestion.objects.select_related("team").all()
        }

        created = 0
        existing = 0
        sin_cabecera = 0
        sin_clave = 0
        duplicados_source = 0
        preview = []
        seen = set()

        for data in rows:
            cod_albaran = clean(data.get("CodAlbaran"))
            linea = clean_int(data.get("Linea"))

            if not cod_albaran or linea is None:
                sin_clave += 1
                continue

            key = (cod_albaran, linea)
            if key in seen:
                duplicados_source += 1
                continue
            seen.add(key)

            albaran = albaranes.get(cod_albaran)
            if not albaran:
                sin_cabecera += 1
                continue

            exists = AlbaranProveedorLineaGestion.objects.filter(
                albaran=albaran,
                linea=linea,
            ).exists()

            if exists:
                existing += 1
                continue

            payload = {
                "albaran": albaran,
                "linea": linea,
                "cod_articulo_legacy": clean_int(data.get("CodArticulo")),
                "cantidad": clean_decimal(data.get("Cantidad"), "0.0000"),
                "unidad": clean(data.get("Unidad")),
                "cantidad_compra": clean_decimal(data.get("CantidadCompra"), "0.0000"),
                "unidad_compra": clean(data.get("UnidadCompra")),
                "cantidad_x_unidad": clean_decimal(data.get("CantidadxUnidad"), "0.0000"),
                "precio_unitario": clean_decimal(data.get("PrecioUnitario"), "0.0000"),
                "importe_linea": clean_decimal(data.get("ImporteLinea"), "0.00"),
                "facturado": clean_bool(data.get("Facturado")),
                "factura_legacy": clean(data.get("Factura")),
                "en_pedido": clean_bool(data.get("EnPedido")),
                "en_partida": clean_bool(data.get("EnPartida")),
                "fecha_entrega": clean_date(data.get("FechaEntrega")),
                "recepcionado_por": clean(data.get("RecepcionadoPor")),
                "importe_descuento": clean_decimal(data.get("ImporteDescuento"), "0.00"),
                "descuento": clean_decimal(data.get("Descuento"), "0.0000"),
                "id_almacen_legacy": clean_int(data.get("IdAlmacen")),
                "observaciones": clean(data.get("Observaciones")),
                "tipo_recurso": clean(data.get("TipoRecurso")),
                "cantidad_en_partidas": clean_decimal(data.get("CantidadEnPartidas"), "0.0000"),
                "en_almacen": clean_bool(data.get("EnAlmacen")),
                "raw_data": raw_data(data),
            }

            if commit:
                AlbaranProveedorLineaGestion.objects.create(**payload)

            created += 1

            if len(preview) < self.samples:
                preview.append({
                    "team_id": albaran.team_id,
                    "cod_albaran": cod_albaran,
                    "linea": linea,
                    "cod_articulo": payload["cod_articulo_legacy"],
                    "importe": str(payload["importe_linea"]),
                    "facturado": payload["facturado"],
                })

        self.stdout.write("=== ALBARANES LINEAS ===")
        self.stdout.write(f"Fichero: {path}")
        self.stdout.write(f"Candidatos: {len(rows)}")
        self.stdout.write(f"Existentes omitidos: {existing}")
        self.stdout.write(f"Nuevos {'creados' if commit else 'candidatos'}: {created}")
        self.stdout.write(f"Omitidos sin cabecera: {sin_cabecera}")
        self.stdout.write(f"Sin clave: {sin_clave}")
        self.stdout.write(f"Duplicados source: {duplicados_source}")
        self.stdout.write(f"Muestra: {preview}")
        self.stdout.write("")

        return {
            "candidatos": len(rows),
            "existentes": existing,
            "nuevos": created,
            "omitidos_sin_cabecera": sin_cabecera,
            "sin_clave": sin_clave,
            "duplicados_source": duplicados_source,
        }


from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.gestion.models import EmpresaGestionLegacy, Proveedor


DEFAULT_FILE = "/app/imports/access_sync_2026-05-19/tblProveedores.xlsx"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_bool(value):
    if value in (True, 1, "1", "true", "True", "Sí", "SI", "si", "SÍ"):
        return True
    return False


class Command(BaseCommand):
    help = "Importa proveedores legacy en todos los Teams mapeados desde EmpresaGestionLegacy"

    def add_arguments(self, parser):
        parser.add_argument("--file", default=DEFAULT_FILE)
        parser.add_argument("--commit", action="store_true")

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        commit = options["commit"]

        if not file_path.exists():
            raise CommandError(f"No existe el fichero: {file_path}")

        empresas = list(
            EmpresaGestionLegacy.objects.select_related("team").order_by("legacy_id_empresa")
        )

        if not empresas:
            raise CommandError("No hay EmpresaGestionLegacy importadas. Importa primero tblParametros.")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [clean(c.value) for c in ws[1]]

        required = [
            "IdProveedor",
            "NombreComercial",
            "NombreFiscal",
            "Direccion",
            "CodPostal",
            "Poblacion",
            "Provincia",
            "Pais",
            "CIF",
            "Email",
            "Telefono",
            "ContactoComercial",
            "TelContactoComercial",
            "ContactoAdmin",
            "TelContactoAdmin",
            "SP_Iva",
            "Observaciones",
            "SubContrata",
            "CodObra",
            "FueraListado",
        ]

        missing = [h for h in required if h not in headers]
        if missing:
            raise CommandError(f"Faltan columnas: {missing}")

        created = 0
        updated = 0
        skipped = 0
        errors = []
        preview = []

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                data = dict(zip(headers, row))

                legacy_id = data.get("IdProveedor")
                nombre_comercial = clean(data.get("NombreComercial"))

                if not legacy_id or not nombre_comercial:
                    skipped += len(empresas)
                    errors.append(f"Fila {row_num}: sin IdProveedor o NombreComercial")
                    continue

                payload_base = {
                    "nombre_comercial": nombre_comercial,
                    "nombre_fiscal": clean(data.get("NombreFiscal")),
                    "direccion": clean(data.get("Direccion")),
                    "cod_postal": clean(data.get("CodPostal")),
                    "poblacion": clean(data.get("Poblacion")),
                    "provincia": clean(data.get("Provincia")),
                    "pais": clean(data.get("Pais")),
                    "cif": clean(data.get("CIF")),
                    "email": clean(data.get("Email")),
                    "telefono": clean(data.get("Telefono")),
                    "contacto_comercial": clean(data.get("ContactoComercial")),
                    "tel_contacto_comercial": clean(data.get("TelContactoComercial")),
                    "contacto_admin": clean(data.get("ContactoAdmin")),
                    "tel_contacto_admin": clean(data.get("TelContactoAdmin")),
                    "sp_iva": clean_bool(data.get("SP_Iva")),
                    "observaciones": clean(data.get("Observaciones")),
                    "es_subcontrata": clean_bool(data.get("SubContrata")),
                    "cod_obra_legacy": clean(data.get("CodObra")),
                    "fuera_listado": clean_bool(data.get("FueraListado")),
                    "activo": not clean_bool(data.get("FueraListado")),
                    "raw_data": {k: clean(v) for k, v in data.items()},
                }

                for empresa in empresas:
                    obj, was_created = Proveedor.objects.update_or_create(
                        team=empresa.team,
                        legacy_id_proveedor=int(legacy_id),
                        defaults=payload_base,
                    )

                    if was_created:
                        created += 1
                    else:
                        updated += 1

                    if len(preview) < 8:
                        preview.append({
                            "team": str(empresa.team),
                            "legacy_id": int(legacy_id),
                            "nombre": nombre_comercial,
                            "cif": payload_base["cif"],
                        })

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write("=== IMPORT PROVEEDORES LEGACY ===")
        self.stdout.write(f"Fichero: {file_path}")
        self.stdout.write(f"Modo: {'COMMIT REAL' if commit else 'DRY-RUN'}")
        self.stdout.write(f"Empresas/Teams destino: {len(empresas)}")
        self.stdout.write(f"Creados: {created}")
        self.stdout.write(f"Actualizados: {updated}")
        self.stdout.write(f"Omitidos: {skipped}")

        self.stdout.write("")
        self.stdout.write("=== MUESTRA ===")
        for item in preview:
            self.stdout.write(str(item))

        if errors:
            self.stdout.write("")
            self.stdout.write("=== ERRORES / AVISOS ===")
            for e in errors[:20]:
                self.stdout.write(e)

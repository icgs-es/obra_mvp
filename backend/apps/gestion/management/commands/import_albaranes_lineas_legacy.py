from pathlib import Path
from decimal import Decimal, InvalidOperation
from datetime import datetime, date

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.gestion.models import (
    AlbaranProveedorGestion,
    AlbaranProveedorLineaGestion,
)


DEFAULT_FILE = "/app/imports/access_sync_2026-05-19/tblAlbaranesLineas.xlsx"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except Exception:
        return None


def clean_decimal(value, places="0.0000"):
    if value in (None, ""):
        return Decimal(places)
    try:
        return Decimal(str(value)).quantize(Decimal(places))
    except (InvalidOperation, ValueError):
        return Decimal(places)


def clean_bool(value):
    if value in (True, 1, "1", "true", "True", "Sí", "SI", "si", "SÍ"):
        return True
    return False


def clean_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


class Command(BaseCommand):
    help = "Importa líneas de albaranes legacy desde tblAlbaranesLineas.xlsx"

    def add_arguments(self, parser):
        parser.add_argument("--file", default=DEFAULT_FILE)
        parser.add_argument("--commit", action="store_true")

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        commit = options["commit"]

        if not file_path.exists():
            raise CommandError(f"No existe el fichero: {file_path}")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [clean(c.value) for c in ws[1]]

        required = [
            "CodAlbaran",
            "Linea",
            "CodArticulo",
            "Cantidad",
            "Unidad",
            "CantidadCompra",
            "UnidadCompra",
            "CantidadxUnidad",
            "PrecioUnitario",
            "ImporteLinea",
            "Facturado",
            "Factura",
            "EnPedido",
            "EnPartida",
            "FechaEntrega",
            "RecepcionadoPor",
            "ImporteDescuento",
            "Descuento",
            "IdAlmacen",
            "Observaciones",
            "TipoRecurso",
            "CantidadEnPartidas",
            "EnAlmacen",
        ]

        missing = [h for h in required if h not in headers]
        if missing:
            raise CommandError(f"Faltan columnas: {missing}")

        albaranes = {
            a.cod_albaran: a
            for a in AlbaranProveedorGestion.objects.select_related("team").all()
        }

        created = 0
        updated = 0
        skipped_no_albaran = 0
        skipped_no_linea = 0
        preview = []

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                data = dict(zip(headers, row))

                cod_albaran = clean(data.get("CodAlbaran"))
                linea = clean_int(data.get("Linea"))

                albaran = albaranes.get(cod_albaran)

                if not albaran:
                    skipped_no_albaran += 1
                    continue

                if linea is None:
                    skipped_no_linea += 1
                    continue

                payload = {
                    "cod_articulo_legacy": clean_int(data.get("CodArticulo")),
                    "cantidad": clean_decimal(data.get("Cantidad")),
                    "unidad": clean(data.get("Unidad")),
                    "cantidad_compra": clean_decimal(data.get("CantidadCompra")),
                    "unidad_compra": clean(data.get("UnidadCompra")),
                    "cantidad_x_unidad": clean_decimal(data.get("CantidadxUnidad")),
                    "precio_unitario": clean_decimal(data.get("PrecioUnitario")),
                    "importe_linea": clean_decimal(data.get("ImporteLinea"), "0.00"),
                    "facturado": clean_bool(data.get("Facturado")),
                    "factura_legacy": clean(data.get("Factura")),
                    "en_pedido": clean_bool(data.get("EnPedido")),
                    "en_partida": clean_bool(data.get("EnPartida")),
                    "fecha_entrega": clean_date(data.get("FechaEntrega")),
                    "recepcionado_por": clean(data.get("RecepcionadoPor")),
                    "importe_descuento": clean_decimal(data.get("ImporteDescuento"), "0.00"),
                    "descuento": clean_decimal(data.get("Descuento")),
                    "id_almacen_legacy": clean_int(data.get("IdAlmacen")),
                    "observaciones": clean(data.get("Observaciones")),
                    "tipo_recurso": clean(data.get("TipoRecurso")),
                    "cantidad_en_partidas": clean_decimal(data.get("CantidadEnPartidas")),
                    "en_almacen": clean_bool(data.get("EnAlmacen")),
                    "raw_data": {k: clean(v) for k, v in data.items()},
                }

                obj, was_created = AlbaranProveedorLineaGestion.objects.update_or_create(
                    albaran=albaran,
                    linea=linea,
                    defaults=payload,
                )

                if was_created:
                    created += 1
                else:
                    updated += 1

                if len(preview) < 10:
                    preview.append({
                        "team": str(albaran.team),
                        "cod_albaran": cod_albaran,
                        "linea": linea,
                        "cod_articulo": payload["cod_articulo_legacy"],
                        "cantidad": str(payload["cantidad"]),
                        "importe": str(payload["importe_linea"]),
                        "facturado": payload["facturado"],
                    })

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write("=== IMPORT ALBARANES LINEAS LEGACY ===")
        self.stdout.write(f"Fichero: {file_path}")
        self.stdout.write(f"Modo: {'COMMIT REAL' if commit else 'DRY-RUN'}")
        self.stdout.write(f"Creadas: {created}")
        self.stdout.write(f"Actualizadas: {updated}")
        self.stdout.write(f"Omitidas sin cabecera albarán: {skipped_no_albaran}")
        self.stdout.write(f"Omitidas sin número de línea: {skipped_no_linea}")

        self.stdout.write("")
        self.stdout.write("=== MUESTRA ===")
        for item in preview:
            self.stdout.write(str(item))

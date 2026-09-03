from pathlib import Path
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.gestion.models import (
    FacturaProveedorGestion,
    FacturaProveedorLineaGestion,
    AlbaranProveedorGestion,
)


DEFAULT_FILE = "/app/imports/access_sync_2026-05-19/tblFacturasLineas.xlsx"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except Exception:
        return None


def clean_decimal(value, places="0.0000"):
    if value in (None, ""):
        return Decimal(places)
    try:
        return Decimal(str(value)).quantize(Decimal(places))
    except (InvalidOperation, ValueError):
        return Decimal(places)


def clean_bool(value):
    if value in (True, 1, "1", "true", "True", "Sí", "SI", "si", "SÍ"):
        return True
    return False


class Command(BaseCommand):
    help = "Importa líneas de facturas legacy desde tblFacturasLineas.xlsx"

    def add_arguments(self, parser):
        parser.add_argument("--file", default=DEFAULT_FILE)
        parser.add_argument("--commit", action="store_true")

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        commit = options["commit"]

        if not file_path.exists():
            raise CommandError(f"No existe el fichero: {file_path}")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [clean(c.value) for c in ws[1]]

        required = [
            "CodFactura",
            "Linea",
            "CodArticulo",
            "CodAlbaran",
            "LineaAlbaran",
            "Cantidad",
            "PrecioUnitario",
            "ImporteLinea",
            "ImporteDescuento",
            "Descuento",
            "EnPartida",
            "CantidadEnPartidas",
            "EnAlmacen",
        ]

        missing = [h for h in required if h not in headers]
        if missing:
            raise CommandError(f"Faltan columnas: {missing}")

        facturas = {
            f.cod_factura: f
            for f in FacturaProveedorGestion.objects.select_related("team").all()
        }

        albaranes = {
            (a.team_id, a.cod_albaran): a
            for a in AlbaranProveedorGestion.objects.all()
        }

        created = 0
        updated = 0
        skipped_no_factura = 0
        skipped_no_linea = 0
        albaran_missing = 0
        preview = []

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                data = dict(zip(headers, row))

                cod_factura = clean(data.get("CodFactura"))
                linea = clean_int(data.get("Linea"))

                factura = facturas.get(cod_factura)
                if not factura:
                    skipped_no_factura += 1
                    continue

                if linea is None:
                    skipped_no_linea += 1
                    continue

                cod_albaran = clean(data.get("CodAlbaran"))
                albaran = None

                if cod_albaran:
                    albaran = albaranes.get((factura.team_id, cod_albaran))
                    if not albaran:
                        albaran_missing += 1

                payload = {
                    "albaran": albaran,
                    "cod_articulo_legacy": clean_int(data.get("CodArticulo")),
                    "cod_albaran_legacy": cod_albaran,
                    "linea_albaran_legacy": clean_int(data.get("LineaAlbaran")),
                    "cantidad": clean_decimal(data.get("Cantidad")),
                    "precio_unitario": clean_decimal(data.get("PrecioUnitario")),
                    "importe_linea": clean_decimal(data.get("ImporteLinea"), "0.00"),
                    "importe_descuento": clean_decimal(data.get("ImporteDescuento"), "0.00"),
                    "descuento": clean_decimal(data.get("Descuento")),
                    "en_partida": clean_bool(data.get("EnPartida")),
                    "cantidad_en_partidas": clean_decimal(data.get("CantidadEnPartidas")),
                    "en_almacen": clean_bool(data.get("EnAlmacen")),
                    "raw_data": {k: clean(v) for k, v in data.items()},
                }

                obj, was_created = FacturaProveedorLineaGestion.objects.update_or_create(
                    factura=factura,
                    linea=linea,
                    defaults=payload,
                )

                if was_created:
                    created += 1
                else:
                    updated += 1

                if len(preview) < 10:
                    preview.append({
                        "team": str(factura.team),
                        "cod_factura": cod_factura,
                        "linea": linea,
                        "cod_albaran": cod_albaran,
                        "albaran_link": bool(albaran),
                        "importe": str(payload["importe_linea"]),
                    })

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write("=== IMPORT FACTURAS LINEAS LEGACY ===")
        self.stdout.write(f"Fichero: {file_path}")
        self.stdout.write(f"Modo: {'COMMIT REAL' if commit else 'DRY-RUN'}")
        self.stdout.write(f"Creadas: {created}")
        self.stdout.write(f"Actualizadas: {updated}")
        self.stdout.write(f"Omitidas sin cabecera factura: {skipped_no_factura}")
        self.stdout.write(f"Omitidas sin número de línea: {skipped_no_linea}")
        self.stdout.write(f"CodAlbaran sin cabecera importada para el mismo Team: {albaran_missing}")

        self.stdout.write("")
        self.stdout.write("=== MUESTRA ===")
        for item in preview:
            self.stdout.write(str(item))

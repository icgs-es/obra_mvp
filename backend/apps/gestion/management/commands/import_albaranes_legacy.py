from pathlib import Path
from decimal import Decimal, InvalidOperation
from datetime import datetime, date

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.gestion.models import (
    AlbaranProveedorGestion,
    EmpresaGestionLegacy,
    Proveedor,
)


DEFAULT_FILE = "/app/imports/access_sync_2026-05-19/tblAlbaranes.xlsx"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except Exception:
        return None


def clean_decimal(value):
    if value in (None, ""):
        return Decimal("0.00")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def clean_bool(value):
    if value in (True, 1, "1", "true", "True", "Sí", "SI", "si", "SÍ"):
        return True
    return False


def clean_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


class Command(BaseCommand):
    help = "Importa cabeceras de albaranes legacy desde tblAlbaranes.xlsx"

    def add_arguments(self, parser):
        parser.add_argument("--file", default=DEFAULT_FILE)
        parser.add_argument("--commit", action="store_true")

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        commit = options["commit"]

        if not file_path.exists():
            raise CommandError(f"No existe el fichero: {file_path}")

        empresas = {
            e.legacy_id_empresa: e
            for e in EmpresaGestionLegacy.objects.select_related("team").all()
        }

        if not empresas:
            raise CommandError("No hay EmpresaGestionLegacy importadas.")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [clean(c.value) for c in ws[1]]

        required = [
            "CodObra",
            "CodAlbaran",
            "Presupuesto",
            "CodPresupuesto",
            "NumAlbaranProveedor",
            "FechaAlbaran",
            "ImporteAlbaran",
            "Descripcion",
            "FechaEntregaMercaderia",
            "Recepcionadopor",
            "OK_Presupuesto",
            "CodProveedor",
            "AutorizadoJefeObra",
            "Archivo",
            "AsignadoPartidaObra",
            "Empresa",
            "Situacion",
            "AsignadoFactura",
            "ImporteAsignadoFactura",
            "LineasAsignadas",
        ]

        missing = [h for h in required if h not in headers]
        if missing:
            raise CommandError(f"Faltan columnas: {missing}")

        created = 0
        updated = 0
        skipped_no_empresa = 0
        skipped_empresa_unknown = 0
        skipped_no_codalbaran = 0
        proveedor_missing = 0
        preview = []
        avisos = []

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                data = dict(zip(headers, row))

                cod_albaran = clean(data.get("CodAlbaran"))
                empresa_raw = clean_int(data.get("Empresa"))
                cod_proveedor = clean_int(data.get("CodProveedor"))

                if not cod_albaran:
                    skipped_no_codalbaran += 1
                    avisos.append(f"Fila {row_num}: sin CodAlbaran")
                    continue

                if empresa_raw in (None, 0):
                    skipped_no_empresa += 1
                    continue

                empresa_legacy = empresas.get(empresa_raw)
                if not empresa_legacy:
                    skipped_empresa_unknown += 1
                    avisos.append(f"Fila {row_num}: Empresa {empresa_raw} no existe en EmpresaGestionLegacy")
                    continue

                proveedor = None
                if cod_proveedor is not None:
                    proveedor = Proveedor.objects.filter(
                        team=empresa_legacy.team,
                        legacy_id_proveedor=cod_proveedor,
                    ).first()

                    if not proveedor:
                        proveedor_missing += 1
                        avisos.append(
                            f"Fila {row_num}: proveedor legacy {cod_proveedor} no existe para team {empresa_legacy.team_id}"
                        )

                payload = {
                    "empresa_legacy": empresa_legacy,
                    "proveedor": proveedor,
                    "cod_obra_legacy": clean(data.get("CodObra")),
                    "cod_proveedor_legacy": cod_proveedor,
                    "empresa_legacy_raw": empresa_raw,
                    "num_albaran_proveedor": clean(data.get("NumAlbaranProveedor")),
                    "fecha_albaran": clean_date(data.get("FechaAlbaran")),
                    "fecha_entrega_mercaderia": clean_date(data.get("FechaEntregaMercaderia")),
                    "importe_albaran": clean_decimal(data.get("ImporteAlbaran")),
                    "importe_asignado_factura": clean_decimal(data.get("ImporteAsignadoFactura")),
                    "descripcion": clean(data.get("Descripcion")),
                    "recepcionado_por": clean(data.get("Recepcionadopor")),
                    "presupuesto": clean_bool(data.get("Presupuesto")),
                    "cod_presupuesto_legacy": clean(data.get("CodPresupuesto")),
                    "ok_presupuesto": clean_bool(data.get("OK_Presupuesto")),
                    "autorizado_jefe_obra": clean_bool(data.get("AutorizadoJefeObra")),
                    "asignado_partida_obra": clean_bool(data.get("AsignadoPartidaObra")),
                    "asignado_factura": clean_bool(data.get("AsignadoFactura")),
                    "lineas_asignadas": clean_int(data.get("LineasAsignadas")) or 0,
                    "situacion": clean(data.get("Situacion")),
                    "archivo": clean(data.get("Archivo")),
                    "raw_data": {k: clean(v) for k, v in data.items()},
                }

                obj, was_created = AlbaranProveedorGestion.objects.update_or_create(
                    team=empresa_legacy.team,
                    cod_albaran=cod_albaran,
                    defaults=payload,
                )

                if was_created:
                    created += 1
                else:
                    updated += 1

                if len(preview) < 10:
                    preview.append({
                        "team": str(empresa_legacy.team),
                        "cod_albaran": cod_albaran,
                        "empresa": empresa_raw,
                        "cod_proveedor": cod_proveedor,
                        "proveedor": proveedor.nombre_comercial if proveedor else None,
                        "importe": str(payload["importe_albaran"]),
                        "asignado_factura": payload["asignado_factura"],
                    })

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write("=== IMPORT ALBARANES LEGACY ===")
        self.stdout.write(f"Fichero: {file_path}")
        self.stdout.write(f"Modo: {'COMMIT REAL' if commit else 'DRY-RUN'}")
        self.stdout.write(f"Creados: {created}")
        self.stdout.write(f"Actualizados: {updated}")
        self.stdout.write(f"Omitidos sin Empresa/Empresa 0: {skipped_no_empresa}")
        self.stdout.write(f"Omitidos Empresa desconocida: {skipped_empresa_unknown}")
        self.stdout.write(f"Omitidos sin CodAlbaran: {skipped_no_codalbaran}")
        self.stdout.write(f"Proveedor no encontrado: {proveedor_missing}")

        self.stdout.write("")
        self.stdout.write("=== MUESTRA ===")
        for item in preview:
            self.stdout.write(str(item))

        if avisos:
            self.stdout.write("")
            self.stdout.write("=== AVISOS PRIMEROS 30 ===")
            for aviso in avisos[:30]:
                self.stdout.write(aviso)

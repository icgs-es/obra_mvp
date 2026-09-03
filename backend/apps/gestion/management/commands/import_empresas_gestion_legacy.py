from pathlib import Path

import openpyxl
from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.gestion.models import EmpresaGestionLegacy


DEFAULT_FILE = "/app/imports/access_sync_2026-05-19/tblParametros.xlsx"

LEGACY_TO_TEAM = {
    1: 4,  # ADRI MARTI INVESTMENT S.L. -> ADRI MARTIN
    2: 1,  # INVERADRIDE GESTION S.L -> INVERADRIDE
    3: 2,  # INMOADRIDE COMPANY S.L. -> INMOADRIDE
    4: 3,  # INMOZASA WORLD -> INMOSAZA
}


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_int(value):
    if value in (None, ""):
        return 0
    return int(value)


class Command(BaseCommand):
    help = "Importa tblParametros.xlsx como mapeo Empresa legacy -> Team"

    def add_arguments(self, parser):
        parser.add_argument("--file", default=DEFAULT_FILE)
        parser.add_argument("--commit", action="store_true")

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        commit = options["commit"]

        if not file_path.exists():
            raise CommandError(f"No existe el fichero: {file_path}")

        Team = apps.get_model("usuarios", "Team")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [clean(c.value) for c in ws[1]]
        required = [
            "IdEmpresa",
            "NombreEmpresa",
            "CIFEmpresa",
            "DireccionEmpresa",
            "PoblacionEmpresa",
            "ProvinciaEmpresa",
            "CodPostalEmpresa",
            "PeriodoGestion",
            "UltCodigoFactura",
            "PrefijoFactura",
            "UltCodigoAlbaran",
            "PrefijoAlbaran",
            "ObraDefecto",
            "PrefijoPedido",
            "UltCodigoPedido",
        ]

        missing = [h for h in required if h not in headers]
        if missing:
            raise CommandError(f"Faltan columnas: {missing}")

        created = 0
        updated = 0
        errors = []
        preview = []

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                data = dict(zip(headers, row))

                legacy_id = clean_int(data.get("IdEmpresa"))
                team_id = LEGACY_TO_TEAM.get(legacy_id)

                if not team_id:
                    errors.append(f"Fila {row_num}: IdEmpresa {legacy_id} sin mapeo a Team")
                    continue

                team = Team.objects.filter(id=team_id).first()
                if not team:
                    errors.append(f"Fila {row_num}: Team {team_id} no existe")
                    continue

                payload = {
                    "team": team,
                    "nombre_empresa": clean(data.get("NombreEmpresa")),
                    "cif_empresa": clean(data.get("CIFEmpresa")),
                    "direccion_empresa": clean(data.get("DireccionEmpresa")),
                    "poblacion_empresa": clean(data.get("PoblacionEmpresa")),
                    "provincia_empresa": clean(data.get("ProvinciaEmpresa")),
                    "cod_postal_empresa": clean(data.get("CodPostalEmpresa")),
                    "periodo_gestion": clean(data.get("PeriodoGestion")),
                    "ult_codigo_factura": clean_int(data.get("UltCodigoFactura")),
                    "prefijo_factura": clean(data.get("PrefijoFactura")),
                    "ult_codigo_albaran": clean_int(data.get("UltCodigoAlbaran")),
                    "prefijo_albaran": clean(data.get("PrefijoAlbaran")),
                    "obra_defecto_legacy": clean_int(data.get("ObraDefecto")),
                    "prefijo_pedido": clean(data.get("PrefijoPedido")),
                    "ult_codigo_pedido": clean_int(data.get("UltCodigoPedido")),
                    "raw_data": {k: clean(v) for k, v in data.items()},
                }

                obj, was_created = EmpresaGestionLegacy.objects.update_or_create(
                    legacy_id_empresa=legacy_id,
                    defaults=payload,
                )

                if was_created:
                    created += 1
                else:
                    updated += 1

                preview.append(
                    {
                        "legacy_id": legacy_id,
                        "legacy_nombre": payload["nombre_empresa"],
                        "team_id": team.id,
                        "team": str(team),
                        "cif": payload["cif_empresa"],
                    }
                )

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write("=== IMPORT EMPRESAS GESTION LEGACY ===")
        self.stdout.write(f"Fichero: {file_path}")
        self.stdout.write(f"Modo: {'COMMIT REAL' if commit else 'DRY-RUN'}")
        self.stdout.write(f"Creados: {created}")
        self.stdout.write(f"Actualizados: {updated}")
        self.stdout.write("")

        self.stdout.write("=== MAPEO ===")
        for item in preview:
            self.stdout.write(str(item))

        if errors:
            self.stdout.write("")
            self.stdout.write("=== ERRORES ===")
            for e in errors:
                self.stdout.write(e)

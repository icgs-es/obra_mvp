from collections import Counter
from decimal import Decimal, InvalidOperation
import re

from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from usuarios.models import Team
from rrhh.models import Empleado, GrupoTrabajo, EmpleadoGrupoTrabajo


EMPRESA_TO_TEAM = {
    "INVERADRIDE GESTION S.L.": "INVERADRIDE",
    "INVERADRIDE GESTIÓN S.L.": "INVERADRIDE",
    "ADRI MARTIN INVESTMENTS S.L.": "ADRI MARTIN",
    "INMOSAZA WORLD, S.L.": "INMOSAZA",
    "INMOADRIDE COMPANY, S.L.": "INMOADRIDE",
}


def norm(value):
    if value is None:
        return ""
    value = str(value).strip()
    value = re.sub(r"\s+", " ", value)
    return value


def norm_key(value):
    return norm(value).upper()


def name_key(value):
    value = norm(value).upper().replace(",", " ")
    value = re.sub(r"\s+", " ", value)
    return value


def decimal_or_none(value):
    value = norm(value)
    if not value:
        return None
    value = value.replace("€", "").replace("%", "").replace(" ", "")
    value = value.replace(".", "").replace(",", ".")
    try:
        return Decimal(value)
    except InvalidOperation:
        return None


def date_or_none(value):
    # openpyxl devuelve date/datetime cuando la celda es fecha real.
    # Algunas columnas de SeaTable, como BAJA, pueden venir como booleano True/False.
    # Eso no debe convertirse en fecha.
    from datetime import date, datetime

    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    value = norm(value)
    if not value:
        return None

    # Intento básico para fechas tipo dd/mm/yyyy o yyyy-mm-dd.
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass

    return None


def classify_area(puesto, profesion):
    text = f"{puesto} {profesion}".upper()

    if any(x in text for x in ["COMERCIAL", "EXPANSION", "EXPANSIÓN"]):
        return Empleado.AreaPrincipal.COMERCIAL

    if any(x in text for x in ["ADMIN", "ADMON", "GESTION", "GESTIÓN", "AUXILIAR"]):
        return Empleado.AreaPrincipal.ADMINISTRACION

    if any(x in text for x in ["ARQUITECT", "DISEÑ", "DISEN", "INGENIERO"]):
        return Empleado.AreaPrincipal.ARQUITECTURA

    if any(x in text for x in ["LIMPIEZA", "SERVICIOS"]):
        return Empleado.AreaPrincipal.SERVICIOS

    if any(x in text for x in [
        "OFICIAL", "PEON", "PEÓN", "ENCARGADO", "JEFE DE OBRA",
        "ALBAÑIL", "ALBANIL", "PLADUR", "PINTOR", "FONTANERO",
        "ELECTRICISTA", "DEMOLICION", "DEMOLICIÓN", "MAQUINISTA",
        "ALMACEN", "ALMACÉN", "VERTICALES", "MULTISERVICIOS"
    ]):
        return Empleado.AreaPrincipal.OBRA

    return Empleado.AreaPrincipal.OTRO


def tipo_relacion_from(puesto, profesion):
    text = f"{puesto} {profesion}".upper()
    if "AUTONOMO" in text or "AUTÓNOMO" in text:
        return Empleado.TipoRelacion.AUTONOMO
    return Empleado.TipoRelacion.PROPIO


def group_name_for_area(area):
    return {
        Empleado.AreaPrincipal.OBRA: "Obra",
        Empleado.AreaPrincipal.ADMINISTRACION: "Administración",
        Empleado.AreaPrincipal.COMERCIAL: "Comercial",
        Empleado.AreaPrincipal.GERENCIA: "Gerencia",
        Empleado.AreaPrincipal.ARQUITECTURA: "Arquitectura",
        Empleado.AreaPrincipal.SERVICIOS: "Servicios",
        Empleado.AreaPrincipal.OTRO: None,
    }.get(area)


class Command(BaseCommand):
    help = "Importa carga inicial de RRHH desde Excel exportado de SeaTable."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument("--default-team-id", type=int, default=None)
        parser.add_argument("--commit", action="store_true")

    def handle(self, *args, **options):
        xlsx = options["file"]
        default_team_id = options["default_team_id"]
        commit = options["commit"]

        wb = load_workbook(xlsx, data_only=True)
        ws = wb[wb.sheetnames[0]]

        headers = [norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        idx = {h.upper(): i + 1 for i, h in enumerate(headers) if h}

        def get(row, col):
            c = idx.get(col.upper())
            if not c:
                return ""
            return norm(ws.cell(row, c).value)

        def raw(row, col):
            c = idx.get(col.upper())
            if not c:
                return None
            return ws.cell(row, c).value

        teams_by_name = {norm_key(t.name): t for t in Team.objects.all()}
        default_team = Team.objects.filter(id=default_team_id).first() if default_team_id else None

        created = 0
        updated = 0
        matched_name = 0
        skipped_no_team = 0
        skipped_no_name = 0
        memberships_created = 0

        area_counter = Counter()
        team_counter = Counter()
        relation_counter = Counter()
        situation_counter = Counter()

        with transaction.atomic():
            for row in range(2, ws.max_row + 1):
                nombre = get(row, "NOMBRE")
                if not nombre:
                    skipped_no_name += 1
                    continue

                empresa = get(row, "EMPRESA")
                team_name = EMPRESA_TO_TEAM.get(norm_key(empresa), "")
                team = teams_by_name.get(norm_key(team_name)) if team_name else None

                if not team and not empresa and default_team:
                    team = default_team

                if not team:
                    skipped_no_team += 1
                    continue

                puesto = get(row, "PUESTO")
                profesion = get(row, "PROFESION")
                area = classify_area(puesto, profesion)
                tipo_relacion = tipo_relacion_from(puesto, profesion)

                situacion_raw = norm_key(get(row, "SITUACION"))
                situacion = Empleado.Situacion.BAJA if situacion_raw == "BAJA" else Empleado.Situacion.ACTIVO

                area_counter[area] += 1
                team_counter[team.name] += 1
                relation_counter[tipo_relacion] += 1
                situation_counter[situacion] += 1

                existing = None

                # 1) Si ya existe por referencia SeaTable
                ref = f"seatable_row:{row}"
                existing = Empleado.objects.filter(
                    team=team,
                    origen="seatable",
                    referencia_externa=ref,
                ).first()

                # 2) Si no existe, buscar por nombre normalizado dentro del team
                if not existing:
                    for e in Empleado.objects.filter(team=team):
                        if name_key(e.nombre_completo) == name_key(nombre):
                            existing = e
                            matched_name += 1
                            break

                raw_data = {
                    "source": "seatable",
                    "row": row,
                    "empresa": empresa,
                    "puesto": puesto,
                    "profesion": profesion,
                    "situacion": get(row, "SITUACION"),
                }

                values = {
                    "nombre_completo": nombre,
                    "nif_nie": get(row, "NIF/NIE"),
                    "telefono": get(row, "TELEFONO"),
                    "empresa_empleadora": empresa or team.name,
                    "tipo_relacion": tipo_relacion,
                    "area_principal": area,
                    "puesto": puesto,
                    "profesion": profesion,
                    "situacion": situacion,
                    "fecha_alta": date_or_none(raw(row, "FECHA ALTA")),
                    "fecha_baja": date_or_none(raw(row, "BAJA")),
                    "coste_bruto_nomina": decimal_or_none(get(row, "COSTE BRUTO NOMINA")),
                    "sueldo": decimal_or_none(get(row, "SUELDO")),
                    "coste_hora": decimal_or_none(get(row, "COSTE HORA")),
                    "coeficiente": decimal_or_none(get(row, "COEFICIENTE")),
                    "retencion": decimal_or_none(get(row, "RETENCION")),
                    "precio_bruto_hora": decimal_or_none(get(row, "PRECIO BRUTO HRA")),
                    "coste_bruto_ss": decimal_or_none(get(row, "COSTE BRUTO S S")),
                    "es_fichable": True,
                    "es_planificable_obra": area == Empleado.AreaPrincipal.OBRA,
                    "activo": situacion == Empleado.Situacion.ACTIVO,
                    "raw_data": raw_data,
                }

                if existing:
                    for field, value in values.items():
                        setattr(existing, field, value)
                    if existing.origen != "access_empleado_obra":
                        existing.origen = "seatable"
                        existing.referencia_externa = ref
                    existing.save()
                    empleado = existing
                    updated += 1
                else:
                    empleado = Empleado.objects.create(
                        team=team,
                        origen="seatable",
                        referencia_externa=ref,
                        codigo=f"ST-{row}",
                        **values,
                    )
                    created += 1

                group_name = group_name_for_area(area)
                if group_name:
                    grupo = GrupoTrabajo.objects.filter(team=team, nombre=group_name).first()
                    if grupo:
                        _, m_created = EmpleadoGrupoTrabajo.objects.get_or_create(
                            empleado=empleado,
                            grupo=grupo,
                            defaults={"rol": puesto, "activo": True},
                        )
                        if m_created:
                            memberships_created += 1

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write(f"Modo: {'COMMIT' if commit else 'DRY-RUN'}")
        self.stdout.write(f"Creados: {created}")
        self.stdout.write(f"Actualizados/enriquecidos: {updated}")
        self.stdout.write(f"Coincidencias por nombre: {matched_name}")
        self.stdout.write(f"Membresías creadas: {memberships_created}")
        self.stdout.write(f"Omitidos sin team: {skipped_no_team}")
        self.stdout.write(f"Omitidos sin nombre: {skipped_no_name}")

        self.stdout.write("")
        self.stdout.write("Por Team:")
        for k, v in team_counter.most_common():
            self.stdout.write(f"  {k}: {v}")

        self.stdout.write("")
        self.stdout.write("Por área:")
        for k, v in area_counter.most_common():
            self.stdout.write(f"  {k}: {v}")

        self.stdout.write("")
        self.stdout.write("Por relación:")
        for k, v in relation_counter.most_common():
            self.stdout.write(f"  {k}: {v}")

        self.stdout.write("")
        self.stdout.write("Por situación:")
        for k, v in situation_counter.most_common():
            self.stdout.write(f"  {k}: {v}")

        if not commit:
            self.stdout.write(self.style.WARNING("DRY-RUN: no se guardaron cambios."))
        else:
            self.stdout.write(self.style.SUCCESS("Importación aplicada."))

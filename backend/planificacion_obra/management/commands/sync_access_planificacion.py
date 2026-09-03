from __future__ import annotations

import json
from collections import Counter
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.apps import apps
from django.core.management import call_command
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook


REQUIRED_XLSX = [
    "tblTareas.xlsx",
    "tblTareasRecursosInicial.xlsx",
    "tblTareasRecursos.xlsx",
    "tblRecursos.xlsx",
    "tblRecursoAlmacen.xlsx",
    "tblPartida.xlsx",
]


def s(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).replace("\ufeff", "").strip()


def to_int(value):
    value = s(value)
    if not value:
        return None
    try:
        return int(Decimal(value.replace(",", ".")))
    except (InvalidOperation, ValueError):
        return None


def to_decimal(value):
    value = s(value)
    if not value:
        return None
    try:
        return Decimal(value.replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def to_bool(value):
    if isinstance(value, bool):
        return value
    value = s(value).lower()
    return value in {"true", "1", "sí", "si", "yes", "y", "-1"}


def to_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    txt = s(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(txt, fmt).date()
        except ValueError:
            pass
    return None


def to_time(value):
    if value in (None, ""):
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    txt = s(value)
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(txt, fmt).time()
        except ValueError:
            pass
    return None


def clean_json(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def clean_row(row):
    return {str(k): clean_json(v) for k, v in row.items()}


def read_xlsx(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    try:
        headers = [s(v) for v in next(rows)]
    except StopIteration:
        wb.close()
        return

    for row_number, raw in enumerate(rows, start=2):
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = raw[idx] if idx < len(raw) else None
        row["_xlsx_row_number"] = row_number
        yield row

    wb.close()


def get_team(team_name):
    Team = apps.get_model("usuarios", "Team")
    for field in ["nombre", "name", "razon_social"]:
        if field in {f.name for f in Team._meta.fields}:
            obj = Team.objects.filter(**{field: team_name}).first()
            if obj:
                return obj
    for field in ["nombre", "name", "razon_social"]:
        if field in {f.name for f in Team._meta.fields}:
            obj = Team.objects.filter(**{f"{field}__icontains": team_name}).first()
            if obj:
                return obj
    return None


def tarea_exact_key_from_access(row):
    return (
        to_int(row.get("CodObra")),
        to_int(row.get("CodFase")),
        s(row.get("CodVivienda")),
        s(row.get("Planta")),
        s(row.get("Capitulo")),
        s(row.get("Partida")),
        to_int(row.get("Orden")),
    )


def tarea_exact_key_from_model(t):
    return (
        t.legacy_cod_obra,
        t.legacy_cod_fase,
        s(t.legacy_cod_vivienda),
        s(t.legacy_planta),
        s(t.legacy_capitulo),
        s(t.legacy_partida),
        t.legacy_orden,
    )


def invalid_tarea_key(key):
    return None in [key[0], key[1], key[6]] or not key[3] or not key[4] or not key[5]


def make_legacy_key(key):
    return "|".join("" if v is None else str(v) for v in key)


def normalize_tipo_mov(value):
    txt = s(value).upper()
    return txt or "SIN_TIPO"


class Command(BaseCommand):
    help = "Sincronización Access -> Planificación de Obra. Dry-run por defecto. Commit insert-only con backup."

    def add_arguments(self, parser):
        parser.add_argument("--folder", required=True)
        parser.add_argument("--team", required=True)
        parser.add_argument("--sample", type=int, default=20)
        parser.add_argument("--json-out", default="")
        parser.add_argument("--commit", action="store_true")

    def handle(self, *args, **options):
        folder = Path(options["folder"])
        team_name = options["team"]
        sample_limit = int(options["sample"] or 20)
        json_out = options["json_out"]
        commit = bool(options["commit"])

        if not folder.exists():
            raise CommandError(f"No existe la carpeta: {folder}")

        missing_files = [name for name in REQUIRED_XLSX if not (folder / name).exists()]
        if missing_files:
            raise CommandError(f"Faltan XLSX requeridos: {missing_files}")

        team = get_team(team_name)
        if not team:
            raise CommandError(f"No encuentro Team: {team_name}")

        ObraPlanificacion = apps.get_model("planificacion_obra", "ObraPlanificacion")
        UnidadObra = apps.get_model("planificacion_obra", "UnidadObra")
        CapituloCatalogo = apps.get_model("planificacion_obra", "CapituloCatalogo")
        PartidaCatalogo = apps.get_model("planificacion_obra", "PartidaCatalogo")
        RecursoCatalogo = apps.get_model("planificacion_obra", "RecursoCatalogo")
        EmpleadoObra = apps.get_model("planificacion_obra", "EmpleadoObra")
        AlmacenObra = apps.get_model("planificacion_obra", "AlmacenObra")
        TareaObra = apps.get_model("planificacion_obra", "TareaObra")
        TareaRecursoPrevisto = apps.get_model("planificacion_obra", "TareaRecursoPrevisto")
        TareaRecursoReal = apps.get_model("planificacion_obra", "TareaRecursoReal")
        RecursoAlmacenMovimiento = apps.get_model("planificacion_obra", "RecursoAlmacenMovimiento")

        report = {
            "mode": "COMMIT_INSERT_ONLY" if commit else "DRY_RUN_INSERT_ONLY",
            "commit": commit,
            "folder": str(folder),
            "team": {"id": team.id, "name": str(team)},
            "tables": {},
        }

        # Catálogos base en memoria
        obras = {o.legacy_cod_obra: o for o in ObraPlanificacion.objects.filter(team=team)}
        unidades = {
            (u.legacy_cod_obra, u.legacy_cod_fase, s(u.legacy_cod_vivienda)): u
            for u in UnidadObra.objects.filter(team=team)
        }
        capitulos = {c.codigo: c for c in CapituloCatalogo.objects.filter(team=team)}

        partidas_by_codigo = {}
        partidas_by_cap_codigo = {}
        for p in PartidaCatalogo.objects.filter(team=team).select_related("capitulo"):
            partidas_by_codigo.setdefault(p.codigo, p)
            if p.capitulo_id:
                partidas_by_cap_codigo.setdefault((p.capitulo.codigo, p.codigo), p)

        empleados = {
            e.legacy_id: e for e in EmpleadoObra.objects.filter(team=team).exclude(legacy_id=None)
        }
        almacenes = {a.legacy_id_almacen: a for a in AlmacenObra.objects.filter(team=team)}

        # Recursos
        existing_recursos = set(RecursoCatalogo.objects.filter(team=team).values_list("legacy_id", flat=True))
        access_recursos = set()
        missing_recursos = []

        for row in read_xlsx(folder / "tblRecursos.xlsx"):
            legacy_id = to_int(row.get("IdRecurso"))
            if legacy_id is None:
                continue
            access_recursos.add(legacy_id)
            if legacy_id not in existing_recursos:
                missing_recursos.append(row)

        report["tables"]["tblRecursos"] = {
            "existing_bd": len(existing_recursos),
            "access_ids": len(access_recursos),
            "missing_count": len(missing_recursos),
            "extra_in_bd_count_no_action": len(existing_recursos - access_recursos),
            "missing_sample": [clean_row(r) for r in missing_recursos[:sample_limit]],
        }

        # Tareas
        existing_tarea_exact = set()
        for t in TareaObra.objects.filter(team=team).only(
            "legacy_cod_obra", "legacy_cod_fase", "legacy_cod_vivienda",
            "legacy_planta", "legacy_capitulo", "legacy_partida", "legacy_orden"
        ):
            existing_tarea_exact.add(tarea_exact_key_from_model(t))

        seen_tareas = set()
        duplicate_tareas = 0
        invalid_tareas = 0
        existing_tareas = 0
        missing_tareas = []
        by_cod_obra = Counter()

        for row in read_xlsx(folder / "tblTareas.xlsx"):
            key = tarea_exact_key_from_access(row)
            by_cod_obra[s(row.get("CodObra"))] += 1

            if key in seen_tareas:
                duplicate_tareas += 1
            seen_tareas.add(key)

            if invalid_tarea_key(key):
                invalid_tareas += 1
                continue

            if key in existing_tarea_exact:
                existing_tareas += 1
            else:
                row["_exact_key"] = key
                missing_tareas.append(row)

        report["tables"]["tblTareas"] = {
            "existing_bd_exact_keys": len(existing_tarea_exact),
            "access_rows": len(seen_tareas),
            "existing_by_exact_key": existing_tareas,
            "missing_count": len(missing_tareas),
            "invalid_key": invalid_tareas,
            "duplicate_in_access": duplicate_tareas,
            "by_cod_obra": dict(by_cod_obra),
            "missing_sample": [clean_row(r) for r in missing_tareas[:sample_limit]],
        }

        # Previsto
        existing_prev = set(TareaRecursoPrevisto.objects.filter(team=team).values_list("legacy_row_number", flat=True))
        access_prev = set()
        missing_prev = []

        for row in read_xlsx(folder / "tblTareasRecursosInicial.xlsx"):
            row_number = row["_xlsx_row_number"]
            access_prev.add(row_number)
            if row_number not in existing_prev:
                missing_prev.append(row)

        report["tables"]["tblTareasRecursosInicial"] = {
            "existing_bd": len(existing_prev),
            "access_row_numbers": len(access_prev),
            "missing_count": len(missing_prev),
            "extra_in_bd_count_no_action": len(existing_prev - access_prev),
            "missing_sample": [clean_row(r) for r in missing_prev[:sample_limit]],
        }

        # Real
        existing_real = set(TareaRecursoReal.objects.filter(team=team).values_list("legacy_id_recurso_tarea", flat=True))
        access_real = set()
        missing_real = []

        for row in read_xlsx(folder / "tblTareasRecursos.xlsx"):
            legacy_id = to_int(row.get("IdRecursoTarea"))
            if legacy_id is None:
                continue
            access_real.add(legacy_id)
            if legacy_id not in existing_real:
                missing_real.append(row)

        report["tables"]["tblTareasRecursos"] = {
            "existing_bd": len(existing_real),
            "access_ids": len(access_real),
            "missing_count": len(missing_real),
            "extra_in_bd_count_no_action": len(existing_real - access_real),
            "extra_in_bd_ids_no_action": sorted(list(existing_real - access_real))[:sample_limit],
            "missing_sample": [clean_row(r) for r in missing_real[:sample_limit]],
        }

        # Movimientos almacén
        existing_mov = set(RecursoAlmacenMovimiento.objects.filter(team=team).values_list("legacy_id_movimiento", flat=True))
        access_mov = set()
        missing_mov = []

        for row in read_xlsx(folder / "tblRecursoAlmacen.xlsx"):
            legacy_id = to_int(row.get("IdRecurso"))
            if legacy_id is None:
                continue
            access_mov.add(legacy_id)
            if legacy_id not in existing_mov:
                missing_mov.append(row)

        report["tables"]["tblRecursoAlmacen"] = {
            "existing_bd": len(existing_mov),
            "access_ids": len(access_mov),
            "missing_count": len(missing_mov),
            "extra_in_bd_count_no_action": len(existing_mov - access_mov),
            "missing_sample": [clean_row(r) for r in missing_mov[:sample_limit]],
        }

        created = {
            "tblRecursos": 0,
            "tblTareas": 0,
            "tblTareasRecursosInicial": 0,
            "tblTareasRecursos": 0,
            "tblRecursoAlmacen": 0,
            "skipped": {},
        }

        backup_dir = None

        if commit:
            ts = timezone.now().strftime("%Y%m%d_%H%M%S")
            backup_dir = Path("/app/backups") / f"planificacion_sync_access_{ts}"
            backup_dir.mkdir(parents=True, exist_ok=True)

            call_command(
                "dumpdata",
                "planificacion_obra",
                output=str(backup_dir / "planificacion_obra_before.json"),
                indent=2,
                verbosity=0,
            )

            with transaction.atomic():
                # 1. Crear recursos catálogo faltantes
                for row in missing_recursos:
                    legacy_id = to_int(row.get("IdRecurso"))
                    if RecursoCatalogo.objects.filter(team=team, legacy_id=legacy_id).exists():
                        continue

                    capitulo = capitulos.get(s(row.get("Capitulo")))

                    RecursoCatalogo.objects.create(
                        team=team,
                        legacy_id=legacy_id,
                        nombre=s(row.get("NombreRecurso")) or f"Recurso {legacy_id}",
                        tipo=s(row.get("Tipo")),
                        unidad=s(row.get("Unidad")),
                        capitulo=capitulo,
                        stock=to_decimal(row.get("Stock")),
                        ultimo_precio_unidad=to_decimal(row.get("UltPrecioUnidad")),
                        precio_unidad_uso=to_decimal(row.get("PrecioUnidadUso")),
                        control_stock=to_bool(row.get("ControlStock")),
                        observaciones=s(row.get("Observaciones")),
                        raw_data=clean_row(row),
                    )
                    created["tblRecursos"] += 1

                recursos = {
                    r.legacy_id: r
                    for r in RecursoCatalogo.objects.filter(team=team).exclude(legacy_id=None)
                }

                # 2. Crear tareas faltantes
                for row in missing_tareas:
                    key = row["_exact_key"]
                    cod_obra, cod_fase, cod_vivienda, planta, cod_capitulo, cod_partida, orden = key

                    if TareaObra.objects.filter(
                        team=team,
                        legacy_cod_obra=cod_obra,
                        legacy_cod_fase=cod_fase,
                        legacy_cod_vivienda=cod_vivienda,
                        legacy_planta=planta,
                        legacy_capitulo=cod_capitulo,
                        legacy_partida=cod_partida,
                        legacy_orden=orden,
                    ).exists():
                        continue

                    obra = obras.get(cod_obra)
                    if not obra:
                        created["skipped"]["tareas_sin_obra"] = created["skipped"].get("tareas_sin_obra", 0) + 1
                        continue

                    unidad = unidades.get((cod_obra, cod_fase, cod_vivienda))
                    capitulo = capitulos.get(cod_capitulo)
                    partida = partidas_by_cap_codigo.get((cod_capitulo, cod_partida)) or partidas_by_codigo.get(cod_partida)

                    TareaObra.objects.create(
                        team=team,
                        legacy_key=make_legacy_key(key),
                        obra=obra,
                        unidad_obra=unidad,
                        capitulo=capitulo,
                        partida=partida,
                        legacy_cod_obra=cod_obra,
                        legacy_cod_fase=cod_fase,
                        legacy_cod_vivienda=cod_vivienda,
                        legacy_planta=planta,
                        legacy_capitulo=cod_capitulo,
                        legacy_partida=cod_partida,
                        legacy_orden=orden,
                        programacion=s(row.get("Programacion")),
                        porcentaje_completado=to_decimal(row.get("PorcentajeCompletado")),
                        inicio_tarea=to_date(row.get("InicioTarea")),
                        fin_tarea=to_date(row.get("FinTarea")),
                        dias=to_decimal(row.get("Dias")),
                        horas=to_decimal(row.get("Horas")),
                        inicio_real=to_date(row.get("InicioReal")),
                        fin_real=to_date(row.get("FinReal")),
                        dias_reales=to_decimal(row.get("DiasReales")),
                        horas_reales=to_decimal(row.get("HorasReales")),
                        inicio_estimado=to_date(row.get("InicioEstimado")),
                        fin_estimado=to_date(row.get("FinEstimado")),
                        dias_estimados=to_decimal(row.get("DiasEstimados")),
                        horas_estimadas=to_decimal(row.get("HorasEstimadas")),
                        personas_a_utilizar=to_decimal(row.get("Personasautilizar")),
                        personas_utilizadas=to_decimal(row.get("Personasutilizadas")),
                        importe_tarea=to_decimal(row.get("ImporteTarea")),
                        importe_tarea_real=to_decimal(row.get("ImporteTareaReal")),
                        importe_tarea_estimado=to_decimal(row.get("ImporteTareaEstimado")),
                        tipo_partida=s(row.get("TipoPartida")),
                        unidad=s(row.get("Unidad")),
                        cantidad=to_decimal(row.get("Cantidad")),
                        precio_unidad=to_decimal(row.get("PrecioUnidad")),
                        con_incidencias=to_bool(row.get("ConIncidencias")),
                        observaciones=s(row.get("Observaciones")),
                        raw_data=clean_row(row),
                    )
                    created["tblTareas"] += 1

                tareas_no_cap = {}
                for t in TareaObra.objects.filter(team=team):
                    tareas_no_cap.setdefault(
                        (
                            t.legacy_cod_obra,
                            t.legacy_cod_fase,
                            s(t.legacy_cod_vivienda),
                            s(t.legacy_planta),
                            s(t.legacy_partida),
                        ),
                        t,
                    )

                # 3. Crear recursos previstos faltantes
                for row in missing_prev:
                    row_number = row["_xlsx_row_number"]
                    if TareaRecursoPrevisto.objects.filter(team=team, legacy_row_number=row_number).exists():
                        continue

                    cod_obra = to_int(row.get("CodObra"))
                    cod_fase = to_int(row.get("CodFase"))
                    cod_vivienda = s(row.get("CodVivienda"))
                    planta = s(row.get("Planta"))
                    cod_partida = s(row.get("CodPartida"))
                    id_recurso = to_int(row.get("IdRecurso"))
                    orden_recurso = to_int(row.get("Orden"))

                    unidad = unidades.get((cod_obra, cod_fase, cod_vivienda))
                    partida = partidas_by_codigo.get(cod_partida)
                    recurso = recursos.get(id_recurso)
                    tarea = tareas_no_cap.get((cod_obra, cod_fase, cod_vivienda, planta, cod_partida))

                    TareaRecursoPrevisto.objects.create(
                        team=team,
                        legacy_row_number=row_number,
                        tarea_obra=tarea,
                        unidad_obra=unidad,
                        partida=partida,
                        recurso=recurso,
                        legacy_cod_obra=cod_obra,
                        legacy_cod_fase=cod_fase,
                        legacy_cod_vivienda=cod_vivienda,
                        legacy_planta=planta,
                        legacy_cod_partida=cod_partida,
                        legacy_id_recurso=id_recurso,
                        legacy_id_recurso_old=to_int(row.get("IdRecurso_old")),
                        legacy_orden_recurso=orden_recurso,
                        unidad=s(row.get("Unidad")),
                        precio_unidad=to_decimal(row.get("PrecioUnidad")),
                        cantidad=to_decimal(row.get("Cantidad")),
                        costo_recurso=to_decimal(row.get("CostoRecurso")),
                        control_suministros=to_bool(row.get("ControlSuministros")),
                        avisar=to_int(row.get("Avisar")),
                        fecha_estimada_entrega=to_date(row.get("FechaEstimadaEntrega")),
                        raw_data=clean_row(row),
                    )
                    created["tblTareasRecursosInicial"] += 1

                # 4. Crear recursos reales faltantes
                # Insert-only: crea IdRecursoTarea faltantes. No actualiza existentes.
                movimientos = {
                    m.legacy_id_movimiento: m
                    for m in RecursoAlmacenMovimiento.objects.filter(team=team).exclude(legacy_id_movimiento=None)
                }

                for row in missing_real:
                    legacy_id = to_int(row.get("IdRecursoTarea"))
                    if legacy_id is None:
                        created["skipped"]["recursos_reales_sin_id"] = created["skipped"].get("recursos_reales_sin_id", 0) + 1
                        continue

                    if TareaRecursoReal.objects.filter(team=team, legacy_id_recurso_tarea=legacy_id).exists():
                        continue

                    cod_obra = to_int(row.get("CodObra"))
                    cod_fase = to_int(row.get("CodFase"))
                    cod_vivienda = s(row.get("CodVivienda"))
                    planta = s(row.get("Planta"))
                    cod_capitulo = s(row.get("Capitulo"))
                    cod_partida = s(row.get("Partida"))
                    id_recurso = to_int(row.get("IdRecurso"))
                    tipo_recurso = s(row.get("TipoRecurso"))
                    legacy_personal_raw = to_int(row.get("Personal"))
                    legacy_id_mov = to_int(row.get("IdMovimientoAlmacen"))
                    labor_types = {"M.O. ADM.", "PER. CONT.", "M.O. CONT."}
                    effective_personal = id_recurso if tipo_recurso in labor_types else legacy_personal_raw

                    unidad = unidades.get((cod_obra, cod_fase, cod_vivienda))
                    partida = partidas_by_cap_codigo.get((cod_capitulo, cod_partida)) or partidas_by_codigo.get(cod_partida)
                    recurso = recursos.get(id_recurso)
                    empleado = empleados.get(effective_personal) if effective_personal not in (None, 0) else None
                    movimiento = movimientos.get(legacy_id_mov)
                    tarea = tareas_no_cap.get((cod_obra, cod_fase, cod_vivienda, planta, cod_partida))

                    TareaRecursoReal.objects.create(
                        team=team,
                        legacy_id_recurso_tarea=legacy_id,
                        tarea_obra=tarea,
                        unidad_obra=unidad,
                        partida=partida,
                        recurso=recurso,
                        empleado=empleado,
                        movimiento_almacen=movimiento,
                        legacy_cod_obra=cod_obra,
                        legacy_cod_fase=cod_fase,
                        legacy_cod_vivienda=cod_vivienda,
                        legacy_planta=planta,
                        legacy_capitulo=cod_capitulo,
                        legacy_partida=cod_partida,
                        legacy_id_recurso=id_recurso,
                        legacy_tipo_recurso=tipo_recurso,
                        legacy_personal=effective_personal,
                        legacy_id_movimiento_almacen=legacy_id_mov,
                        legacy_orden_recurso=to_int(row.get("Orden")),
                        unidad=s(row.get("Unidad")),
                        cantidad=to_decimal(row.get("Cantidad")),
                        precio_unidad=to_decimal(row.get("PrecioUnidad")),
                        dias=to_decimal(row.get("Dias")),
                        dias_reales=to_decimal(row.get("DiasReales")),
                        horas=to_decimal(row.get("Horas")),
                        horas_reales=to_decimal(row.get("HorasReales")),
                        inicio_recurso_real=to_date(row.get("InicioRecursoReal")),
                        fin_recurso_real=to_date(row.get("FinRecursoReal")),
                        costo_recurso=to_decimal(row.get("CostoRecurso")),
                        costo_recurso_real=to_decimal(row.get("CostoRecursoReal")),
                        control_suministros=to_bool(row.get("ControlSuministros")),
                        avisar=to_int(row.get("Avisar")),
                        id_proveedor=s(row.get("IdProveedor")),
                        cod_albaran=s(row.get("CodAlbaran")),
                        num_linea_albaran=to_int(row.get("NumLineaAlbaran")),
                        cod_factura=s(row.get("CodFactura")),
                        num_linea_factura=to_int(row.get("NumLineaFactura")),
                        observaciones=s(row.get("Observaciones")),
                        raw_data=clean_row(row),
                    )
                    created["tblTareasRecursos"] += 1

                # 5. Crear movimientos almacén faltantes
                for row in missing_mov:
                    legacy_id_mov = to_int(row.get("IdRecurso"))
                    if RecursoAlmacenMovimiento.objects.filter(team=team, legacy_id_movimiento=legacy_id_mov).exists():
                        continue

                    legacy_id_almacen = s(row.get("IdAlmacen"))
                    legacy_cod_recurso = to_int(row.get("CodRecurso"))
                    legacy_cod_obra = to_int(row.get("CodObra"))
                    legacy_cod_fase = to_int(row.get("CodFase"))
                    legacy_cod_vivienda = s(row.get("Vivienda"))
                    legacy_partida = s(row.get("Partida"))
                    legacy_cod_personal = to_int(row.get("CodPersonal"))

                    almacen = almacenes.get(legacy_id_almacen)
                    recurso = recursos.get(legacy_cod_recurso)
                    obra = obras.get(legacy_cod_obra)
                    unidad = unidades.get((legacy_cod_obra, legacy_cod_fase, legacy_cod_vivienda))
                    empleado = empleados.get(legacy_cod_personal)
                    partida = partidas_by_codigo.get(legacy_partida)

                    RecursoAlmacenMovimiento.objects.create(
                        team=team,
                        legacy_id_movimiento=legacy_id_mov,
                        almacen=almacen,
                        recurso=recurso,
                        obra=obra,
                        unidad_obra=unidad,
                        empleado=empleado,
                        partida=partida,
                        legacy_id_almacen=legacy_id_almacen,
                        legacy_cod_recurso=legacy_cod_recurso,
                        legacy_cod_obra=legacy_cod_obra,
                        legacy_cod_fase=legacy_cod_fase,
                        legacy_cod_vivienda=legacy_cod_vivienda,
                        legacy_planta=s(row.get("Planta")),
                        legacy_capitulo=s(row.get("Capitulo")),
                        legacy_partida=legacy_partida,
                        legacy_cod_personal=legacy_cod_personal,
                        unidad=s(row.get("Unidad")),
                        cantidad=to_decimal(row.get("Cantidad")),
                        quedan=to_decimal(row.get("Quedan")),
                        fecha_movimiento=to_date(row.get("FechaMovimiento")),
                        hora_movimiento=to_time(row.get("HoraMovimiento")),
                        tipo_movimiento=normalize_tipo_mov(row.get("TipoMovimiento")),
                        tipo_movimiento_raw=s(row.get("TipoMovimiento")),
                        cod_proveedor=s(row.get("CodProveedor")),
                        cod_albaran=s(row.get("CodAlbaran")),
                        linea=to_int(row.get("Linea")),
                        cod_factura=s(row.get("CodFactura")),
                        en_partida=to_bool(row.get("EnPartida")),
                        vehiculo=s(row.get("Vehiculo")),
                        kilometraje=to_decimal(row.get("Kilometraje")),
                        observaciones=s(row.get("Observaciones")),
                        raw_data=clean_row(row),
                    )
                    created["tblRecursoAlmacen"] += 1

            report["created"] = created
            report["backup_dir"] = str(backup_dir)

        self.stdout.write("")
        title = "COMMIT INSERT-ONLY" if commit else "DRY-RUN INSERT-ONLY"
        self.stdout.write(self.style.SUCCESS(f"=== SYNC ACCESS PLANIFICACION · {title} ==="))
        self.stdout.write(f"Folder: {folder}")
        self.stdout.write(f"Team: {team}")
        self.stdout.write("Modo: " + ("INSERTA faltantes. No actualiza. No borra." if commit else "NO modifica base de datos"))
        self.stdout.write("")

        for table, data in report["tables"].items():
            self.stdout.write(f"{table}: missing_count={data.get('missing_count', 0)}")

        if commit:
            self.stdout.write("")
            self.stdout.write("=== CREADOS ===")
            for k, v in created.items():
                self.stdout.write(f"{k}: {v}")
            self.stdout.write(f"Backup: {backup_dir}")

        if json_out:
            out = Path(json_out)
            out.parent.mkdir(parents=True, exist_ok=True)
            out.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
            self.stdout.write("")
            self.stdout.write(f"JSON guardado en: {out}")

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("OK: sync finalizado."))

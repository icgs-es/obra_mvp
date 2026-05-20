from pathlib import Path

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Importa almacenes de obra desde tblAlmacen.xlsx hacia AlmacenObra."

    def add_arguments(self, parser):
        parser.add_argument("base_path", type=str)
        parser.add_argument("--team-id", type=int, default=1)
        parser.add_argument("--commit", action="store_true")

    def txt(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def clean_int(self, value):
        if value in (None, ""):
            return None
        try:
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return int(value)
        except Exception:
            return None

    def clean_bool(self, value):
        if isinstance(value, bool):
            return value
        if value in (None, ""):
            return False
        return str(value).strip().lower() in ("true", "1", "-1", "sí", "si", "yes")

    def read_rows(self, path):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if any(v is not None for v in data.values()):
                yield data

    def changed(self, obj, defaults):
        for key, value in defaults.items():
            current = getattr(obj, key)
            if hasattr(current, "pk") and hasattr(value, "pk"):
                if current.pk != value.pk:
                    return True
            elif current != value:
                return True
        return False

    def handle(self, *args, **options):
        base_path = Path(options["base_path"])
        team_id = options["team_id"]
        commit = options["commit"]

        file_path = base_path / "tblAlmacen.xlsx"
        if not file_path.exists():
            raise CommandError(f"No existe: {file_path}")

        Team = apps.get_model("usuarios", "Team")
        ObraPlanificacion = apps.get_model("planificacion_obra", "ObraPlanificacion")
        AlmacenObra = apps.get_model("planificacion_obra", "AlmacenObra")

        team = Team.objects.filter(id=team_id).first()
        if not team:
            raise CommandError(f"No existe Team con id={team_id}")

        modo = "COMMIT REAL" if commit else "SIMULACION"
        self.stdout.write(self.style.WARNING(f"Modo: {modo}"))
        self.stdout.write(f"Team destino: {team.id} · {team.name}")

        rows = list(self.read_rows(file_path))

        obras = {
            o.legacy_cod_obra: o
            for o in ObraPlanificacion.objects.filter(team=team)
        }

        crear = actualizar = sin_cambios = 0
        sin_obra = 0
        sin_id = 0
        sin_nombre = 0

        with transaction.atomic():
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("=== ALMACENES OBRA ==="))

            for row in rows:
                legacy_id = self.txt(row.get("IdAlmacen"))
                nombre = self.txt(row.get("NombreAlmacen"))
                cod_obra = self.clean_int(row.get("CodObra"))

                if not legacy_id:
                    sin_id += 1
                    continue

                if not nombre:
                    sin_nombre += 1
                    continue

                # Regla de saneamiento Access:
                # El almacén A11 corresponde a oficina y debe quedar asociado a OBRA 2 / ALTOVELOO.
                if legacy_id == "A11" and cod_obra not in obras:
                    cod_obra = 2

                obra = obras.get(cod_obra)
                if not obra:
                    sin_obra += 1
                    continue

                defaults = {
                    "obra": obra,
                    "legacy_id_almacen": legacy_id,
                    "nombre": nombre,
                    "ubicacion": self.txt(row.get("Ubicacion")),
                    "descuenta_stock": self.clean_bool(row.get("DescuentaStock")),
                }

                obj = AlmacenObra.objects.filter(
                    team=team,
                    obra=obra,
                    legacy_id_almacen=legacy_id,
                ).first()

                if obj is None:
                    crear += 1
                    self.stdout.write(
                        f"CREAR Almacén {legacy_id} => {obra.nombre} · {nombre}"
                    )
                    AlmacenObra.objects.create(
                        team=team,
                        **defaults,
                    )
                else:
                    if self.changed(obj, defaults):
                        actualizar += 1
                        self.stdout.write(
                            f"ACTUALIZAR Almacén {legacy_id} => {obra.nombre} · {nombre}"
                        )
                        for k, v in defaults.items():
                            setattr(obj, k, v)
                        obj.save()
                    else:
                        sin_cambios += 1

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== RESUMEN ALMACENES OBRA ==="))
        self.stdout.write(f"Filas leídas: {len(rows)}")
        self.stdout.write(f"Crear: {crear}")
        self.stdout.write(f"Actualizar: {actualizar}")
        self.stdout.write(f"Sin cambios: {sin_cambios}")
        self.stdout.write(f"Sin obra: {sin_obra}")
        self.stdout.write(f"Sin IdAlmacen: {sin_id}")
        self.stdout.write(f"Sin NombreAlmacen: {sin_nombre}")

        if not commit:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("SIMULACION: no se ha guardado nada."))
        else:
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("IMPORTACION APLICADA."))

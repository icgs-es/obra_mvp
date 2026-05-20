from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Importa movimientos de almacén desde tblRecursoAlmacen.xlsx."

    def add_arguments(self, parser):
        parser.add_argument("base_path", type=str)
        parser.add_argument("--team-id", type=int, default=1)
        parser.add_argument("--commit", action="store_true")

    def txt(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def clean_int(self, value):
        if value in (None, ""):
            return None
        try:
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return int(value)
        except Exception:
            return None

    def clean_decimal(self, value):
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return None

    def clean_date(self, value):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None

    def clean_time(self, value):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.time()
        if isinstance(value, time):
            return value
        return None

    def clean_bool(self, value):
        if isinstance(value, bool):
            return value
        if value in (None, ""):
            return False
        return str(value).strip().lower() in ("true", "1", "-1", "sí", "si", "yes")

    def normalize_tipo(self, value):
        raw = self.txt(value).upper()
        if raw == "ENTRADA":
            return "ENTRADA"
        if raw == "SALIDA":
            return "SALIDA"
        if raw == "CONTROL STOCK":
            return "CONTROL_STOCK"
        if raw == "ROTURA":
            return "ROTURA"
        return "OTRO"

    def json_safe(self, data):
        safe = {}
        for k, v in data.items():
            if isinstance(v, (datetime, date, time)):
                safe[k] = v.isoformat()
            elif isinstance(v, Decimal):
                safe[k] = str(v)
            else:
                safe[k] = v
        return safe

    def read_rows(self, path):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if any(v is not None for v in data.values()):
                yield data

    def changed(self, obj, defaults):
        for key, value in defaults.items():
            current = getattr(obj, key)
            if hasattr(current, "pk") and hasattr(value, "pk"):
                if current.pk != value.pk:
                    return True
            elif current != value:
                return True
        return False

    def handle(self, *args, **options):
        base_path = Path(options["base_path"])
        team_id = options["team_id"]
        commit = options["commit"]

        file_path = base_path / "tblRecursoAlmacen.xlsx"
        if not file_path.exists():
            raise CommandError(f"No existe: {file_path}")

        Team = apps.get_model("usuarios", "Team")
        ObraPlanificacion = apps.get_model("planificacion_obra", "ObraPlanificacion")
        UnidadObra = apps.get_model("planificacion_obra", "UnidadObra")
        AlmacenObra = apps.get_model("planificacion_obra", "AlmacenObra")
        RecursoCatalogo = apps.get_model("planificacion_obra", "RecursoCatalogo")
        EmpleadoObra = apps.get_model("planificacion_obra", "EmpleadoObra")
        PartidaCatalogo = apps.get_model("planificacion_obra", "PartidaCatalogo")
        RecursoAlmacenMovimiento = apps.get_model("planificacion_obra", "RecursoAlmacenMovimiento")

        team = Team.objects.filter(id=team_id).first()
        if not team:
            raise CommandError(f"No existe Team con id={team_id}")

        modo = "COMMIT REAL" if commit else "SIMULACION"
        self.stdout.write(self.style.WARNING(f"Modo: {modo}"))
        self.stdout.write(f"Team destino: {team.id} · {team.name}")

        rows = list(self.read_rows(file_path))

        obras = {
            o.legacy_cod_obra: o
            for o in ObraPlanificacion.objects.filter(team=team)
        }

        almacenes = {
            (a.obra.legacy_cod_obra, a.legacy_id_almacen): a
            for a in AlmacenObra.objects.select_related("obra").filter(team=team)
        }

        recursos = {
            r.legacy_id: r
            for r in RecursoCatalogo.objects.filter(team=team)
        }

        empleados = {
            e.legacy_id: e
            for e in EmpleadoObra.objects.filter(team=team)
        }

        unidades = {
            (
                u.legacy_cod_obra,
                u.legacy_cod_fase,
                self.txt(u.legacy_cod_vivienda),
            ): u
            for u in UnidadObra.objects.filter(team=team)
        }

        partidas = {
            p.codigo: p
            for p in PartidaCatalogo.objects.filter(team=team)
        }

        crear = actualizar = sin_cambios = 0
        sin_id = 0
        sin_obra = 0
        sin_almacen = 0
        sin_recurso = 0
        sin_unidad_con_datos = 0
        sin_partida_exacta = 0

        with transaction.atomic():
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("=== MOVIMIENTOS ALMACEN ==="))

            for row in rows:
                legacy_id_movimiento = self.clean_int(row.get("IdRecurso"))
                if legacy_id_movimiento is None:
                    sin_id += 1
                    continue

                legacy_id_almacen = self.txt(row.get("IdAlmacen"))
                legacy_cod_recurso = self.clean_int(row.get("CodRecurso"))
                legacy_cod_obra = self.clean_int(row.get("CodObra"))
                legacy_cod_fase = self.clean_int(row.get("CodFase"))
                legacy_cod_vivienda = self.txt(row.get("Vivienda"))
                legacy_cod_personal = self.clean_int(row.get("CodPersonal"))
                legacy_partida = self.txt(row.get("Partida"))

                obra = obras.get(legacy_cod_obra)
                if not obra:
                    sin_obra += 1

                almacen = almacenes.get((legacy_cod_obra, legacy_id_almacen))
                if not almacen:
                    sin_almacen += 1

                recurso = recursos.get(legacy_cod_recurso)
                if not recurso:
                    sin_recurso += 1

                empleado = empleados.get(legacy_cod_personal) if legacy_cod_personal is not None else None

                unidad_obra = None
                if legacy_cod_fase not in (None, 0) or legacy_cod_vivienda not in ("", "0"):
                    unidad_obra = unidades.get((legacy_cod_obra, legacy_cod_fase, legacy_cod_vivienda))
                    if not unidad_obra:
                        sin_unidad_con_datos += 1

                partida = None
                if legacy_partida:
                    partida = partidas.get(legacy_partida)
                    if not partida:
                        sin_partida_exacta += 1

                tipo_raw = self.txt(row.get("TipoMovimiento"))

                defaults = {
                    "almacen": almacen,
                    "recurso": recurso,
                    "obra": obra,
                    "unidad_obra": unidad_obra,
                    "empleado": empleado,
                    "partida": partida,

                    "legacy_id_almacen": legacy_id_almacen,
                    "legacy_cod_recurso": legacy_cod_recurso,
                    "legacy_cod_obra": legacy_cod_obra,
                    "legacy_cod_fase": legacy_cod_fase,
                    "legacy_cod_vivienda": legacy_cod_vivienda,
                    "legacy_planta": self.txt(row.get("Planta")),
                    "legacy_capitulo": self.txt(row.get("Capitulo")),
                    "legacy_partida": legacy_partida,
                    "legacy_cod_personal": legacy_cod_personal,

                    "unidad": self.txt(row.get("Unidad")),
                    "cantidad": self.clean_decimal(row.get("Cantidad")),
                    "quedan": self.clean_decimal(row.get("Quedan")),

                    "fecha_movimiento": self.clean_date(row.get("FechaMovimiento")),
                    "hora_movimiento": self.clean_time(row.get("HoraMovimiento")),

                    "tipo_movimiento": self.normalize_tipo(tipo_raw),
                    "tipo_movimiento_raw": tipo_raw,

                    "cod_proveedor": self.txt(row.get("CodProveedor")),
                    "cod_albaran": self.txt(row.get("CodAlbaran")),
                    "linea": self.clean_int(row.get("Linea")),
                    "cod_factura": self.txt(row.get("CodFactura")),

                    "en_partida": self.clean_bool(row.get("EnPartida")),
                    "vehiculo": self.txt(row.get("Vehiculo")),
                    "kilometraje": self.clean_decimal(row.get("Kilometraje")),

                    "observaciones": self.txt(row.get("Observaciones")),
                    "raw_data": self.json_safe(row),
                }

                obj = RecursoAlmacenMovimiento.objects.filter(
                    team=team,
                    legacy_id_movimiento=legacy_id_movimiento,
                ).first()

                if obj is None:
                    crear += 1
                    if crear <= 40:
                        self.stdout.write(
                            f"CREAR Mov {legacy_id_movimiento} => {legacy_id_almacen} · {legacy_cod_recurso} · {tipo_raw}"
                        )
                    RecursoAlmacenMovimiento.objects.create(
                        team=team,
                        legacy_id_movimiento=legacy_id_movimiento,
                        **defaults,
                    )
                else:
                    if self.changed(obj, defaults):
                        actualizar += 1
                        if actualizar <= 40:
                            self.stdout.write(f"ACTUALIZAR Mov {legacy_id_movimiento}")
                        for k, v in defaults.items():
                            setattr(obj, k, v)
                        obj.save()
                    else:
                        sin_cambios += 1

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== RESUMEN MOVIMIENTOS ALMACEN ==="))
        self.stdout.write(f"Filas leídas: {len(rows)}")
        self.stdout.write(f"Crear: {crear}")
        self.stdout.write(f"Actualizar: {actualizar}")
        self.stdout.write(f"Sin cambios: {sin_cambios}")
        self.stdout.write(f"Sin IdRecurso/movimiento: {sin_id}")
        self.stdout.write(f"Sin obra enlazada: {sin_obra}")
        self.stdout.write(f"Sin almacén enlazado: {sin_almacen}")
        self.stdout.write(f"Sin recurso catálogo enlazado: {sin_recurso}")
        self.stdout.write(f"Sin unidad cuando hay datos: {sin_unidad_con_datos}")
        self.stdout.write(f"Sin partida exacta cuando hay datos: {sin_partida_exacta}")

        if not commit:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("SIMULACION: no se ha guardado nada."))
        else:
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("IMPORTACION APLICADA."))

from pathlib import Path

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Importa unidades de obra desde tblFaseViviendas + tblObraFases."

    def add_arguments(self, parser):
        parser.add_argument("base_path", type=str)
        parser.add_argument("--team-id", type=int, default=1)
        parser.add_argument("--commit", action="store_true")

    def clean_code(self, value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()

    def clean_text(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def read_rows(self, path):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if any(v is not None for v in data.values()):
                yield data

    def handle(self, *args, **options):
        base_path = Path(options["base_path"])
        team_id = options["team_id"]
        commit = options["commit"]

        viviendas_file = base_path / "tblFaseViviendas.xlsx"
        fases_file = base_path / "tblObraFases.xlsx"

        if not viviendas_file.exists():
            raise CommandError(f"No existe: {viviendas_file}")
        if not fases_file.exists():
            raise CommandError(f"No existe: {fases_file}")

        Team = apps.get_model("usuarios", "Team")
        Obra = apps.get_model("core", "Obra")
        UnidadObra = apps.get_model("planificacion_obra", "UnidadObra")

        team = Team.objects.filter(id=team_id).first()
        if not team:
            raise CommandError(f"No existe Team con id={team_id}")

        fases = {}
        for r in self.read_rows(fases_file):
            cod_obra = self.clean_code(r.get("CodObra"))
            cod_fase = self.clean_code(r.get("CodFase"))
            nombre_fase = self.clean_text(r.get("NombreFase"))

            if cod_obra and cod_fase:
                fases[(cod_obra, cod_fase)] = nombre_fase or f"FASE {cod_fase}"

        modo = "COMMIT REAL" if commit else "SIMULACION"
        self.stdout.write(self.style.WARNING(f"Modo: {modo}"))
        self.stdout.write(f"Team destino: {team.id} · {team.name}")

        crear = 0
        actualizar = 0
        sin_cambios = 0
        sin_obra = 0
        incompletas = 0

        rows = list(self.read_rows(viviendas_file))

        with transaction.atomic():
            for r in rows:
                cod_obra = self.clean_code(r.get("CodObra"))
                cod_fase = self.clean_code(r.get("CodFase"))
                cod_vivienda = self.clean_code(r.get("CodVivienda"))
                planta = self.clean_text(r.get("Nivel"))

                if not cod_obra or not cod_fase or not cod_vivienda:
                    incompletas += 1
                    continue

                obra = Obra.objects.filter(codigo=cod_obra).first()
                if not obra:
                    sin_obra += 1
                    self.stdout.write(self.style.ERROR(
                        f"SIN OBRA: CodObra={cod_obra} CodFase={cod_fase} CodVivienda={cod_vivienda}"
                    ))
                    continue

                edificio = fases.get((cod_obra, cod_fase), f"FASE {cod_fase}")
                vivienda = cod_vivienda

                obj = UnidadObra.objects.filter(
                    team=team,
                    obra=obra,
                    edificio=edificio,
                    vivienda=vivienda,
                    planta=planta,
                ).first()

                defaults = {
                    "legacy_cod_obra": int(cod_obra) if cod_obra.isdigit() else None,
                    "legacy_cod_fase": int(cod_fase) if cod_fase.isdigit() else None,
                    "legacy_cod_vivienda": cod_vivienda,
                }

                if obj is None:
                    crear += 1
                    if crear <= 30:
                        self.stdout.write(
                            f"CREAR Obra {cod_obra} · {edificio} · Vivienda {vivienda} · Planta {planta or '-'}"
                        )

                    if commit:
                        UnidadObra.objects.create(
                            team=team,
                            obra=obra,
                            edificio=edificio,
                            vivienda=vivienda,
                            planta=planta,
                            **defaults,
                        )
                else:
                    changed = (
                        obj.legacy_cod_obra != defaults["legacy_cod_obra"]
                        or obj.legacy_cod_fase != defaults["legacy_cod_fase"]
                        or obj.legacy_cod_vivienda != defaults["legacy_cod_vivienda"]
                    )

                    if changed:
                        actualizar += 1
                        if actualizar <= 30:
                            self.stdout.write(
                                f"ACTUALIZAR Obra {cod_obra} · {edificio} · Vivienda {vivienda} · Planta {planta or '-'}"
                            )

                        if commit:
                            obj.legacy_cod_obra = defaults["legacy_cod_obra"]
                            obj.legacy_cod_fase = defaults["legacy_cod_fase"]
                            obj.legacy_cod_vivienda = defaults["legacy_cod_vivienda"]
                            obj.save(update_fields=[
                                "legacy_cod_obra",
                                "legacy_cod_fase",
                                "legacy_cod_vivienda",
                                "actualizado_en",
                            ])
                    else:
                        sin_cambios += 1

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== RESUMEN UNIDADES ==="))
        self.stdout.write(f"Filas leídas: {len(rows)}")
        self.stdout.write(f"Crear: {crear}")
        self.stdout.write(f"Actualizar: {actualizar}")
        self.stdout.write(f"Sin cambios: {sin_cambios}")
        self.stdout.write(f"Incompletas: {incompletas}")
        self.stdout.write(f"Sin obra: {sin_obra}")

        if not commit:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("SIMULACION: no se ha guardado nada."))
        else:
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("IMPORTACION APLICADA."))

from pathlib import Path
import re

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Importa catálogos base de Access: tblObras y tblCapitulo."

    def add_arguments(self, parser):
        parser.add_argument(
            "base_path",
            type=str,
            help="Carpeta donde están los Excel exportados desde Access.",
        )
        parser.add_argument(
            "--commit",
            action="store_true",
            help="Aplica cambios reales. Sin este flag solo simula.",
        )

    def read_rows(self, path):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if any(v is not None for v in data.values()):
                yield data

    def clean_code(self, value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()

    def clean_text(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def chapter_order(self, codigo):
        match = re.search(r"\d+", codigo or "")
        if not match:
            return 0
        return int(match.group(0))

    def handle(self, *args, **options):
        base_path = Path(options["base_path"])
        commit = options["commit"]

        obras_file = base_path / "tblObras.xlsx"
        caps_file = base_path / "tblCapitulo.xlsx"

        if not obras_file.exists():
            raise CommandError(f"No existe: {obras_file}")
        if not caps_file.exists():
            raise CommandError(f"No existe: {caps_file}")

        Obra = apps.get_model("core", "Obra")
        Capitulo = apps.get_model("core", "Capitulo")

        modo = "COMMIT REAL" if commit else "SIMULACION"
        self.stdout.write(self.style.WARNING(f"Modo: {modo}"))

        obras_rows = list(self.read_rows(obras_file))
        caps_rows = list(self.read_rows(caps_file))

        obras_create = 0
        obras_update = 0
        obras_ok = 0

        caps_create = 0
        caps_update = 0
        caps_ok = 0
        caps_sin_obra = 0

        access_obra_codes = set()

        with transaction.atomic():
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("=== OBRAS ==="))

            for r in obras_rows:
                codigo = self.clean_code(r.get("CodObra"))
                nombre = self.clean_text(r.get("NombreObra"))

                if not codigo or not nombre:
                    continue

                access_obra_codes.add(codigo)
                obj = Obra.objects.filter(codigo=codigo).first()

                if obj is None:
                    obras_create += 1
                    self.stdout.write(f"CREAR Obra {codigo} => {nombre}")
                    if commit:
                        Obra.objects.create(codigo=codigo, nombre=nombre)
                else:
                    if obj.nombre != nombre:
                        obras_update += 1
                        self.stdout.write(f"ACTUALIZAR Obra {codigo}: {obj.nombre} => {nombre}")
                        if commit:
                            obj.nombre = nombre
                            obj.save(update_fields=["nombre", "updated_at"])
                    else:
                        obras_ok += 1

            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("=== CAPITULOS ==="))

            for r in caps_rows:
                cod_obra = self.clean_code(r.get("CodObra"))
                codigo = self.clean_code(r.get("Codigo"))
                nombre = self.clean_text(r.get("NombreCapitulo"))

                if not cod_obra or not codigo or not nombre:
                    continue

                obra = Obra.objects.filter(codigo=cod_obra).first()

                if obra is None:
                    if not commit and cod_obra in access_obra_codes:
                        caps_create += 1
                        self.stdout.write(f"CREAR Capitulo Obra {cod_obra} · {codigo} => {nombre}")
                        continue

                    caps_sin_obra += 1
                    self.stdout.write(self.style.ERROR(f"SIN OBRA: CodObra={cod_obra} Capitulo={codigo} {nombre}"))
                    continue

                obj = Capitulo.objects.filter(obra=obra, codigo=codigo).first()
                orden = self.chapter_order(codigo)

                if obj is None:
                    caps_create += 1
                    self.stdout.write(f"CREAR Capitulo Obra {cod_obra} · {codigo} => {nombre}")
                    if commit:
                        Capitulo.objects.create(
                            obra=obra,
                            codigo=codigo,
                            nombre=nombre,
                            orden=orden,
                            presupuesto_plan=0,
                        )
                else:
                    changed = obj.nombre != nombre or obj.orden != orden

                    if changed:
                        caps_update += 1
                        self.stdout.write(
                            f"ACTUALIZAR Capitulo Obra {cod_obra} · {codigo}: {obj.nombre} => {nombre}"
                        )
                        if commit:
                            obj.nombre = nombre
                            obj.orden = orden
                            obj.save(update_fields=["nombre", "orden", "updated_at"])
                    else:
                        caps_ok += 1

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== RESUMEN ==="))
        self.stdout.write(f"Obras crear: {obras_create}")
        self.stdout.write(f"Obras actualizar: {obras_update}")
        self.stdout.write(f"Obras sin cambios: {obras_ok}")
        self.stdout.write(f"Capítulos crear: {caps_create}")
        self.stdout.write(f"Capítulos actualizar: {caps_update}")
        self.stdout.write(f"Capítulos sin cambios: {caps_ok}")
        self.stdout.write(f"Capítulos sin obra: {caps_sin_obra}")

        if not commit:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("SIMULACION: no se ha guardado nada."))
        else:
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("IMPORTACION APLICADA."))

from pathlib import Path
from decimal import Decimal, InvalidOperation

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Importa tblPartida.xlsx desde Access hacia planificacion_obra.PartidaObra."

    def add_arguments(self, parser):
        parser.add_argument("base_path", type=str)
        parser.add_argument("--team-id", type=int, default=1)
        parser.add_argument("--commit", action="store_true")

    def clean_code(self, value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()

    def clean_text(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def clean_decimal(self, value):
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None

    def read_rows(self, path):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if any(v is not None for v in data.values()):
                yield data

    def handle(self, *args, **options):
        base_path = Path(options["base_path"])
        team_id = options["team_id"]
        commit = options["commit"]

        partidas_file = base_path / "tblPartida.xlsx"

        if not partidas_file.exists():
            raise CommandError(f"No existe: {partidas_file}")

        Team = apps.get_model("usuarios", "Team")
        Obra = apps.get_model("core", "Obra")
        Capitulo = apps.get_model("core", "Capitulo")
        PartidaObra = apps.get_model("planificacion_obra", "PartidaObra")

        team = Team.objects.filter(id=team_id).first()
        if not team:
            raise CommandError(f"No existe Team con id={team_id}")

        modo = "COMMIT REAL" if commit else "SIMULACION"
        self.stdout.write(self.style.WARNING(f"Modo: {modo}"))
        self.stdout.write(f"Team destino: {team.id} · {team.name}")

        crear = 0
        actualizar = 0
        sin_cambios = 0
        sin_obra = 0
        sin_capitulo = 0
        incompletas = 0

        rows = list(self.read_rows(partidas_file))

        with transaction.atomic():
            for r in rows:
                cod_obra = self.clean_code(r.get("CodObra"))
                cod_capitulo = self.clean_code(r.get("CodCapitulo"))
                cod_partida = self.clean_code(r.get("CodPartida"))
                nombre = self.clean_text(r.get("NombrePartida"))

                tipo_partida = self.clean_text(r.get("TipoPartida"))
                unidad = self.clean_text(r.get("Unidad"))
                dias_material = self.clean_decimal(r.get("DiasMaterial"))

                if not cod_obra or not cod_capitulo or not cod_partida or not nombre:
                    incompletas += 1
                    continue

                obra = Obra.objects.filter(codigo=cod_obra).first()
                if not obra:
                    sin_obra += 1
                    self.stdout.write(self.style.ERROR(
                        f"SIN OBRA: CodObra={cod_obra} Partida={cod_partida} {nombre}"
                    ))
                    continue

                capitulo = Capitulo.objects.filter(
                    obra=obra,
                    codigo=cod_capitulo,
                ).first()

                if not capitulo:
                    sin_capitulo += 1
                    self.stdout.write(self.style.ERROR(
                        f"SIN CAPITULO: Obra={cod_obra} Capitulo={cod_capitulo} Partida={cod_partida} {nombre}"
                    ))
                    continue

                obj = PartidaObra.objects.filter(
                    team=team,
                    obra=obra,
                    capitulo=capitulo,
                    codigo=cod_partida,
                ).first()

                if obj is None:
                    crear += 1
                    if crear <= 20:
                        self.stdout.write(f"CREAR {cod_obra} · {cod_capitulo} · {cod_partida} => {nombre}")

                    if commit:
                        PartidaObra.objects.create(
                            team=team,
                            obra=obra,
                            capitulo=capitulo,
                            codigo=cod_partida,
                            nombre=nombre,
                            tipo_partida=tipo_partida,
                            unidad=unidad,
                            dias_material=dias_material,
                        )
                else:
                    changed = (
                        obj.nombre != nombre
                        or obj.tipo_partida != tipo_partida
                        or obj.unidad != unidad
                        or obj.dias_material != dias_material
                    )

                    if changed:
                        actualizar += 1
                        if actualizar <= 20:
                            self.stdout.write(f"ACTUALIZAR {cod_obra} · {cod_capitulo} · {cod_partida} => {nombre}")

                        if commit:
                            obj.nombre = nombre
                            obj.tipo_partida = tipo_partida
                            obj.unidad = unidad
                            obj.dias_material = dias_material
                            obj.save(update_fields=[
                                "nombre",
                                "tipo_partida",
                                "unidad",
                                "dias_material",
                                "actualizado_en",
                            ])
                    else:
                        sin_cambios += 1

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== RESUMEN PARTIDAS ==="))
        self.stdout.write(f"Filas leídas: {len(rows)}")
        self.stdout.write(f"Crear: {crear}")
        self.stdout.write(f"Actualizar: {actualizar}")
        self.stdout.write(f"Sin cambios: {sin_cambios}")
        self.stdout.write(f"Incompletas: {incompletas}")
        self.stdout.write(f"Sin obra: {sin_obra}")
        self.stdout.write(f"Sin capítulo: {sin_capitulo}")

        if not commit:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("SIMULACION: no se ha guardado nada."))
        else:
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("IMPORTACION APLICADA."))

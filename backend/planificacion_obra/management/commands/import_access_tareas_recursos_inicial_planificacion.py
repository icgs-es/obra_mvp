from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Importa recursos previstos desde tblTareasRecursosInicial.xlsx."

    def add_arguments(self, parser):
        parser.add_argument("base_path", type=str)
        parser.add_argument("--team-id", type=int, default=1)
        parser.add_argument("--commit", action="store_true")

    def txt(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def clean_int(self, value):
        if value in (None, ""):
            return None
        try:
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return int(value)
        except Exception:
            return None

    def clean_decimal(self, value):
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return None

    def clean_date(self, value):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None

    def clean_bool(self, value):
        if isinstance(value, bool):
            return value
        if value in (None, ""):
            return False
        return str(value).strip().lower() in ("true", "1", "-1", "sí", "si", "yes")

    def json_safe(self, data):
        safe = {}
        for k, v in data.items():
            if isinstance(v, (datetime, date, time)):
                safe[k] = v.isoformat()
            elif isinstance(v, Decimal):
                safe[k] = str(v)
            else:
                safe[k] = v
        return safe

    def read_rows(self, path):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for excel_row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))
            if any(v is not None for v in data.values()):
                yield excel_row_number, data

    def changed(self, obj, defaults):
        for key, value in defaults.items():
            current = getattr(obj, key)
            if hasattr(current, "pk") and hasattr(value, "pk"):
                if current.pk != value.pk:
                    return True
            elif current != value:
                return True
        return False

    def handle(self, *args, **options):
        base_path = Path(options["base_path"])
        team_id = options["team_id"]
        commit = options["commit"]

        file_path = base_path / "tblTareasRecursosInicial.xlsx"
        if not file_path.exists():
            raise CommandError(f"No existe: {file_path}")

        Team = apps.get_model("usuarios", "Team")
        TareaObra = apps.get_model("planificacion_obra", "TareaObra")
        UnidadObra = apps.get_model("planificacion_obra", "UnidadObra")
        PartidaCatalogo = apps.get_model("planificacion_obra", "PartidaCatalogo")
        RecursoCatalogo = apps.get_model("planificacion_obra", "RecursoCatalogo")
        TareaRecursoPrevisto = apps.get_model("planificacion_obra", "TareaRecursoPrevisto")

        team = Team.objects.filter(id=team_id).first()
        if not team:
            raise CommandError(f"No existe Team con id={team_id}")

        modo = "COMMIT REAL" if commit else "SIMULACION"
        self.stdout.write(self.style.WARNING(f"Modo: {modo}"))
        self.stdout.write(f"Team destino: {team.id} · {team.name}")
        self.stdout.write("TareaObra se enlaza SIN usar Orden: CodObra + CodFase + CodVivienda + Planta + CodPartida.")
        self.stdout.write("Orden se conserva como legacy_orden_recurso.")

        rows = list(self.read_rows(file_path))

        tareas = {
            (
                t.legacy_cod_obra,
                t.legacy_cod_fase,
                self.txt(t.legacy_cod_vivienda),
                self.txt(t.legacy_planta),
                self.txt(t.legacy_partida),
            ): t
            for t in TareaObra.objects.filter(team=team)
        }

        unidades = {
            (
                u.legacy_cod_obra,
                u.legacy_cod_fase,
                self.txt(u.legacy_cod_vivienda),
            ): u
            for u in UnidadObra.objects.filter(team=team)
        }

        partidas = {
            p.codigo: p
            for p in PartidaCatalogo.objects.filter(team=team)
        }

        recursos = {
            r.legacy_id: r
            for r in RecursoCatalogo.objects.filter(team=team)
        }

        crear = actualizar = sin_cambios = 0
        sin_tarea = sin_unidad = sin_partida = sin_recurso = 0

        with transaction.atomic():
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("=== RECURSOS PREVISTOS DE TAREAS ==="))

            for legacy_row_number, row in rows:
                cod_obra = self.clean_int(row.get("CodObra"))
                cod_fase = self.clean_int(row.get("CodFase"))
                cod_vivienda = self.txt(row.get("CodVivienda"))
                planta = self.txt(row.get("Planta"))
                cod_partida = self.txt(row.get("CodPartida"))
                id_recurso = self.clean_int(row.get("IdRecurso"))
                orden_recurso = self.clean_int(row.get("Orden"))

                tarea_key = (
                    cod_obra,
                    cod_fase,
                    cod_vivienda,
                    planta,
                    cod_partida,
                )

                tarea_obra = tareas.get(tarea_key)
                if not tarea_obra:
                    sin_tarea += 1

                unidad_obra = unidades.get((cod_obra, cod_fase, cod_vivienda))
                if not unidad_obra:
                    sin_unidad += 1

                partida = partidas.get(cod_partida)
                if not partida:
                    sin_partida += 1

                recurso = recursos.get(id_recurso)
                if not recurso:
                    sin_recurso += 1

                defaults = {
                    "tarea_obra": tarea_obra,
                    "unidad_obra": unidad_obra,
                    "partida": partida,
                    "recurso": recurso,

                    "legacy_cod_obra": cod_obra,
                    "legacy_cod_fase": cod_fase,
                    "legacy_cod_vivienda": cod_vivienda,
                    "legacy_planta": planta,
                    "legacy_cod_partida": cod_partida,
                    "legacy_id_recurso": id_recurso,
                    "legacy_id_recurso_old": self.clean_int(row.get("IdRecurso_old")),
                    "legacy_orden_recurso": orden_recurso,

                    "unidad": self.txt(row.get("Unidad")),
                    "precio_unidad": self.clean_decimal(row.get("PrecioUnidad")),
                    "cantidad": self.clean_decimal(row.get("Cantidad")),
                    "costo_recurso": self.clean_decimal(row.get("CostoRecurso")),

                    "control_suministros": self.clean_bool(row.get("ControlSuministros")),
                    "avisar": self.clean_int(row.get("Avisar")),
                    "fecha_estimada_entrega": self.clean_date(row.get("FechaEstimadaEntrega")),

                    "raw_data": self.json_safe(row),
                }

                obj = TareaRecursoPrevisto.objects.filter(
                    team=team,
                    legacy_row_number=legacy_row_number,
                ).first()

                if obj is None:
                    crear += 1
                    if crear <= 40:
                        self.stdout.write(
                            f"CREAR Previsto fila {legacy_row_number} => {cod_obra}|{cod_fase}|{cod_vivienda}|{planta}|{cod_partida} · rec {id_recurso}"
                        )
                    TareaRecursoPrevisto.objects.create(
                        team=team,
                        legacy_row_number=legacy_row_number,
                        **defaults,
                    )
                else:
                    if self.changed(obj, defaults):
                        actualizar += 1
                        if actualizar <= 40:
                            self.stdout.write(f"ACTUALIZAR Previsto fila {legacy_row_number}")
                        for k, v in defaults.items():
                            setattr(obj, k, v)
                        obj.save()
                    else:
                        sin_cambios += 1

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== RESUMEN RECURSOS PREVISTOS ==="))
        self.stdout.write(f"Filas leídas: {len(rows)}")
        self.stdout.write(f"Crear: {crear}")
        self.stdout.write(f"Actualizar: {actualizar}")
        self.stdout.write(f"Sin cambios: {sin_cambios}")
        self.stdout.write(f"Sin tarea enlazada: {sin_tarea}")
        self.stdout.write(f"Sin unidad enlazada: {sin_unidad}")
        self.stdout.write(f"Sin partida enlazada: {sin_partida}")
        self.stdout.write(f"Sin recurso catálogo enlazado: {sin_recurso}")

        if not commit:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("SIMULACION: no se ha guardado nada."))
        else:
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("IMPORTACION APLICADA."))

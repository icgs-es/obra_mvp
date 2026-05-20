from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Importa catálogos genéricos de planificación: capítulos, partidas y recursos."

    def add_arguments(self, parser):
        parser.add_argument("base_path", type=str)
        parser.add_argument("--team-id", type=int, default=1)
        parser.add_argument("--commit", action="store_true")

    def read_rows(self, path):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if any(v is not None for v in data.values()):
                yield data

    def txt(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def clean_int(self, value):
        if value in (None, ""):
            return None
        try:
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return int(value)
        except (TypeError, ValueError):
            return None

    def clean_decimal(self, value):
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return None

    def clean_bool(self, value):
        if isinstance(value, bool):
            return value
        if value in (None, ""):
            return False
        return str(value).strip().lower() in ("true", "1", "sí", "si", "yes", "-1")

    def json_safe(self, data):
        safe = {}
        for k, v in data.items():
            if isinstance(v, (datetime, date)):
                safe[k] = v.isoformat()
            elif isinstance(v, Decimal):
                safe[k] = str(v)
            else:
                safe[k] = v
        return safe

    def handle(self, *args, **options):
        base_path = Path(options["base_path"])
        team_id = options["team_id"]
        commit = options["commit"]

        capitulos_file = base_path / "tblCapitulo.xlsx"
        partidas_file = base_path / "tblPartida.xlsx"
        recursos_file = base_path / "tblRecursos.xlsx"

        for f in (capitulos_file, partidas_file, recursos_file):
            if not f.exists():
                raise CommandError(f"No existe: {f}")

        Team = apps.get_model("usuarios", "Team")
        CapituloCatalogo = apps.get_model("planificacion_obra", "CapituloCatalogo")
        PartidaCatalogo = apps.get_model("planificacion_obra", "PartidaCatalogo")
        RecursoCatalogo = apps.get_model("planificacion_obra", "RecursoCatalogo")

        team = Team.objects.filter(id=team_id).first()
        if not team:
            raise CommandError(f"No existe Team con id={team_id}")

        modo = "COMMIT REAL" if commit else "SIMULACION"
        self.stdout.write(self.style.WARNING(f"Modo: {modo}"))
        self.stdout.write(f"Team destino: {team.id} · {team.name}")
        self.stdout.write("CodObra será ignorado en capítulos, partidas y recursos.")

        capitulos_rows = list(self.read_rows(capitulos_file))
        partidas_rows = list(self.read_rows(partidas_file))
        recursos_rows = list(self.read_rows(recursos_file))

        cap_crear = cap_actualizar = cap_ok = 0
        part_crear = part_actualizar = part_ok = part_sin_capitulo = 0
        rec_crear = rec_actualizar = rec_ok = rec_sin_id = 0

        with transaction.atomic():
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("=== CAPITULOS GENERICOS ==="))

            for r in capitulos_rows:
                codigo = self.txt(r.get("Codigo"))
                nombre = self.txt(r.get("NombreCapitulo"))

                if not codigo:
                    continue

                obj = CapituloCatalogo.objects.filter(team=team, codigo=codigo).first()

                defaults = {
                    "nombre": nombre or codigo,
                    "orden": self.clean_int(codigo.replace("C", "")) or 0,
                    "raw_data": self.json_safe(r),
                }

                if obj is None:
                    cap_crear += 1
                    self.stdout.write(f"CREAR Capítulo {codigo} => {defaults['nombre']}")
                    CapituloCatalogo.objects.create(team=team, codigo=codigo, **defaults)
                else:
                    changed = any(getattr(obj, k) != v for k, v in defaults.items())
                    if changed:
                        cap_actualizar += 1
                        self.stdout.write(f"ACTUALIZAR Capítulo {codigo} => {defaults['nombre']}")
                        for k, v in defaults.items():
                            setattr(obj, k, v)
                        obj.save()
                    else:
                        cap_ok += 1

            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("=== PARTIDAS GENERICAS ==="))

            for r in partidas_rows:
                cod_capitulo = self.txt(r.get("CodCapitulo"))
                cod_partida = self.txt(r.get("CodPartida"))

                if not cod_capitulo or not cod_partida:
                    continue

                capitulo = CapituloCatalogo.objects.filter(team=team, codigo=cod_capitulo).first()
                if not capitulo:
                    part_sin_capitulo += 1
                    self.stdout.write(self.style.ERROR(f"SIN CAPITULO: {cod_capitulo} · {cod_partida}"))
                    continue

                obj = PartidaCatalogo.objects.filter(
                    team=team,
                    capitulo=capitulo,
                    codigo=cod_partida,
                ).first()

                defaults = {
                    "nombre": self.txt(r.get("NombrePartida")) or cod_partida,
                    "tipo_partida": self.txt(r.get("TipoPartida")),
                    "unidad": self.txt(r.get("Unidad")),
                    "dias_material": self.clean_decimal(r.get("DiasMaterial")),
                    "raw_data": self.json_safe(r),
                }

                if obj is None:
                    part_crear += 1
                    if part_crear <= 25:
                        self.stdout.write(f"CREAR Partida {cod_capitulo} · {cod_partida} => {defaults['nombre']}")
                    PartidaCatalogo.objects.create(
                        team=team,
                        capitulo=capitulo,
                        codigo=cod_partida,
                        **defaults,
                    )
                else:
                    changed = any(getattr(obj, k) != v for k, v in defaults.items())
                    if changed:
                        part_actualizar += 1
                        if part_actualizar <= 25:
                            self.stdout.write(f"ACTUALIZAR Partida {cod_capitulo} · {cod_partida}")
                        for k, v in defaults.items():
                            setattr(obj, k, v)
                        obj.save()
                    else:
                        part_ok += 1

            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("=== RECURSOS GENERICOS ==="))

            for r in recursos_rows:
                legacy_id = self.clean_int(r.get("IdRecurso"))
                if legacy_id is None:
                    rec_sin_id += 1
                    continue

                cod_capitulo = self.txt(r.get("Capitulo"))
                capitulo = None
                if cod_capitulo:
                    capitulo = CapituloCatalogo.objects.filter(team=team, codigo=cod_capitulo).first()

                obj = RecursoCatalogo.objects.filter(team=team, legacy_id=legacy_id).first()

                defaults = {
                    "nombre": self.txt(r.get("NombreRecurso")) or str(legacy_id),
                    "tipo": self.txt(r.get("Tipo")),
                    "unidad": self.txt(r.get("Unidad")),
                    "capitulo": capitulo,
                    "stock": self.clean_decimal(r.get("Stock")),
                    "ultimo_precio_unidad": self.clean_decimal(r.get("UltPrecioUnidad")),
                    "precio_unidad_uso": self.clean_decimal(r.get("PrecioUnidadUso")),
                    "control_stock": self.clean_bool(r.get("ControlStock")),
                    "observaciones": self.txt(r.get("Observaciones")),
                    "raw_data": self.json_safe(r),
                }

                if obj is None:
                    rec_crear += 1
                    if rec_crear <= 25:
                        self.stdout.write(f"CREAR Recurso {legacy_id} => {defaults['nombre']}")
                    RecursoCatalogo.objects.create(
                        team=team,
                        legacy_id=legacy_id,
                        **defaults,
                    )
                else:
                    changed = any(getattr(obj, k) != v for k, v in defaults.items())
                    if changed:
                        rec_actualizar += 1
                        if rec_actualizar <= 25:
                            self.stdout.write(f"ACTUALIZAR Recurso {legacy_id} => {defaults['nombre']}")
                        for k, v in defaults.items():
                            setattr(obj, k, v)
                        obj.save()
                    else:
                        rec_ok += 1

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== RESUMEN CATALOGOS GENERICOS ==="))
        self.stdout.write(f"Capítulos leídos: {len(capitulos_rows)}")
        self.stdout.write(f"Capítulos crear: {cap_crear}")
        self.stdout.write(f"Capítulos actualizar: {cap_actualizar}")
        self.stdout.write(f"Capítulos sin cambios: {cap_ok}")

        self.stdout.write("")
        self.stdout.write(f"Partidas leídas: {len(partidas_rows)}")
        self.stdout.write(f"Partidas crear: {part_crear}")
        self.stdout.write(f"Partidas actualizar: {part_actualizar}")
        self.stdout.write(f"Partidas sin cambios: {part_ok}")
        self.stdout.write(f"Partidas sin capítulo: {part_sin_capitulo}")

        self.stdout.write("")
        self.stdout.write(f"Recursos leídos: {len(recursos_rows)}")
        self.stdout.write(f"Recursos crear: {rec_crear}")
        self.stdout.write(f"Recursos actualizar: {rec_actualizar}")
        self.stdout.write(f"Recursos sin cambios: {rec_ok}")
        self.stdout.write(f"Recursos sin IdRecurso: {rec_sin_id}")

        if not commit:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("SIMULACION: no se ha guardado nada."))
        else:
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("IMPORTACION APLICADA."))

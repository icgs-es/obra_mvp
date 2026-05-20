from __future__ import annotations

import json
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook


def s(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).replace("\ufeff", "").strip()


def to_int(value):
    value = s(value)
    if not value:
        return None
    try:
        return int(Decimal(value.replace(",", ".")))
    except (InvalidOperation, ValueError):
        return None


def to_decimal(value):
    value = s(value)
    if not value:
        return None
    try:
        return Decimal(value.replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def to_date(value):
    if value is None or s(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    raw = s(value).split(".")[0]
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def to_bool(value):
    raw = s(value).lower()
    return raw in ("true", "1", "sí", "si", "yes")


def read_xlsx(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    try:
        headers = [s(v) for v in next(rows)]
    except StopIteration:
        wb.close()
        return

    for raw in rows:
        row = {}
        for idx, header in enumerate(headers):
            if header:
                row[header] = raw[idx] if idx < len(raw) else None
        yield row

    wb.close()


def field_names(model):
    return {f.name for f in model._meta.fields}


def has_field(model, name):
    return name in field_names(model)


def set_if_field(model, data, field, value):
    if has_field(model, field):
        data[field] = value


def get_team(team_name):
    Team = apps.get_model("usuarios", "Team")
    fields = field_names(Team)

    for field in ("nombre", "name", "razon_social"):
        if field in fields:
            obj = Team.objects.filter(**{field: team_name}).first()
            if obj:
                return obj

    for field in ("nombre", "name", "razon_social"):
        if field in fields:
            obj = Team.objects.filter(**{f"{field}__icontains": team_name}).first()
            if obj:
                return obj

    return None


def legacy_key_for_tarea(key):
    return "TAREA|" + "|".join("" if v is None else str(v) for v in key)


def tarea_key(row):
    return (
        to_int(row.get("CodObra")),
        to_int(row.get("CodFase")),
        s(row.get("CodVivienda")),
        s(row.get("Planta")),
        s(row.get("Capitulo")),
        s(row.get("Partida")),
        to_int(row.get("Orden")),
    )


def tarea_filter_from_key(key):
    return {
        "legacy_cod_obra": key[0],
        "legacy_cod_fase": key[1],
        "legacy_cod_vivienda": key[2],
        "legacy_planta": key[3],
        "legacy_capitulo": key[4],
        "legacy_partida": key[5],
        "legacy_orden": key[6],
    }


def find_tarea(TareaObra, team, key):
    legacy_key = legacy_key_for_tarea(key)

    if has_field(TareaObra, "legacy_key"):
        obj = TareaObra.objects.filter(team=team, legacy_key=legacy_key).first()
        if obj:
            return obj

    return TareaObra.objects.filter(team=team, **tarea_filter_from_key(key)).first()


def find_similar_tarea(TareaObra, team, key):
    base = TareaObra.objects.filter(
        team=team,
        legacy_cod_obra=key[0],
        legacy_cod_fase=key[1],
        legacy_cod_vivienda=key[2],
        legacy_planta=key[3],
    )

    obj = base.filter(
        legacy_capitulo=key[4],
        legacy_partida=key[5],
    ).first()

    if obj:
        return obj

    return base.first()


def copy_fk_context_from_similar(TareaObra, data, similar):
    if not similar:
        return

    for f in TareaObra._meta.fields:
        if not getattr(f, "many_to_one", False):
            continue

        name = f.name

        if name in ("team", "capitulo", "partida"):
            continue

        value = getattr(similar, name, None)

        if value is not None:
            data[name] = value


def update_model_instance(obj, data):
    changed = []

    for key, value in data.items():
        if key == "team":
            continue

        if getattr(obj, key, None) != value:
            setattr(obj, key, value)
            changed.append(key)

    if changed:
        obj.save(update_fields=changed)

    return changed


def assign_tarea_fields(TareaObra, data, row):
    mappings_int = {
        "legacy_cod_obra": "CodObra",
        "legacy_cod_fase": "CodFase",
        "legacy_orden": "Orden",
        "porcentaje_completado": "PorcentajeCompletado",
        "personas_a_utilizar": "Personasautilizar",
        "personas_utilizadas": "Personasutilizadas",
    }

    mappings_str = {
        "legacy_cod_vivienda": "CodVivienda",
        "legacy_planta": "Planta",
        "legacy_capitulo": "Capitulo",
        "legacy_partida": "Partida",
        "programacion": "Programacion",
        "tipo_partida": "TipoPartida",
        "unidad": "Unidad",
        "observaciones": "Observaciones",
        "partida_predecesora": "PartidaPredecesora",
        "archivos": "Archivos",
    }

    mappings_decimal = {
        "hora_predeterminada": "HoraPredeterminada",
        "dias": "Dias",
        "horas": "Horas",
        "dias_reales": "DiasReales",
        "horas_reales": "HorasReales",
        "importe_tarea": "ImporteTarea",
        "importe_tarea_real": "ImporteTareaReal",
        "desvio": "Desvio",
        "cantidad": "Cantidad",
        "precio_unidad": "PrecioUnidad",
        "desvio_fi": "DesvioFI",
        "desvio_ff": "DesvioFF",
        "desvio_importes": "DesvioImportes",
        "duracion_real": "DuracionReal",
        "personas_utilizar_estimado": "Personasutilizarestimado",
        "horas_estimadas": "HorasEstimadas",
        "importe_tarea_estimado": "ImporteTareaEstimado",
        "dias_estimados": "DiasEstimados",
    }

    mappings_date = {
        "inicio_tarea": "InicioTarea",
        "fin_tarea": "FinTarea",
        "inicio_real": "InicioReal",
        "fin_real": "FinReal",
        "inicio_estimado": "InicioEstimado",
        "fin_estimado": "FinEstimado",
    }

    mappings_bool = {
        "seleccionar": "Seleccionar",
        "con_incidencias": "ConIncidencias",
    }

    for field, source in mappings_int.items():
        set_if_field(TareaObra, data, field, to_int(row.get(source)))

    for field, source in mappings_str.items():
        set_if_field(TareaObra, data, field, s(row.get(source)))

    for field, source in mappings_decimal.items():
        set_if_field(TareaObra, data, field, to_decimal(row.get(source)))

    for field, source in mappings_date.items():
        set_if_field(TareaObra, data, field, to_date(row.get(source)))

    for field, source in mappings_bool.items():
        set_if_field(TareaObra, data, field, to_bool(row.get(source)))


class Command(BaseCommand):
    help = "Sincroniza tblPartida.xlsx y tblTareas.xlsx desde Access hacia Portal INTASA. Idempotente. Por defecto dry-run."

    def add_arguments(self, parser):
        parser.add_argument("--folder", required=True)
        parser.add_argument("--team", required=True)
        parser.add_argument("--apply", action="store_true")
        parser.add_argument("--sample", type=int, default=10)
        parser.add_argument("--json-out", default="")

    def handle(self, *args, **options):
        folder = Path(options["folder"])
        team_name = options["team"]
        apply = bool(options["apply"])
        sample_limit = max(1, int(options["sample"] or 10))

        if not folder.exists():
            raise CommandError(f"No existe carpeta: {folder}")

        partida_file = folder / "tblPartida.xlsx"
        tareas_file = folder / "tblTareas.xlsx"

        if not partida_file.exists():
            raise CommandError(f"No existe: {partida_file}")

        if not tareas_file.exists():
            raise CommandError(f"No existe: {tareas_file}")

        team = get_team(team_name)

        if not team:
            raise CommandError(f"No encuentro Team: {team_name}")

        CapituloCatalogo = apps.get_model("planificacion_obra", "CapituloCatalogo")
        PartidaCatalogo = apps.get_model("planificacion_obra", "PartidaCatalogo")
        TareaObra = apps.get_model("planificacion_obra", "TareaObra")

        stats = {
            "modo": "APPLY" if apply else "DRY-RUN",
            "partidas_leidas": 0,
            "partidas_creadas": 0,
            "partidas_actualizadas": 0,
            "partidas_sin_cambios": 0,
            "capitulos_creados": 0,
            "tareas_leidas": 0,
            "tareas_creadas": 0,
            "tareas_actualizadas": 0,
            "tareas_sin_cambios": 0,
            "tareas_saltadas_sin_contexto": 0,
            "tareas_saltadas_clave_invalida": 0,
        }

        samples = {
            "partidas_creadas": [],
            "partidas_actualizadas": [],
            "tareas_creadas": [],
            "tareas_actualizadas": [],
            "tareas_saltadas": [],
        }

        def add_sample(name, item):
            if len(samples[name]) < sample_limit:
                samples[name].append(item)

        def run():
            for row in read_xlsx(partida_file):
                stats["partidas_leidas"] += 1

                cap_codigo = s(row.get("CodCapitulo"))
                part_codigo = s(row.get("CodPartida"))

                if not cap_codigo or not part_codigo:
                    continue

                capitulo = CapituloCatalogo.objects.filter(team=team, codigo=cap_codigo).first()

                if not capitulo:
                    stats["capitulos_creados"] += 1
                    if apply:
                        data = {"team": team, "codigo": cap_codigo}
                        set_if_field(CapituloCatalogo, data, "nombre", cap_codigo)
                        capitulo = CapituloCatalogo.objects.create(**data)

                partida = None

                if capitulo:
                    partida = PartidaCatalogo.objects.filter(
                        team=team,
                        capitulo=capitulo,
                        codigo=part_codigo,
                    ).first()

                pdata = {
                    "team": team,
                    "capitulo": capitulo,
                    "codigo": part_codigo,
                }

                set_if_field(PartidaCatalogo, pdata, "nombre", s(row.get("NombrePartida")))
                set_if_field(PartidaCatalogo, pdata, "tipo", s(row.get("TipoPartida")))
                set_if_field(PartidaCatalogo, pdata, "unidad", s(row.get("Unidad")))
                set_if_field(PartidaCatalogo, pdata, "dias_material", to_decimal(row.get("DiasMaterial")))
                set_if_field(PartidaCatalogo, pdata, "legacy_cod_obra", to_int(row.get("CodObra")))

                if partida:
                    changed = []

                    for k, v in pdata.items():
                        if k in ("team", "capitulo", "codigo"):
                            continue
                        if hasattr(partida, k) and getattr(partida, k) != v:
                            changed.append(k)

                    if changed:
                        stats["partidas_actualizadas"] += 1
                        add_sample("partidas_actualizadas", {"capitulo": cap_codigo, "partida": part_codigo, "changed": changed})

                        if apply:
                            update_model_instance(partida, pdata)
                    else:
                        stats["partidas_sin_cambios"] += 1
                else:
                    stats["partidas_creadas"] += 1
                    add_sample("partidas_creadas", {
                        "capitulo": cap_codigo,
                        "partida": part_codigo,
                        "nombre": s(row.get("NombrePartida")),
                    })

                    if apply:
                        PartidaCatalogo.objects.create(**pdata)

            for row in read_xlsx(tareas_file):
                stats["tareas_leidas"] += 1

                key = tarea_key(row)

                if None in (key[0], key[1], key[6]) or not key[3] or not key[4] or not key[5]:
                    stats["tareas_saltadas_clave_invalida"] += 1
                    continue

                existing = find_tarea(TareaObra, team, key)

                capitulo = CapituloCatalogo.objects.filter(team=team, codigo=key[4]).first()
                partida = None

                if capitulo:
                    partida = PartidaCatalogo.objects.filter(team=team, capitulo=capitulo, codigo=key[5]).first()

                if not capitulo or not partida:
                    stats["tareas_saltadas_sin_contexto"] += 1
                    add_sample("tareas_saltadas", {"key": key, "motivo": "No existe capítulo/partida"})
                    continue

                data = {}

                if not existing:
                    similar = find_similar_tarea(TareaObra, team, key)

                    if not similar:
                        stats["tareas_saltadas_sin_contexto"] += 1
                        add_sample("tareas_saltadas", {"key": key, "motivo": "No hay tarea similar para copiar contexto FK"})
                        continue

                    data["team"] = team
                    copy_fk_context_from_similar(TareaObra, data, similar)

                if has_field(TareaObra, "capitulo"):
                    data["capitulo"] = capitulo

                if has_field(TareaObra, "partida"):
                    data["partida"] = partida

                assign_tarea_fields(TareaObra, data, row)

                if has_field(TareaObra, "legacy_key"):
                    data["legacy_key"] = legacy_key_for_tarea(key)

                if existing:
                    changed = []

                    for k, v in data.items():
                        if k == "team":
                            continue
                        if hasattr(existing, k) and getattr(existing, k) != v:
                            changed.append(k)

                    if changed:
                        stats["tareas_actualizadas"] += 1
                        add_sample("tareas_actualizadas", {"key": key, "changed": changed[:12]})

                        if apply:
                            update_model_instance(existing, data)
                    else:
                        stats["tareas_sin_cambios"] += 1
                else:
                    stats["tareas_creadas"] += 1
                    add_sample("tareas_creadas", {
                        "key": key,
                        "legacy_key": data.get("legacy_key", ""),
                        "inicio_tarea": s(row.get("InicioTarea")),
                        "fin_tarea": s(row.get("FinTarea")),
                        "inicio_real": s(row.get("InicioReal")),
                        "fin_real": s(row.get("FinReal")),
                    })

                    if apply:
                        TareaObra.objects.create(**data)

        if apply:
            with transaction.atomic():
                run()
        else:
            run()

        result = {
            "folder": str(folder),
            "team": str(team),
            "apply": apply,
            "stats": stats,
            "samples": samples,
        }

        self.stdout.write("")
        self.stdout.write("=== SYNC ACCESS PARTIDAS + TAREAS ===")
        self.stdout.write(f"Folder: {folder}")
        self.stdout.write(f"Team: {team}")
        self.stdout.write(f"Modo: {'APPLY - modifica BD' if apply else 'DRY-RUN - no modifica BD'}")
        self.stdout.write("")

        for k, v in stats.items():
            self.stdout.write(f"{k}: {v}")

        for name, items in samples.items():
            if items:
                self.stdout.write("")
                self.stdout.write(name + ":")
                for item in items:
                    self.stdout.write(json.dumps(item, ensure_ascii=False, default=str))

        if options["json_out"]:
            out = Path(options["json_out"])
            out.parent.mkdir(parents=True, exist_ok=True)
            out.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
            self.stdout.write("")
            self.stdout.write(f"JSON guardado en: {out}")

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("OK: sync finalizado."))

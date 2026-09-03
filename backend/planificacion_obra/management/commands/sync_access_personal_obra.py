from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.apps import apps
from django.core.management import call_command
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook


LABOR_TYPES = {"M.O. ADM.", "PER. CONT.", "M.O. CONT."}


def s(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).replace("\ufeff", "").strip()


def to_int(value):
    value = s(value)
    if not value:
        return None
    try:
        return int(Decimal(value.replace(",", ".")))
    except (InvalidOperation, ValueError):
        return None


def to_decimal(value):
    value = s(value)
    if not value:
        return None
    try:
        return Decimal(value.replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def to_bool(value):
    if isinstance(value, bool):
        return value
    return s(value).lower() in {"true", "1", "sí", "si", "yes", "y", "-1"}


def to_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    txt = s(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(txt, fmt).date()
        except ValueError:
            pass
    return None


def clean_json(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def clean_row(row):
    return {str(k): clean_json(v) for k, v in row.items()}


def read_xlsx(path: Path):
    if not path.exists():
        raise CommandError(f"No existe el archivo: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    try:
        headers = [s(v) for v in next(rows)]
    except StopIteration:
        wb.close()
        return

    for idx, row in enumerate(rows, start=2):
        data = dict(zip(headers, row))
        data["_xlsx_row_number"] = idx
        yield data

    wb.close()


def map_tipo(data):
    tipo = s(data.get("Tipo")).upper()
    subcontrata = to_bool(data.get("SUBCONTRATA"))
    if subcontrata or "CONT" in tipo:
        return "CONTRATADO"
    return "ADMINISTRADA"


def map_categoria(data):
    cargo = s(data.get("CARGO")).upper()
    if "JEFE" in cargo:
        return "JEFE_OBRA"
    if "ENCARG" in cargo:
        return "ENCARGADO"
    if "OFICIAL 1" in cargo or "OFICIAL 1ª" in cargo:
        return "OFICIAL_1"
    if "OFICIAL 2" in cargo or "OFICIAL 2ª" in cargo:
        return "OFICIAL_2"
    if "PEON" in cargo or "PEÓN" in cargo:
        return "PEON"
    return "OTRO"


def map_situacion(data):
    estado = s(data.get("ESTADO")).upper()
    if estado == "BAJA":
        return "BAJA"
    if estado == "VACACIONES":
        return "VACACIONES"
    return "ACTIVO"


def build_empleado_payload(data):
    legacy_id = to_int(data.get("IdPersonal"))
    return {
        "nombre": s(data.get("NOMBRE")) or f"Personal {legacy_id}",
        "tipo": map_tipo(data),
        "categoria": map_categoria(data),
        "situacion": map_situacion(data),
        "fecha_alta": to_date(data.get("Altadesde")),
        "fecha_baja": to_date(data.get("Bajadesde")),
        "precio_hora": to_decimal(data.get("PRECIO_HORA")),
        "empresa_origen": s(data.get("EMPRESA")) or s(data.get("EMPRESA_OLD")),
        "observaciones": (
            f"CARGO={s(data.get('CARGO'))}; "
            f"OFICIO={s(data.get('OFICIO'))}; "
            f"Equipo={s(data.get('Equipo'))}; "
            f"CodObra={s(data.get('CodObra'))}"
        ),
        "raw_data": clean_row(data),
    }


class Command(BaseCommand):
    help = "Sincroniza tblPersonalObra y reconcilia TareaRecursoReal de mano de obra."

    def add_arguments(self, parser):
        parser.add_argument("--folder", required=True)
        parser.add_argument("--team", required=True)
        parser.add_argument("--commit", action="store_true")
        parser.add_argument("--reconcile-reales", action="store_true")
        parser.add_argument("--sample", type=int, default=20)
        parser.add_argument("--json-out", default="")

    def handle(self, *args, **options):
        Team = apps.get_model("usuarios", "Team")
        EmpleadoObra = apps.get_model("planificacion_obra", "EmpleadoObra")
        TareaRecursoReal = apps.get_model("planificacion_obra", "TareaRecursoReal")

        folder = Path(options["folder"])
        file_path = folder / "tblPersonalObra.xlsx"
        team_name = options["team"]
        commit = bool(options["commit"])
        reconcile_reales = bool(options["reconcile_reales"])
        sample = int(options["sample"] or 20)

        try:
            team = Team.objects.get(name=team_name)
        except Team.DoesNotExist as exc:
            raise CommandError(f"No existe Team con name={team_name}") from exc

        rows = list(read_xlsx(file_path))
        access = {}

        for row in rows:
            legacy_id = to_int(row.get("IdPersonal"))
            if legacy_id is None:
                continue
            access[legacy_id] = row

        empleados = {
            e.legacy_id: e
            for e in EmpleadoObra.objects.filter(team=team).exclude(legacy_id=None)
        }

        missing = []
        changed = []

        for legacy_id, row in sorted(access.items()):
            emp = empleados.get(legacy_id)
            payload = build_empleado_payload(row)

            if emp is None:
                missing.append({"legacy_id": legacy_id, "payload": payload})
                continue

            diffs = {}
            for field, new_value in payload.items():
                old_value = getattr(emp, field)
                if str(old_value or "") != str(new_value or ""):
                    diffs[field] = {
                        "old": clean_json(old_value),
                        "new": clean_json(new_value),
                    }

            if diffs:
                changed.append({
                    "legacy_id": legacy_id,
                    "nombre_actual": emp.nombre,
                    "nombre_access": payload["nombre"],
                    "diff_fields": list(diffs.keys()),
                    "payload": payload,
                })

        qs_null = TareaRecursoReal.objects.filter(
            team=team,
            legacy_tipo_recurso__in=list(LABOR_TYPES),
            empleado__isnull=True,
        )
        qs_generic = TareaRecursoReal.objects.filter(
            team=team,
            legacy_tipo_recurso__in=list(LABOR_TYPES),
            empleado__legacy_id=0,
        )
        qs_reconcile = (qs_null | qs_generic).distinct()

        reconcilable = []
        without_match = 0

        for r in qs_reconcile.only(
            "id",
            "legacy_id_recurso_tarea",
            "legacy_id_recurso",
            "empleado_id",
            "legacy_tipo_recurso",
        ):
            emp = empleados.get(r.legacy_id_recurso)
            if emp and emp.legacy_id != 0:
                reconcilable.append({
                    "id": r.id,
                    "legacy_id_recurso_tarea": r.legacy_id_recurso_tarea,
                    "legacy_id_recurso": r.legacy_id_recurso,
                    "empleado_id": emp.id,
                    "empleado_nombre": emp.nombre,
                })
            else:
                without_match += 1

        report = {
            "mode": "COMMIT" if commit else "DRY_RUN",
            "folder": str(folder),
            "file": str(file_path),
            "team": {"id": team.id, "name": str(team)},
            "access_personal_count": len(access),
            "existing_empleados_count": len(empleados),
            "missing_count": len(missing),
            "changed_count": len(changed),
            "reconcile_reales": reconcile_reales,
            "reconcilable_reales_count": len(reconcilable),
            "without_match_count": without_match,
            "missing_sample": [
                {"legacy_id": x["legacy_id"], "nombre": x["payload"]["nombre"]}
                for x in missing[:sample]
            ],
            "changed_sample": [
                {
                    "legacy_id": x["legacy_id"],
                    "nombre_actual": x["nombre_actual"],
                    "nombre_access": x["nombre_access"],
                    "diff_fields": x["diff_fields"],
                }
                for x in changed[:sample]
            ],
            "reconcilable_sample": reconcilable[:sample],
            "created": {
                "empleados": 0,
                "empleados_updated": 0,
                "reales_reconciled": 0,
            },
            "backup": None,
        }

        self.stdout.write("")
        self.stdout.write("=== SYNC ACCESS PERSONAL OBRA ===")
        self.stdout.write(f"Folder: {folder}")
        self.stdout.write(f"Team: {team}")
        self.stdout.write(f"Modo: {'COMMIT' if commit else 'DRY-RUN'}")
        self.stdout.write(f"Personal Access: {len(access)}")
        self.stdout.write(f"Empleados BD: {len(empleados)}")
        self.stdout.write(f"Faltantes: {len(missing)}")
        self.stdout.write(f"Cambios detectados: {len(changed)}")
        self.stdout.write(f"Reales corregibles: {len(reconcilable)}")
        self.stdout.write(f"Sin match: {without_match}")

        if commit:
            backup_dir = Path("/app/backups") / f"personalobra_sync_{timezone.now().strftime('%Y%m%d_%H%M%S')}"
            backup_dir.mkdir(parents=True, exist_ok=True)
            report["backup"] = str(backup_dir)

            call_command(
                "dumpdata",
                "planificacion_obra",
                output=str(backup_dir / "planificacion_obra_before.json"),
                indent=2,
                verbosity=0,
            )

            with transaction.atomic():
                for item in missing:
                    obj = EmpleadoObra.objects.create(
                        team=team,
                        legacy_id=item["legacy_id"],
                        **item["payload"],
                    )
                    empleados[item["legacy_id"]] = obj
                    report["created"]["empleados"] += 1

                for item in changed:
                    emp = EmpleadoObra.objects.get(team=team, legacy_id=item["legacy_id"])
                    for field, value in item["payload"].items():
                        setattr(emp, field, value)
                    emp.save()
                    empleados[item["legacy_id"]] = emp
                    report["created"]["empleados_updated"] += 1

                if reconcile_reales:
                    for item in reconcilable:
                        emp = empleados.get(item["legacy_id_recurso"])
                        if not emp:
                            continue
                        updated = TareaRecursoReal.objects.filter(
                            team=team,
                            id=item["id"],
                        ).update(
                            empleado=emp,
                            legacy_personal=emp.legacy_id,
                            updated_at=timezone.now(),
                        )
                        report["created"]["reales_reconciled"] += updated

        if options["json_out"]:
            out = Path(options["json_out"])
            out.parent.mkdir(parents=True, exist_ok=True)
            out.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
            self.stdout.write(f"JSON guardado en: {out}")

        self.stdout.write("")
        self.stdout.write("=== RESULTADO ===")
        self.stdout.write(json.dumps(report["created"], indent=2, ensure_ascii=False))
        self.stdout.write("OK")

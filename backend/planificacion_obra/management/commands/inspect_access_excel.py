from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Inspecciona cabeceras y primeras filas de un Excel exportado desde Access."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)
        parser.add_argument("--rows", type=int, default=5)

    def handle(self, *args, **options):
        path = Path(options["xlsx_path"])
        sample_rows = options["rows"]

        if not path.exists():
            raise CommandError(f"No existe el fichero: {path}")

        wb = load_workbook(path, data_only=True, read_only=True)

        self.stdout.write(self.style.SUCCESS(f"Archivo: {path}"))
        self.stdout.write(f"Hojas: {', '.join(wb.sheetnames)}")

        for ws in wb.worksheets:
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS(f"=== HOJA: {ws.title} ==="))
            self.stdout.write(f"Filas máximas: {ws.max_row}")
            self.stdout.write(f"Columnas máximas: {ws.max_column}")

            rows = ws.iter_rows(values_only=True)
            headers = next(rows, None)

            if not headers:
                self.stdout.write("Hoja vacía.")
                continue

            headers = [str(h).strip() if h is not None else "" for h in headers]

            self.stdout.write("")
            self.stdout.write("CABECERAS:")
            for i, h in enumerate(headers, start=1):
                self.stdout.write(f"{i:02d}. {h}")

            self.stdout.write("")
            self.stdout.write(f"PRIMERAS {sample_rows} FILAS:")
            for idx, row in enumerate(rows, start=1):
                if idx > sample_rows:
                    break

                data = dict(zip(headers, row))
                self.stdout.write(f"--- fila {idx + 1} ---")
                for key, value in data.items():
                    if key:
                        self.stdout.write(f"{key}: {value}")

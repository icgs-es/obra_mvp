from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Importa base de planificación desde Access: obras, fases y unidades."

    def add_arguments(self, parser):
        parser.add_argument("base_path", type=str)
        parser.add_argument("--team-id", type=int, default=1)
        parser.add_argument("--commit", action="store_true")

    def read_rows(self, path):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if any(v is not None for v in data.values()):
                yield data

    def clean_text(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def clean_int(self, value):
        if value in (None, ""):
            return None
        try:
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return int(value)
        except (TypeError, ValueError):
            return None

    def clean_decimal(self, value):
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return None

    def clean_date(self, value):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None

    def clean_bool(self, value):
        if isinstance(value, bool):
            return value
        if value in (None, ""):
            return False
        return str(value).strip().lower() in ("true", "1", "sí", "si", "yes")

    def json_safe(self, data):
        safe = {}
        for k, v in data.items():
            if isinstance(v, (datetime, date)):
                safe[k] = v.isoformat()
            elif isinstance(v, Decimal):
                safe[k] = str(v)
            else:
                safe[k] = v
        return safe

    def handle(self, *args, **options):
        base_path = Path(options["base_path"])
        team_id = options["team_id"]
        commit = options["commit"]

        obras_file = base_path / "tblObras.xlsx"
        fases_file = base_path / "tblObraFases.xlsx"
        viviendas_file = base_path / "tblFaseViviendas.xlsx"

        for f in (obras_file, fases_file, viviendas_file):
            if not f.exists():
                raise CommandError(f"No existe: {f}")

        Team = apps.get_model("usuarios", "Team")
        ObraPlanificacion = apps.get_model("planificacion_obra", "ObraPlanificacion")
        FaseObra = apps.get_model("planificacion_obra", "FaseObra")
        UnidadObra = apps.get_model("planificacion_obra", "UnidadObra")

        team = Team.objects.filter(id=team_id).first()
        if not team:
            raise CommandError(f"No existe Team con id={team_id}")

        modo = "COMMIT REAL" if commit else "SIMULACION"
        self.stdout.write(self.style.WARNING(f"Modo: {modo}"))
        self.stdout.write(f"Team destino: {team.id} · {team.name}")

        obras_rows = list(self.read_rows(obras_file))
        fases_rows = list(self.read_rows(fases_file))
        viviendas_rows = list(self.read_rows(viviendas_file))

        obras_crear = obras_actualizar = obras_ok = 0
        fases_crear = fases_actualizar = fases_ok = fases_sin_obra = 0
        unidades_crear = unidades_actualizar = unidades_ok = unidades_sin_obra = 0

        with transaction.atomic():
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("=== OBRAS ==="))

            for r in obras_rows:
                cod_obra = self.clean_int(r.get("CodObra"))
                nombre = self.clean_text(r.get("NombreObra"))

                if cod_obra is None or not nombre:
                    continue

                codigo = str(cod_obra)

                obj = ObraPlanificacion.objects.filter(
                    team=team,
                    legacy_cod_obra=cod_obra,
                ).first()

                defaults = {
                    "codigo": codigo,
                    "nombre": nombre,
                    "descripcion": self.clean_text(r.get("Descripcion")),
                    "direccion": self.clean_text(r.get("DireccionObra")),
                    "poblacion": self.clean_text(r.get("PoblacionObra")),
                    "provincia": self.clean_text(r.get("Provincia")),
                    "aparejador": self.clean_text(r.get("Aparejador")),
                    "jefe_obra": self.clean_text(r.get("JefeObra")),
                    "fecha_inicio": self.clean_date(r.get("FechaInicio")),
                    "fecha_fin": self.clean_date(r.get("FechaFin")),
                    "importe_obra": self.clean_decimal(r.get("ImporteObra")),
                    "total_viviendas": self.clean_int(r.get("TotalViviendas")),
                    "raw_data": self.json_safe(r),
                }

                if obj is None:
                    obras_crear += 1
                    self.stdout.write(f"CREAR Obra {codigo} => {nombre}")
                    obj = ObraPlanificacion.objects.create(
                        team=team,
                        legacy_cod_obra=cod_obra,
                        **defaults,
                    )
                else:
                    changed = any(getattr(obj, k) != v for k, v in defaults.items())
                    if changed:
                        obras_actualizar += 1
                        self.stdout.write(f"ACTUALIZAR Obra {codigo} => {nombre}")
                        for k, v in defaults.items():
                            setattr(obj, k, v)
                        obj.save()
                    else:
                        obras_ok += 1

            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("=== FASES / EDIFICIOS ==="))

            for r in fases_rows:
                cod_obra = self.clean_int(r.get("CodObra"))
                cod_fase = self.clean_int(r.get("CodFase"))
                nombre = self.clean_text(r.get("NombreFase")) or f"FASE {cod_fase}"

                if cod_obra is None or cod_fase is None:
                    continue

                obra = ObraPlanificacion.objects.filter(
                    team=team,
                    legacy_cod_obra=cod_obra,
                ).first()

                if not obra:
                    fases_sin_obra += 1
                    self.stdout.write(self.style.ERROR(f"SIN OBRA: fase {cod_obra} · {cod_fase}"))
                    continue

                obj = FaseObra.objects.filter(
                    team=team,
                    obra=obra,
                    legacy_cod_fase=cod_fase,
                ).first()

                defaults = {
                    "nombre": nombre,
                    "cantidad_viviendas": self.clean_int(r.get("CantidadViviendas")),
                    "num_vivienda_inicial": self.clean_int(r.get("NumViviendaInicial")),
                    "num_vivienda_final": self.clean_int(r.get("NumViviendaFinal")),
                    "vivienda_lateral": self.clean_bool(r.get("ViviendaLateral")),
                    "cantidad_viviendas_laterales": self.clean_int(r.get("CantidadViviendaLaterales")),
                    "zona_comun": self.clean_bool(r.get("ZonaComun")),
                    "observaciones": self.clean_text(r.get("Observaciones")),
                    "raw_data": self.json_safe(r),
                }

                if obj is None:
                    fases_crear += 1
                    if fases_crear <= 20:
                        self.stdout.write(f"CREAR Fase Obra {cod_obra} · {cod_fase} => {nombre}")
                    obj = FaseObra.objects.create(
                        team=team,
                        obra=obra,
                        legacy_cod_fase=cod_fase,
                        **defaults,
                    )
                else:
                    changed = any(getattr(obj, k) != v for k, v in defaults.items())
                    if changed:
                        fases_actualizar += 1
                        if fases_actualizar <= 20:
                            self.stdout.write(f"ACTUALIZAR Fase Obra {cod_obra} · {cod_fase} => {nombre}")
                        for k, v in defaults.items():
                            setattr(obj, k, v)
                        obj.save()
                    else:
                        fases_ok += 1

            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("=== UNIDADES ==="))

            for r in viviendas_rows:
                cod_obra = self.clean_int(r.get("CodObra"))
                cod_fase = self.clean_int(r.get("CodFase"))
                cod_vivienda = self.clean_text(r.get("CodVivienda"))
                nivel = self.clean_text(r.get("Nivel"))

                if cod_obra is None or cod_fase is None or not cod_vivienda:
                    continue

                obra = ObraPlanificacion.objects.filter(
                    team=team,
                    legacy_cod_obra=cod_obra,
                ).first()

                if not obra:
                    unidades_sin_obra += 1
                    self.stdout.write(self.style.ERROR(f"SIN OBRA: unidad {cod_obra} · {cod_fase} · {cod_vivienda}"))
                    continue

                fase = FaseObra.objects.filter(
                    team=team,
                    obra=obra,
                    legacy_cod_fase=cod_fase,
                ).first()

                edificio = fase.nombre if fase else f"FASE {cod_fase}"

                obj = UnidadObra.objects.filter(
                    team=team,
                    obra=obra,
                    legacy_cod_fase=cod_fase,
                    legacy_cod_vivienda=cod_vivienda,
                    nivel=nivel,
                ).first()

                defaults = {
                    "fase": fase,
                    "legacy_cod_obra": cod_obra,
                    "edificio": edificio,
                    "vivienda": cod_vivienda,
                    "tipo": self.clean_text(r.get("Tipo")),
                    "observaciones": self.clean_text(r.get("Observaciones")),
                    "raw_data": self.json_safe(r),
                }

                if obj is None:
                    unidades_crear += 1
                    if unidades_crear <= 30:
                        self.stdout.write(
                            f"CREAR Unidad Obra {cod_obra} · {edificio} · Viv {cod_vivienda} · Planta {planta or '-'}"
                        )
                    obj = UnidadObra.objects.create(
                        team=team,
                        obra=obra,
                        legacy_cod_fase=cod_fase,
                        legacy_cod_vivienda=cod_vivienda,
                        nivel=nivel,
                        **defaults,
                    )
                else:
                    changed = any(getattr(obj, k) != v for k, v in defaults.items())
                    if changed:
                        unidades_actualizar += 1
                        if unidades_actualizar <= 30:
                            self.stdout.write(
                                f"ACTUALIZAR Unidad Obra {cod_obra} · {edificio} · Viv {cod_vivienda} · Planta {planta or '-'}"
                            )
                        for k, v in defaults.items():
                            setattr(obj, k, v)
                        obj.save()
                    else:
                        unidades_ok += 1

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== RESUMEN BASE PLANIFICACION ==="))
        self.stdout.write(f"Obras leídas: {len(obras_rows)}")
        self.stdout.write(f"Obras crear: {obras_crear}")
        self.stdout.write(f"Obras actualizar: {obras_actualizar}")
        self.stdout.write(f"Obras sin cambios: {obras_ok}")
        self.stdout.write("")
        self.stdout.write(f"Fases leídas: {len(fases_rows)}")
        self.stdout.write(f"Fases crear: {fases_crear}")
        self.stdout.write(f"Fases actualizar: {fases_actualizar}")
        self.stdout.write(f"Fases sin cambios: {fases_ok}")
        self.stdout.write(f"Fases sin obra: {fases_sin_obra}")
        self.stdout.write("")
        self.stdout.write(f"Unidades leídas: {len(viviendas_rows)}")
        self.stdout.write(f"Unidades crear: {unidades_crear}")
        self.stdout.write(f"Unidades actualizar: {unidades_actualizar}")
        self.stdout.write(f"Unidades sin cambios: {unidades_ok}")
        self.stdout.write(f"Unidades sin obra: {unidades_sin_obra}")

        if not commit:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("SIMULACION: no se ha guardado nada."))
        else:
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("IMPORTACION APLICADA."))

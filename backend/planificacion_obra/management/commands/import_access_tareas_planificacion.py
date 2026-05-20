from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Importa tareas de obra desde tblTareas.xlsx hacia TareaObra."

    def add_arguments(self, parser):
        parser.add_argument("base_path", type=str)
        parser.add_argument("--team-id", type=int, default=1)
        parser.add_argument("--commit", action="store_true")

    def read_rows(self, path):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if any(v is not None for v in data.values()):
                yield data

    def txt(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def clean_int(self, value):
        if value in (None, ""):
            return None
        try:
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return int(value)
        except (TypeError, ValueError):
            return None

    def clean_decimal(self, value):
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return None

    def clean_date(self, value):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None

    def clean_bool(self, value):
        if isinstance(value, bool):
            return value
        if value in (None, ""):
            return False
        return str(value).strip().lower() in ("true", "1", "-1", "sí", "si", "yes")

    def json_safe(self, data):
        safe = {}
        for k, v in data.items():
            if isinstance(v, (datetime, date, time)):
                safe[k] = v.isoformat()
            elif isinstance(v, Decimal):
                safe[k] = str(v)
            else:
                safe[k] = v
        return safe

    def changed(self, obj, defaults):
        for key, value in defaults.items():
            current = getattr(obj, key)
            if hasattr(current, "pk") and hasattr(value, "pk"):
                if current.pk != value.pk:
                    return True
            elif current != value:
                return True
        return False

    def handle(self, *args, **options):
        base_path = Path(options["base_path"])
        team_id = options["team_id"]
        commit = options["commit"]

        tareas_file = base_path / "tblTareas.xlsx"
        if not tareas_file.exists():
            raise CommandError(f"No existe: {tareas_file}")

        Team = apps.get_model("usuarios", "Team")
        ObraPlanificacion = apps.get_model("planificacion_obra", "ObraPlanificacion")
        UnidadObra = apps.get_model("planificacion_obra", "UnidadObra")
        CapituloCatalogo = apps.get_model("planificacion_obra", "CapituloCatalogo")
        PartidaCatalogo = apps.get_model("planificacion_obra", "PartidaCatalogo")
        TareaObra = apps.get_model("planificacion_obra", "TareaObra")

        team = Team.objects.filter(id=team_id).first()
        if not team:
            raise CommandError(f"No existe Team con id={team_id}")

        modo = "COMMIT REAL" if commit else "SIMULACION"
        self.stdout.write(self.style.WARNING(f"Modo: {modo}"))
        self.stdout.write(f"Team destino: {team.id} · {team.name}")
        self.stdout.write("UnidadObra se enlaza por CodObra + CodFase + CodVivienda.")
        self.stdout.write("tblTareas.Planta se conserva en TareaObra.legacy_planta.")

        rows = list(self.read_rows(tareas_file))

        obras = {
            o.legacy_cod_obra: o
            for o in ObraPlanificacion.objects.filter(team=team)
        }

        unidades = {
            (
                u.legacy_cod_obra,
                u.legacy_cod_fase,
                self.txt(u.legacy_cod_vivienda),
            ): u
            for u in UnidadObra.objects.filter(team=team)
        }

        capitulos = {
            c.codigo: c
            for c in CapituloCatalogo.objects.filter(team=team)
        }

        partidas = {
            (p.capitulo.codigo, p.codigo): p
            for p in PartidaCatalogo.objects.select_related("capitulo").filter(team=team)
        }

        crear = actualizar = sin_cambios = 0
        sin_obra = sin_unidad = sin_capitulo = sin_partida = 0

        with transaction.atomic():
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("=== TAREAS OBRA ==="))

            for row in rows:
                cod_obra = self.clean_int(row.get("CodObra"))
                cod_fase = self.clean_int(row.get("CodFase"))
                cod_vivienda = self.txt(row.get("CodVivienda"))
                planta_tarea = self.txt(row.get("Planta"))
                cod_capitulo = self.txt(row.get("Capitulo"))
                cod_partida = self.txt(row.get("Partida"))
                orden = self.clean_int(row.get("Orden"))

                legacy_key = f"{cod_obra}|{cod_fase}|{cod_vivienda}|{planta_tarea}|{cod_capitulo}|{cod_partida}|{orden}"

                obra = obras.get(cod_obra)
                unidad = unidades.get((cod_obra, cod_fase, cod_vivienda))
                capitulo = capitulos.get(cod_capitulo)
                partida = partidas.get((cod_capitulo, cod_partida))

                if not obra:
                    sin_obra += 1
                    continue
                if not unidad:
                    sin_unidad += 1
                    continue
                if not capitulo:
                    sin_capitulo += 1
                    continue
                if not partida:
                    sin_partida += 1
                    continue

                obj = TareaObra.objects.filter(team=team, legacy_key=legacy_key).first()

                defaults = {
                    "obra": obra,
                    "unidad_obra": unidad,
                    "capitulo": capitulo,
                    "partida": partida,

                    "legacy_cod_obra": cod_obra,
                    "legacy_cod_fase": cod_fase,
                    "legacy_cod_vivienda": cod_vivienda,
                    "legacy_planta": planta_tarea,
                    "legacy_capitulo": cod_capitulo,
                    "legacy_partida": cod_partida,
                    "legacy_orden": orden,

                    "programacion": self.txt(row.get("Programacion")),
                    "porcentaje_completado": self.clean_decimal(row.get("PorcentajeCompletado")),

                    "inicio_tarea": self.clean_date(row.get("InicioTarea")),
                    "fin_tarea": self.clean_date(row.get("FinTarea")),
                    "dias": self.clean_decimal(row.get("Dias")),
                    "horas": self.clean_decimal(row.get("Horas")),

                    "inicio_real": self.clean_date(row.get("InicioReal")),
                    "fin_real": self.clean_date(row.get("FinReal")),
                    "dias_reales": self.clean_decimal(row.get("DiasReales")),
                    "horas_reales": self.clean_decimal(row.get("HorasReales")),

                    "inicio_estimado": self.clean_date(row.get("InicioEstimado")),
                    "fin_estimado": self.clean_date(row.get("FinEstimado")),
                    "dias_estimados": self.clean_decimal(row.get("DiasEstimados")),
                    "horas_estimadas": self.clean_decimal(row.get("HorasEstimadas")),

                    "personas_a_utilizar": self.clean_decimal(row.get("Personasautilizar")),
                    "personas_utilizadas": self.clean_decimal(row.get("Personasutilizadas")),

                    "importe_tarea": self.clean_decimal(row.get("ImporteTarea")),
                    "importe_tarea_real": self.clean_decimal(row.get("ImporteTareaReal")),
                    "importe_tarea_estimado": self.clean_decimal(row.get("ImporteTareaEstimado")),

                    "tipo_partida": self.txt(row.get("TipoPartida")),
                    "unidad": self.txt(row.get("Unidad")),
                    "cantidad": self.clean_decimal(row.get("Cantidad")),
                    "precio_unidad": self.clean_decimal(row.get("PrecioUnidad")),

                    "con_incidencias": self.clean_bool(row.get("ConIncidencias")),
                    "observaciones": self.txt(row.get("Observaciones")),
                    "raw_data": self.json_safe(row),
                }

                if obj is None:
                    crear += 1
                    if crear <= 40:
                        self.stdout.write(
                            f"CREAR Tarea {legacy_key} => {obra.nombre} · {cod_capitulo} · {cod_partida}"
                        )
                    TareaObra.objects.create(
                        team=team,
                        legacy_key=legacy_key,
                        **defaults,
                    )
                else:
                    if self.changed(obj, defaults):
                        actualizar += 1
                        if actualizar <= 40:
                            self.stdout.write(f"ACTUALIZAR Tarea {legacy_key}")
                        for k, v in defaults.items():
                            setattr(obj, k, v)
                        obj.save()
                    else:
                        sin_cambios += 1

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== RESUMEN TAREAS OBRA ==="))
        self.stdout.write(f"Filas leídas: {len(rows)}")
        self.stdout.write(f"Crear: {crear}")
        self.stdout.write(f"Actualizar: {actualizar}")
        self.stdout.write(f"Sin cambios: {sin_cambios}")
        self.stdout.write(f"Sin obra: {sin_obra}")
        self.stdout.write(f"Sin unidad: {sin_unidad}")
        self.stdout.write(f"Sin capítulo: {sin_capitulo}")
        self.stdout.write(f"Sin partida: {sin_partida}")

        if not commit:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("SIMULACION: no se ha guardado nada."))
        else:
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("IMPORTACION APLICADA."))

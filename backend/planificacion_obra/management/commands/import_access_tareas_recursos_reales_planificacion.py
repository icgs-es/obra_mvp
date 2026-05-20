from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Importa recursos reales/ejecutados desde tblTareasRecursos.xlsx."

    def add_arguments(self, parser):
        parser.add_argument("base_path", type=str)
        parser.add_argument("--team-id", type=int, default=1)
        parser.add_argument("--commit", action="store_true")

    def txt(self, value):
        return "" if value is None else str(value).strip()

    def clean_int(self, value):
        if value in (None, ""):
            return None
        try:
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return int(value)
        except Exception:
            return None

    def clean_decimal(self, value):
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return None

    def clean_date(self, value):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None

    def clean_bool(self, value):
        if isinstance(value, bool):
            return value
        if value in (None, ""):
            return False
        return str(value).strip().lower() in ("true", "1", "-1", "sí", "si", "yes")

    def json_safe(self, data):
        safe = {}
        for k, v in data.items():
            if isinstance(v, (datetime, date, time)):
                safe[k] = v.isoformat()
            elif isinstance(v, Decimal):
                safe[k] = str(v)
            else:
                safe[k] = v
        return safe

    def read_rows(self, path):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if any(v is not None for v in data.values()):
                yield data

    def changed(self, obj, defaults):
        for key, value in defaults.items():
            current = getattr(obj, key)
            if hasattr(current, "pk") and hasattr(value, "pk"):
                if current.pk != value.pk:
                    return True
            elif current != value:
                return True
        return False

    def handle(self, *args, **options):
        base_path = Path(options["base_path"])
        team_id = options["team_id"]
        commit = options["commit"]

        file_path = base_path / "tblTareasRecursos.xlsx"
        if not file_path.exists():
            raise CommandError(f"No existe: {file_path}")

        Team = apps.get_model("usuarios", "Team")
        TareaObra = apps.get_model("planificacion_obra", "TareaObra")
        UnidadObra = apps.get_model("planificacion_obra", "UnidadObra")
        PartidaCatalogo = apps.get_model("planificacion_obra", "PartidaCatalogo")
        RecursoCatalogo = apps.get_model("planificacion_obra", "RecursoCatalogo")
        EmpleadoObra = apps.get_model("planificacion_obra", "EmpleadoObra")
        RecursoAlmacenMovimiento = apps.get_model("planificacion_obra", "RecursoAlmacenMovimiento")
        TareaRecursoReal = apps.get_model("planificacion_obra", "TareaRecursoReal")

        team = Team.objects.filter(id=team_id).first()
        if not team:
            raise CommandError(f"No existe Team con id={team_id}")

        modo = "COMMIT REAL" if commit else "SIMULACION"
        self.stdout.write(self.style.WARNING(f"Modo: {modo}"))
        self.stdout.write(f"Team destino: {team.id} · {team.name}")
        self.stdout.write("TareaObra se enlaza SIN usar Orden: CodObra + CodFase + CodVivienda + Planta + Partida.")
        self.stdout.write("IdRecurso puede enlazar con EmpleadoObra o RecursoCatalogo.")

        rows = list(self.read_rows(file_path))

        tareas = {
            (
                t.legacy_cod_obra,
                t.legacy_cod_fase,
                self.txt(t.legacy_cod_vivienda),
                self.txt(t.legacy_planta),
                self.txt(t.legacy_partida),
            ): t
            for t in TareaObra.objects.filter(team=team)
        }

        unidades = {
            (
                u.legacy_cod_obra,
                u.legacy_cod_fase,
                self.txt(u.legacy_cod_vivienda),
            ): u
            for u in UnidadObra.objects.filter(team=team)
        }

        partidas = {p.codigo: p for p in PartidaCatalogo.objects.filter(team=team)}
        recursos = {r.legacy_id: r for r in RecursoCatalogo.objects.filter(team=team)}
        empleados = {e.legacy_id: e for e in EmpleadoObra.objects.filter(team=team)}
        movimientos = {
            m.legacy_id_movimiento: m
            for m in RecursoAlmacenMovimiento.objects.filter(team=team)
        }

        crear = actualizar = sin_cambios = 0
        sin_id = 0
        sin_tarea = sin_unidad = sin_partida = 0
        sin_recurso_y_sin_empleado = 0
        con_empleado = con_recurso = 0
        con_movimiento = sin_movimiento = 0

        with transaction.atomic():
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("=== RECURSOS REALES DE TAREAS ==="))

            for row in rows:
                legacy_id = self.clean_int(row.get("IdRecursoTarea"))
                if legacy_id is None:
                    sin_id += 1
                    continue

                cod_obra = self.clean_int(row.get("CodObra"))
                cod_fase = self.clean_int(row.get("CodFase"))
                cod_vivienda = self.txt(row.get("CodVivienda"))
                planta = self.txt(row.get("Planta"))
                capitulo = self.txt(row.get("Capitulo"))
                partida_cod = self.txt(row.get("Partida"))
                id_recurso = self.clean_int(row.get("IdRecurso"))
                tipo_recurso = self.txt(row.get("TipoRecurso"))
                id_mov = self.clean_int(row.get("IdMovimientoAlmacen"))

                tarea_obra = tareas.get((cod_obra, cod_fase, cod_vivienda, planta, partida_cod))
                if not tarea_obra:
                    sin_tarea += 1

                unidad_obra = unidades.get((cod_obra, cod_fase, cod_vivienda))
                if not unidad_obra:
                    sin_unidad += 1

                partida = partidas.get(partida_cod)
                if not partida:
                    sin_partida += 1

                empleado = empleados.get(id_recurso)
                recurso = recursos.get(id_recurso)

                if empleado:
                    con_empleado += 1
                if recurso:
                    con_recurso += 1
                if not empleado and not recurso:
                    sin_recurso_y_sin_empleado += 1

                movimiento = None
                if id_mov is not None:
                    movimiento = movimientos.get(id_mov)
                    if movimiento:
                        con_movimiento += 1
                    else:
                        sin_movimiento += 1

                defaults = {
                    "tarea_obra": tarea_obra,
                    "unidad_obra": unidad_obra,
                    "partida": partida,
                    "recurso": recurso,
                    "empleado": empleado,
                    "movimiento_almacen": movimiento,

                    "legacy_cod_obra": cod_obra,
                    "legacy_cod_fase": cod_fase,
                    "legacy_cod_vivienda": cod_vivienda,
                    "legacy_planta": planta,
                    "legacy_capitulo": capitulo,
                    "legacy_partida": partida_cod,
                    "legacy_id_recurso": id_recurso,
                    "legacy_tipo_recurso": tipo_recurso,
                    "legacy_personal": self.clean_int(row.get("Personal")),
                    "legacy_id_movimiento_almacen": id_mov,
                    "legacy_orden_recurso": self.clean_int(row.get("Orden")),

                    "unidad": self.txt(row.get("Unidad")),
                    "cantidad": self.clean_decimal(row.get("Cantidad")),
                    "precio_unidad": self.clean_decimal(row.get("PrecioUnidad")),
                    "dias": self.clean_decimal(row.get("Dias")),
                    "dias_reales": self.clean_decimal(row.get("DiasReales")),
                    "horas": self.clean_decimal(row.get("Horas")),
                    "horas_reales": self.clean_decimal(row.get("HorasReales")),
                    "inicio_recurso_real": self.clean_date(row.get("InicioRecursoReal")),
                    "fin_recurso_real": self.clean_date(row.get("FinRecursoReal")),
                    "costo_recurso": self.clean_decimal(row.get("CostoRecurso")),
                    "costo_recurso_real": self.clean_decimal(row.get("CostoRecursoReal")),
                    "control_suministros": self.clean_bool(row.get("ControlSuministros")),
                    "avisar": self.clean_int(row.get("Avisar")),
                    "id_proveedor": self.txt(row.get("IdProveedor")),
                    "cod_albaran": self.txt(row.get("CodAlbaran")),
                    "num_linea_albaran": self.clean_int(row.get("NumLineaAlbaran")),
                    "cod_factura": self.txt(row.get("CodFactura")),
                    "num_linea_factura": self.clean_int(row.get("NumLineaFactura")),
                    "observaciones": self.txt(row.get("Observaciones")),
                    "raw_data": self.json_safe(row),
                }

                obj = TareaRecursoReal.objects.filter(
                    team=team,
                    legacy_id_recurso_tarea=legacy_id,
                ).first()

                if obj is None:
                    crear += 1
                    if crear <= 40:
                        destino = "empleado" if empleado else ("recurso" if recurso else "legacy")
                        self.stdout.write(
                            f"CREAR Real {legacy_id} => {cod_obra}|{cod_fase}|{cod_vivienda}|{planta}|{partida_cod} · {tipo_recurso} · id {id_recurso} · {destino}"
                        )
                    TareaRecursoReal.objects.create(
                        team=team,
                        legacy_id_recurso_tarea=legacy_id,
                        **defaults,
                    )
                else:
                    if self.changed(obj, defaults):
                        actualizar += 1
                        if actualizar <= 40:
                            self.stdout.write(f"ACTUALIZAR Real {legacy_id}")
                        for k, v in defaults.items():
                            setattr(obj, k, v)
                        obj.save()
                    else:
                        sin_cambios += 1

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== RESUMEN RECURSOS REALES ==="))
        self.stdout.write(f"Filas leídas: {len(rows)}")
        self.stdout.write(f"Crear: {crear}")
        self.stdout.write(f"Actualizar: {actualizar}")
        self.stdout.write(f"Sin cambios: {sin_cambios}")
        self.stdout.write(f"Sin IdRecursoTarea: {sin_id}")
        self.stdout.write(f"Sin tarea enlazada: {sin_tarea}")
        self.stdout.write(f"Sin unidad enlazada: {sin_unidad}")
        self.stdout.write(f"Sin partida enlazada: {sin_partida}")
        self.stdout.write(f"Con empleado enlazado: {con_empleado}")
        self.stdout.write(f"Con recurso catálogo enlazado: {con_recurso}")
        self.stdout.write(f"Sin empleado ni recurso catálogo: {sin_recurso_y_sin_empleado}")
        self.stdout.write(f"Con movimiento almacén enlazado: {con_movimiento}")
        self.stdout.write(f"Sin movimiento almacén cuando viene informado: {sin_movimiento}")

        if not commit:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("SIMULACION: no se ha guardado nada."))
        else:
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("IMPORTACION APLICADA."))

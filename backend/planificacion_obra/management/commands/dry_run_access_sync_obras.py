from __future__ import annotations

import json
from collections import Counter
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook


REQUIRED_XLSX = [
    "tblTareas.xlsx",
    "tblTareasRecursosInicial.xlsx",
    "tblTareasRecursos.xlsx",
    "tblRecursos.xlsx",
    "tblRecursoAlmacen.xlsx",
    "tblPartida.xlsx",
]


def s(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).replace("\ufeff", "").strip()


def to_int(value):
    value = s(value)
    if not value:
        return None
    try:
        return int(Decimal(value.replace(",", ".")))
    except (InvalidOperation, ValueError):
        return None


def to_decimal(value):
    value = s(value)
    if not value:
        return None
    try:
        return Decimal(value.replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def as_date_text(value):
    if value is None or s(value) == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return s(value)


def read_xlsx(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    try:
        headers = [s(v) for v in next(rows)]
    except StopIteration:
        wb.close()
        return

    for raw in rows:
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = raw[idx] if idx < len(raw) else None
        yield row

    wb.close()


def get_team(team_name):
    Team = apps.get_model("usuarios", "Team")
    field_names = {f.name for f in Team._meta.fields}

    for field in ["nombre", "name", "razon_social"]:
        if field in field_names:
            obj = Team.objects.filter(**{field: team_name}).first()
            if obj:
                return obj

    # fallback parcial
    for field in ["nombre", "name", "razon_social"]:
        if field in field_names:
            obj = Team.objects.filter(**{f"{field}__icontains": team_name}).first()
            if obj:
                return obj

    return None


def tarea_exact_key_from_access(row):
    return (
        to_int(row.get("CodObra")),
        to_int(row.get("CodFase")),
        s(row.get("CodVivienda")),
        s(row.get("Planta")),
        s(row.get("Capitulo")),
        s(row.get("Partida")),
        to_int(row.get("Orden")),
    )


def tarea_link_key_from_access(row):
    return (
        to_int(row.get("CodObra")),
        to_int(row.get("CodFase")),
        s(row.get("CodVivienda")),
        s(row.get("Planta")),
        s(row.get("Capitulo")),
        s(row.get("Partida")),
    )


def tarea_link_key_no_capitulo_from_inicial(row):
    return (
        to_int(row.get("CodObra")),
        to_int(row.get("CodFase")),
        s(row.get("CodVivienda")),
        s(row.get("Planta")),
        s(row.get("CodPartida")),
    )


def tarea_exact_key_from_model(t):
    return (
        t.legacy_cod_obra,
        t.legacy_cod_fase,
        s(t.legacy_cod_vivienda),
        s(t.legacy_planta),
        s(t.legacy_capitulo),
        s(t.legacy_partida),
        t.legacy_orden,
    )


def tarea_link_key_from_model(t):
    return (
        t.legacy_cod_obra,
        t.legacy_cod_fase,
        s(t.legacy_cod_vivienda),
        s(t.legacy_planta),
        s(t.legacy_capitulo),
        s(t.legacy_partida),
    )


def load_existing_tarea_keys(TareaObra, team):
    qs = TareaObra.objects.filter(team=team)

    exact = set()
    link = set()
    no_capitulo = set()

    for t in qs.only(
        "legacy_cod_obra",
        "legacy_cod_fase",
        "legacy_cod_vivienda",
        "legacy_planta",
        "legacy_capitulo",
        "legacy_partida",
        "legacy_orden",
    ):
        exact.add(tarea_exact_key_from_model(t))
        lk = tarea_link_key_from_model(t)
        link.add(lk)
        no_capitulo.add((lk[0], lk[1], lk[2], lk[3], lk[5]))

    return exact, link, no_capitulo


def model_has_field(model, field_name):
    return field_name in {f.name for f in model._meta.fields}


class Command(BaseCommand):
    help = "Dry-run XLSX de sincronización Access -> Portal INTASA para obra. No modifica BD."

    def add_arguments(self, parser):
        parser.add_argument("--folder", required=True)
        parser.add_argument("--team", required=True)
        parser.add_argument("--sample", type=int, default=8)
        parser.add_argument("--json-out", default="")

    def handle(self, *args, **options):
        folder = Path(options["folder"])
        team_name = options["team"]
        sample_limit = max(1, int(options["sample"] or 8))

        if not folder.exists():
            raise CommandError(f"No existe carpeta: {folder}")

        missing = [name for name in REQUIRED_XLSX if not (folder / name).exists()]
        if missing:
            raise CommandError("Faltan XLSX requeridos: " + ", ".join(missing))

        team = get_team(team_name)
        if not team:
            raise CommandError(f"No encuentro Team: {team_name}")

        TareaObra = apps.get_model("planificacion_obra", "TareaObra")
        PartidaCatalogo = apps.get_model("planificacion_obra", "PartidaCatalogo")
        TareaRecursoReal = apps.get_model("planificacion_obra", "TareaRecursoReal")
        TareaRecursoPrevisto = apps.get_model("planificacion_obra", "TareaRecursoPrevisto")

        existing_tarea_exact, existing_tarea_link, existing_tarea_no_cap = load_existing_tarea_keys(TareaObra, team)

        existing_partidas = set(
            PartidaCatalogo.objects.filter(team=team).values_list("capitulo__codigo", "codigo")
        )

        existing_real_ids = set()
        if model_has_field(TareaRecursoReal, "legacy_id_recurso_tarea"):
            existing_real_ids = set(
                TareaRecursoReal.objects
                .filter(team=team)
                .exclude(legacy_id_recurso_tarea__isnull=True)
                .values_list("legacy_id_recurso_tarea", flat=True)
            )

        existing_previsto_count = TareaRecursoPrevisto.objects.filter(team=team).count()

        report = {
            "folder": str(folder),
            "team": str(team),
            "cloud": {
                "tareas": len(existing_tarea_exact),
                "partidas": len(existing_partidas),
                "recursos_previstos": existing_previsto_count,
                "recursos_reales_con_legacy_id": len(existing_real_ids),
            },
            "tables": {},
        }

        self.stdout.write("")
        self.stdout.write("=== DRY-RUN ACCESS SYNC OBRAS XLSX ===")
        self.stdout.write(f"Folder: {folder}")
        self.stdout.write(f"Team: {team}")
        self.stdout.write("Modo: NO modifica base de datos")
        self.stdout.write("")
        self.stdout.write(f"Tareas cloud existentes: {len(existing_tarea_exact)}")
        self.stdout.write(f"Partidas cloud existentes: {len(existing_partidas)}")
        self.stdout.write(f"Reales cloud con legacy_id_recurso_tarea: {len(existing_real_ids)}")
        self.stdout.write("")

        # tblPartida
        part_report = {
            "rows": 0,
            "existing": 0,
            "new_candidates": 0,
            "duplicate_in_access": 0,
            "by_cod_obra": {},
            "sample_new": [],
        }
        seen_partidas = set()
        by_cod_obra = Counter()

        for row in read_xlsx(folder / "tblPartida.xlsx"):
            part_report["rows"] += 1
            key = (s(row.get("CodCapitulo")), s(row.get("CodPartida")))
            by_cod_obra[s(row.get("CodObra"))] += 1

            if key in seen_partidas:
                part_report["duplicate_in_access"] += 1
            seen_partidas.add(key)

            if key in existing_partidas:
                part_report["existing"] += 1
            else:
                part_report["new_candidates"] += 1
                if len(part_report["sample_new"]) < sample_limit:
                    part_report["sample_new"].append({
                        "capitulo": key[0],
                        "partida": key[1],
                        "nombre": s(row.get("NombrePartida")),
                        "tipo": s(row.get("TipoPartida")),
                        "unidad": s(row.get("Unidad")),
                        "dias_material": s(row.get("DiasMaterial")),
                        "cod_obra": s(row.get("CodObra")),
                    })

        part_report["by_cod_obra"] = dict(by_cod_obra)
        report["tables"]["tblPartida"] = part_report

        # tblTareas
        tareas_report = {
            "rows": 0,
            "existing_by_exact_key": 0,
            "new_candidates": 0,
            "invalid_key": 0,
            "duplicate_in_access": 0,
            "by_cod_obra": {},
            "sample_new": [],
        }
        seen_tareas = set()
        by_cod_obra = Counter()

        for row in read_xlsx(folder / "tblTareas.xlsx"):
            tareas_report["rows"] += 1
            key = tarea_exact_key_from_access(row)
            by_cod_obra[s(row.get("CodObra"))] += 1

            if key in seen_tareas:
                tareas_report["duplicate_in_access"] += 1
            seen_tareas.add(key)

            if None in [key[0], key[1], key[6]] or not key[3] or not key[4] or not key[5]:
                tareas_report["invalid_key"] += 1
                continue

            if key in existing_tarea_exact:
                tareas_report["existing_by_exact_key"] += 1
            else:
                tareas_report["new_candidates"] += 1
                if len(tareas_report["sample_new"]) < sample_limit:
                    tareas_report["sample_new"].append({
                        "key": key,
                        "programacion": s(row.get("Programacion")),
                        "porcentaje": s(row.get("PorcentajeCompletado")),
                        "inicio_tarea": as_date_text(row.get("InicioTarea")),
                        "fin_tarea": as_date_text(row.get("FinTarea")),
                        "inicio_real": as_date_text(row.get("InicioReal")),
                        "fin_real": as_date_text(row.get("FinReal")),
                        "importe_tarea": s(row.get("ImporteTarea")),
                        "importe_real": s(row.get("ImporteTareaReal")),
                    })

        tareas_report["by_cod_obra"] = dict(by_cod_obra)
        report["tables"]["tblTareas"] = tareas_report

        # tblTareasRecursosInicial
        tri_report = {
            "rows": 0,
            "linked_to_tarea": 0,
            "unlinked_to_tarea": 0,
            "by_id_recurso_top": {},
            "sample_unlinked": [],
        }
        by_recurso = Counter()

        for row in read_xlsx(folder / "tblTareasRecursosInicial.xlsx"):
            tri_report["rows"] += 1
            by_recurso[s(row.get("IdRecurso"))] += 1
            key_no_cap = tarea_link_key_no_capitulo_from_inicial(row)

            if key_no_cap in existing_tarea_no_cap:
                tri_report["linked_to_tarea"] += 1
            else:
                tri_report["unlinked_to_tarea"] += 1
                if len(tri_report["sample_unlinked"]) < sample_limit:
                    tri_report["sample_unlinked"].append({
                        "cod_obra": s(row.get("CodObra")),
                        "fase": s(row.get("CodFase")),
                        "vivienda": s(row.get("CodVivienda")),
                        "planta": s(row.get("Planta")),
                        "partida": s(row.get("CodPartida")),
                        "id_recurso": s(row.get("IdRecurso")),
                        "cantidad": s(row.get("Cantidad")),
                        "costo": s(row.get("CostoRecurso")),
                    })

        tri_report["by_id_recurso_top"] = dict(by_recurso.most_common(20))
        report["tables"]["tblTareasRecursosInicial"] = tri_report

        # tblTareasRecursos real
        tr_report = {
            "rows": 0,
            "existing_by_legacy_id": 0,
            "new_by_legacy_id": 0,
            "linked_to_tarea": 0,
            "unlinked_to_tarea": 0,
            "duplicate_id_in_access": 0,
            "date_min": "",
            "date_max": "",
            "by_tipo_top": {},
            "sample_new": [],
            "sample_unlinked": [],
        }
        seen_real_ids = set()
        by_tipo = Counter()
        date_min = None
        date_max = None

        for row in read_xlsx(folder / "tblTareasRecursos.xlsx"):
            tr_report["rows"] += 1

            access_id = to_int(row.get("IdRecursoTarea"))
            key = tarea_link_key_from_access(row)
            by_tipo[s(row.get("TipoRecurso"))] += 1

            if access_id in seen_real_ids:
                tr_report["duplicate_id_in_access"] += 1
            seen_real_ids.add(access_id)

            if existing_real_ids and access_id in existing_real_ids:
                tr_report["existing_by_legacy_id"] += 1
            else:
                tr_report["new_by_legacy_id"] += 1
                if len(tr_report["sample_new"]) < sample_limit:
                    tr_report["sample_new"].append({
                        "id_recurso_tarea": access_id,
                        "key": key,
                        "id_recurso": s(row.get("IdRecurso")),
                        "tipo": s(row.get("TipoRecurso")),
                        "cantidad": s(row.get("Cantidad")),
                        "precio": s(row.get("PrecioUnidad")),
                        "coste_real": s(row.get("CostoRecursoReal")),
                        "inicio": as_date_text(row.get("InicioRecursoReal")),
                        "fin": as_date_text(row.get("FinRecursoReal")),
                    })

            if key in existing_tarea_link:
                tr_report["linked_to_tarea"] += 1
            else:
                tr_report["unlinked_to_tarea"] += 1
                if len(tr_report["sample_unlinked"]) < sample_limit:
                    tr_report["sample_unlinked"].append({
                        "id_recurso_tarea": access_id,
                        "key": key,
                        "id_recurso": s(row.get("IdRecurso")),
                        "tipo": s(row.get("TipoRecurso")),
                        "cantidad": s(row.get("Cantidad")),
                        "inicio": as_date_text(row.get("InicioRecursoReal")),
                    })

            inicio = as_date_text(row.get("InicioRecursoReal"))
            if inicio:
                if date_min is None or inicio < date_min:
                    date_min = inicio
                if date_max is None or inicio > date_max:
                    date_max = inicio

        tr_report["date_min"] = date_min or ""
        tr_report["date_max"] = date_max or ""
        tr_report["by_tipo_top"] = dict(by_tipo.most_common(20))
        report["tables"]["tblTareasRecursos"] = tr_report

        # tblRecursos
        recursos_report = {
            "rows": 0,
            "by_tipo": {},
            "sample": [],
        }
        by_tipo = Counter()

        for row in read_xlsx(folder / "tblRecursos.xlsx"):
            recursos_report["rows"] += 1
            by_tipo[s(row.get("Tipo"))] += 1
            if len(recursos_report["sample"]) < sample_limit:
                recursos_report["sample"].append({
                    "id_recurso": s(row.get("IdRecurso")),
                    "nombre": s(row.get("NombreRecurso")),
                    "tipo": s(row.get("Tipo")),
                    "unidad": s(row.get("Unidad")),
                    "cod_obra": s(row.get("CodObra")),
                })

        recursos_report["by_tipo"] = dict(by_tipo)
        report["tables"]["tblRecursos"] = recursos_report

        # tblRecursoAlmacen solo inspección por ahora
        almacen_report = {
            "rows": 0,
            "by_tipo_movimiento": {},
            "linked_to_tarea_context": 0,
            "unlinked_to_tarea_context": 0,
            "sample": [],
        }
        by_mov = Counter()

        for row in read_xlsx(folder / "tblRecursoAlmacen.xlsx"):
            almacen_report["rows"] += 1
            by_mov[s(row.get("TipoMovimiento"))] += 1

            key = (
                to_int(row.get("CodObra")),
                to_int(row.get("CodFase")),
                s(row.get("Vivienda")),
                s(row.get("Planta")),
                s(row.get("Capitulo")),
                s(row.get("Partida")),
            )

            if key in existing_tarea_link:
                almacen_report["linked_to_tarea_context"] += 1
            else:
                almacen_report["unlinked_to_tarea_context"] += 1

            if len(almacen_report["sample"]) < sample_limit:
                almacen_report["sample"].append({
                    "id_recurso": s(row.get("IdRecurso")),
                    "almacen": s(row.get("IdAlmacen")),
                    "cod_recurso": s(row.get("CodRecurso")),
                    "tipo_mov": s(row.get("TipoMovimiento")),
                    "cantidad": s(row.get("Cantidad")),
                    "fecha": as_date_text(row.get("FechaMovimiento")),
                    "obra": s(row.get("CodObra")),
                    "fase": s(row.get("CodFase")),
                    "vivienda": s(row.get("Vivienda")),
                    "partida": s(row.get("Partida")),
                })

        almacen_report["by_tipo_movimiento"] = dict(by_mov)
        report["tables"]["tblRecursoAlmacen"] = almacen_report

        self.stdout.write("")
        self.stdout.write("=== RESULTADOS ===")

        for table, info in report["tables"].items():
            self.stdout.write("")
            self.stdout.write(f"--- {table} ---")
            for k, v in info.items():
                if k.startswith("sample"):
                    continue
                self.stdout.write(f"{k}: {v}")

            for sample_key in ["sample_new", "sample_unlinked", "sample"]:
                if info.get(sample_key):
                    self.stdout.write(f"{sample_key}:")
                    for item in info[sample_key]:
                        self.stdout.write(json.dumps(item, ensure_ascii=False, default=str))

        if options["json_out"]:
            out = Path(options["json_out"])
            out.parent.mkdir(parents=True, exist_ok=True)
            out.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
            self.stdout.write("")
            self.stdout.write(f"JSON guardado en: {out}")

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("OK: dry-run XLSX finalizado. No se modificó la base de datos."))
